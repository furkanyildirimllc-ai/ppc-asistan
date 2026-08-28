"""Amazon Bulk Operations dosyasini okur, canli hesabi teshis eder,
duzeltme (Operation=Update) dosyasi uretir.

NEDEN BU MODUL VAR
Konsolun kampanya raporunda Campaign ID kolonu YOKTUR; oradan guncelleme
dosyasi uretilemez. Bulk Operations indirmesi hem gercek ID'leri hem canli
performansi tasir. Tek dogru kaynak budur.

TEMEL KURAL: BUTCE HARCAMA DEGILDIR
Bir kampanyanin harcamasi kazanilan acik artirma sayisiyla sinirlidir.
Bu hesapta olculdu: butcenin ancak %19-28'i harcaniyor. Bu yuzden iki
kisit ayirt edilir ve COZUMLERI ZITTIR:
  - butce kisiti  -> butce artir  (para tikiyor, talep var)
  - talep kisiti  -> teklif artir (para duruyor, acik artirma kazanilmiyor)
Ters yapmak parayi bosa koyar.
"""
import io
import math

import openpyxl

import benchmarks

SHEET = "Sponsored Products Campaigns"

# Bir kampanya gunde en az bu kadar tiklama alabilmeli. Altindaysa
# istatistik uretmeden once ayin sonu gelir - kampanya olu dogar.
MIN_CLICKS_PER_DAY = 5

# Bu kadar tiklamadan az veriyle "kotu" karari verilmez.
MIN_CLICKS_FOR_JUDGMENT = 15

# Sifir siparis karari icin gereken istatistiksel guven.
ZERO_ORDER_CONFIDENCE = 0.80

# Tek seferde teklif en fazla bu kadar dusurulur. Sert kesme kampanyayi
# aciklarin tamamen disina atar ve veri akisi durur.
MAX_BID_CUT = 0.50

MIN_BID = 0.20

# EKONOMIK TAVAN
# Bir tiklama ortalama (AOV x CVR) kadar ciro getirir. Bu deger, %100 ACOS'taki
# maksimum tekliftir - yani reklam harcamasinin satisa esit oldugu nokta.
# Bunun USTUNDE teklif vermek, yapisal olarak zarar satin almaktir; hicbir
# optimizasyon kurtaramaz.
#
# Olculdu: Natural'da tiklama basina ciro $1.47 iken ortanca teklif $3.60'ti
# (=%245 ACOS, yapisal). Gozlenen ACOS %353. Stemcell'de $2.20'ye karsi $2.55.
# Bu, hesaptaki kotu ACOS'un tek basina en buyuk sebebiydi.
DEFAULT_ACOS_CEILING = 1.00


def economic_ceiling(aov, cvr, acos_ceiling=DEFAULT_ACOS_CEILING):
    """benchmarks'a yonlendirir - kural TEK YERDE yasar."""
    return benchmarks.economic_ceiling(aov, cvr, acos_ceiling)


def _f(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def zero_order_confidence(clicks, cvr):
    """benchmarks'a yonlendirir - kural TEK YERDE yasar."""
    return benchmarks.zero_order_confidence(clicks, cvr)


def read_bulk(source):
    """Bulk dosyasini oku. source: dosya yolu ya da bytes."""
    if isinstance(source, (bytes, bytearray)):
        source = io.BytesIO(source)
    wb = openpyxl.load_workbook(source, data_only=True)
    if SHEET not in wb.sheetnames:
        raise ValueError(
            f"'{SHEET}' sayfasi bulunamadi. Bu bir Amazon Bulk Operations "
            f"dosyasi mi? Bulunan sayfalar: {', '.join(wb.sheetnames)}")
    ws = wb[SHEET]
    head = [c.value for c in ws[1]]
    idx = {v: i for i, v in enumerate(head) if v}
    for zorunlu in ("Entity", "Campaign ID", "Operation"):
        if zorunlu not in idx:
            raise ValueError(f"Beklenen kolon yok: {zorunlu}")
    rows = [[c.value for c in r] for r in ws.iter_rows(min_row=2)]
    return {"header": head, "idx": idx, "rows": rows}


def _get(bulk, row, key):
    i = bulk["idx"].get(key)
    return row[i] if i is not None and i < len(row) else None


def diagnose(bulk, target_acos_pct, expected_cvr, fallback_bid=2.00):
    """Her aktif kampanya icin ne yapilmasi gerektigini belirler.

    target_acos_pct: bu markada hedeflenen ACOS (break-even'a gore secilir)
    expected_cvr:    bu markanin OLCULMUS donusum orani (0.03 = %3)
    """
    g = lambda r, k: _get(bulk, r, k)
    ad_gruplari, kampanyalar = {}, {}
    ag_sayisi = {}
    for r in bulk["rows"]:
        e = g(r, "Entity")
        cid = g(r, "Campaign ID")
        if e == "Campaign":
            kampanyalar[cid] = r
        elif e == "Ad Group":
            ad_gruplari[cid] = r
            ag_sayisi[cid] = ag_sayisi.get(cid, 0) + 1

    islemler, uyarilar = [], []
    for cid, kr in kampanyalar.items():
        if g(kr, "State") != "enabled":
            continue
        ad = str(g(kr, "Campaign Name") or "")
        butce = _f(g(kr, "Daily Budget"))
        ar = ad_gruplari.get(cid)

        # Ad group'u olmayan kampanya hicbir zaman yayinlanmaz. Onceki
        # yuklemede Amazon kampanyayi kabul edip ad group'u reddetmis olur.
        if ar is None:
            islemler.append({
                "campaign_id": cid, "ad_group_id": None, "campaign": ad,
                "action": "pause", "budget": butce, "new_budget": butce,
                "bid": None, "new_bid": None,
                "reason": "ad group yok - kampanya hic yayinlanamaz",
                "severity": "kritik"})
            continue

        bid = _f(g(ar, "Ad Group Default Bid"))
        gost = _f(g(kr, "Impressions"))
        tik = _f(g(kr, "Clicks"))
        harcama = _f(g(kr, "Spend"))
        satis = _f(g(kr, "Sales"))
        acos = (harcama / satis * 100) if satis else None

        yeni_bid, yeni_butce, sebep, agirlik = bid, butce, None, "bilgi"

        if bid <= 0:
            yeni_bid = fallback_bid
            sebep = "teklif $0 - kampanya hic calisamaz"
            agirlik = "kritik"
        elif acos is not None and acos > target_acos_pct * 1.3:
            # Teklifi ACOS oraninda dus; ama tek seferde yaridan fazla kesme.
            oran = target_acos_pct / acos
            yeni_bid = max(round(bid * oran, 2),
                           round(bid * MAX_BID_CUT, 2), MIN_BID)
            sebep = f"ACOS %{acos:.0f} -> hedef %{target_acos_pct:.0f}"
            agirlik = "kritik" if acos > target_acos_pct * 2.5 else "uyari"
        elif tik >= MIN_CLICKS_FOR_JUDGMENT and not satis:
            guven = zero_order_confidence(tik, expected_cvr)
            if guven >= ZERO_ORDER_CONFIDENCE:
                yeni_bid = max(round(bid * 0.6, 2), MIN_BID)
                sebep = f"{tik:.0f} tik / 0 siparis (guven %{guven*100:.0f})"
                agirlik = "uyari"
            else:
                # VERI YETERSIZ. Karar vermek yerine bekle. Erken kesmek
                # sanssiz ama iyi bir kampanyayi olduruyor olabilir.
                sebep = (f"{tik:.0f} tik / 0 siparis - veri yetersiz "
                         f"(guven %{guven*100:.0f}), karar icin bekleniyor")
        elif gost == 0:
            sebep = "henuz gosterim yok - cok yeni, dokunulmuyor"

        # BUTCE TABANI HER ZAMAN UYGULANIR - gosterim sartina baglanmaz.
        #
        # HATA GECMISI: taban "gosterim > 0" sartina bagliydi. Mantik
        # doguruydu: $1 butceli kampanya zar zor yayinlanir, gosterim almaz,
        # gosterim almadigi icin "cok yeni" sayilip $1'de birakilirdi.
        # Kampanya sonsuza kadar olu kalirdi.
        #
        # Ayrim su: TEKLIF degisikligi bir performans yargisidir, veri ister.
        # BUTCE TABANI yapisal bir kusurdur - kampanyanin iyi mi kotu mu
        # oldugunu bilmeye gerek yok, $1 butce $2.79 teklifi zaten tasiyamaz.
        # ceil kullaniliyor: round() ile $2.67 x 5 = 13.35 -> 13 olup taban
        # yine saglanmiyordu (4.9 tik/gun).
        taban = math.ceil(yeni_bid * MIN_CLICKS_PER_DAY)
        if yeni_butce < taban:
            yeni_butce = taban
            onceki = f"{butce/yeni_bid:.1f}" if yeni_bid else "0"
            # "cok yeni - dokunulmuyor" notu artik gecersiz; butce degisiyor.
            if sebep and "cok yeni" in sebep:
                sebep = None
            sebep = ((sebep + " + ") if sebep else "") + \
                f"butce teklifi tasimiyordu ({onceki} tik/gun)"
            if agirlik == "bilgi":
                agirlik = "uyari"

        butce_degisti = abs(yeni_butce - butce) > 0.5
        bid_degisti = abs(yeni_bid - bid) > 0.01
        if not (butce_degisti or bid_degisti):
            if sebep:
                uyarilar.append({"campaign": ad, "note": sebep})
            continue

        islemler.append({
            "campaign_id": cid, "ad_group_id": g(ar, "Ad Group ID"),
            "campaign": ad, "action": "update",
            "budget": butce, "new_budget": yeni_butce if butce_degisti else butce,
            "bid": bid, "new_bid": yeni_bid if bid_degisti else bid,
            "clicks": tik, "spend": harcama, "sales": satis, "acos_pct": acos,
            "reason": sebep or "-", "severity": agirlik})

    return {"actions": islemler, "notes": uyarilar,
            "campaigns_live": sum(1 for r in kampanyalar.values()
                                  if _get(bulk, r, "State") == "enabled")}


def build_update(bulk, islemler):
    """Duzeltme dosyasini uretir. Kolon yapisi KAYNAK DOSYAYLA AYNIDIR -
    Amazon eslestirmeyi kolon adina gore yapar, boylece uyumsuzluk olmaz."""
    head, idx = bulk["header"], bulk["idx"]
    g = lambda r, k: _get(bulk, r, k)
    kam = {g(r, "Campaign ID"): r for r in bulk["rows"]
           if g(r, "Entity") == "Campaign"}
    agr = {g(r, "Ad Group ID"): r for r in bulk["rows"]
           if g(r, "Entity") == "Ad Group"}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET
    ws.append(head)

    # Update satiri kimlik alanlarini orijinalden tasimali; aksi halde
    # Amazon satiri eslestiremez.
    TASINAN = ("Product", "Entity", "Campaign ID", "Ad Group ID",
               "Campaign Name", "Ad Group Name", "Targeting Type",
               "State", "Start Date", "Bidding Strategy")

    def satir(kaynak, **alanlar):
        r = [None] * len(head)
        if kaynak is not None:
            for k in TASINAN:
                if k in idx:
                    r[idx[k]] = g(kaynak, k)
        r[idx["Operation"]] = "Update"
        for k, v in alanlar.items():
            if k in idx:
                r[idx[k]] = v
        return r

    sayac = {"butce": 0, "teklif": 0, "kapatma": 0}
    for i in islemler:
        if i["action"] == "pause":
            ws.append(satir(kam.get(i["campaign_id"]), **{"State": "paused"}))
            sayac["kapatma"] += 1
            continue
        if abs(_f(i["new_budget"]) - _f(i["budget"])) > 0.5:
            ws.append(satir(kam.get(i["campaign_id"]),
                            **{"Daily Budget": i["new_budget"]}))
            sayac["butce"] += 1
        if i.get("new_bid") and abs(_f(i["new_bid"]) - _f(i["bid"])) > 0.01:
            ws.append(satir(agr.get(i["ad_group_id"]),
                            **{"Ad Group Default Bid": i["new_bid"]}))
            sayac["teklif"] += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), sayac


def utilization(bulk, days=30):
    """Butce kullanimi: harcama / butce. Butce kisiti mi talep kisiti mi?"""
    g = lambda r, k: _get(bulk, r, k)
    kalem = []
    for r in bulk["rows"]:
        if g(r, "Entity") != "Campaign" or g(r, "State") != "enabled":
            continue
        b = _f(g(r, "Daily Budget"))
        if b <= 0:
            continue
        gunluk = _f(g(r, "Spend")) / max(days, 1)
        kalem.append({"campaign": str(g(r, "Campaign Name") or ""),
                      "budget": b, "daily_spend": round(gunluk, 2),
                      "utilization": round(gunluk / b, 3),
                      "limit": ("butce" if gunluk / b >= 0.8 else
                                "talep" if gunluk / b < 0.3 else "karisik")})
    tb = sum(k["budget"] for k in kalem)
    th = sum(k["daily_spend"] for k in kalem)
    return {"items": kalem, "total_budget": round(tb, 2),
            "total_daily_spend": round(th, 2),
            "utilization": round(th / tb, 3) if tb else 0,
            "budget_limited": sum(1 for k in kalem if k["limit"] == "butce"),
            "demand_limited": sum(1 for k in kalem if k["limit"] == "talep")}
