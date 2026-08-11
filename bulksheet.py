"""Amazon Sponsored Products Bulk Operations Excel formatinda export.

Cıkış: TEK sheet "Sponsored Products Campaigns" - Amazon'un canonical formatı.
Satırlar dogru hierarchical sirada:
  1. Campaign (Create) - yeni harvest kampanyasi
  2. Ad Group (Create) - yeni ad group
  3. Bidding Adjustment (Create) - placement multiplier (opsiyonel)
  4. Keyword (Update) - mevcut kelimede bid guncelleme
  5. Keyword (Create) - yeni exact kelime (harvest)
  6. Product Targeting (Create) - yeni ASIN hedefi
  7. Campaign Negative Keyword (Create) - negatif kelime
  8. Traffic sculpting negatifleri

Parent-ID hatalari icin:
- Kampanya adi ozel karakter icermiyor (parantez yok)
- Ad Group Name'ler bire bir tutarli
- Tum required field'lar dolduruldu
- Ayni upload icinde parent oluşturulup child satırında referans veriliyor
"""
import io
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from datetime import datetime

BULK_HEADERS = [
    "Product", "Entity", "Operation", "Campaign ID", "Ad Group ID",
    "Portfolio ID", "Ad ID (Read only)", "Keyword ID (Read only)",
    "Product Targeting ID (Read only)", "Campaign Name", "Ad Group Name",
    "Campaign Name (Informational only)", "Ad Group Name (Informational only)",
    "Portfolio Name (Informational only)", "Start Date", "End Date",
    "Targeting Type", "State", "Daily Budget", "SKU", "ASIN (Informational only)",
    "Eligibility Status (Informational only)",
    "Reason for Ineligibility (Informational only)",
    "Ad Group Default Bid", "Ad Group Default Bid (Informational only)",
    "Bid", "Keyword Text", "Native Language Keyword", "Native Language Locale",
    "Match Type", "Bidding Strategy",
    "Placement", "Percentage",
    "Product Targeting Expression",
    "Resolved Product Targeting Expression (Informational only)",
    "Impressions", "Clicks", "Click-through Rate", "Spend", "Sales",
    "Orders", "Units", "Conversion Rate", "ACOS", "CPC", "ROAS",
]


def _id_str(v):
    """ID'yi string olarak dondur (float/int/str -> str)."""
    if v is None or v == "":
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    s = str(v).strip()
    return s if s else ""


def _sanitize_name(s):
    """Kampanya/adgroup adindan Amazon'un anlayamayacagi karakterleri temizle.
    
    NOT: Parantezleri SILME — Amazon kampanya adlarinda parantez kullanimi
    normaldir ve silmek ID eslesmesini bozar.
    """
    if not s:
        return s
    s = str(s).strip()
    s = s.replace("\t", " ")
    while "  " in s:
        s = s.replace("  ", " ")
    return s.strip()


def _empty_row():
    return {h: "" for h in BULK_HEADERS}


def build(recs, brand_name, brand=None, report_rows=None):
    """recs -> BytesIO(xlsx). brand: dict. report_rows: raw normalized rows (icin ID map).

    Amazon 2024+ bulk validation Campaign ID + Ad Group ID istiyor.
    report_rows'dan campaign_name -> id ve (campaign_name, ad_group_name) -> id
    map'lerini cikarip her satira yazariz.
    """
    brand = brand or {}
    brand_name = _sanitize_name(brand_name) or "Marka"
    dest_campaign = _sanitize_name(brand.get("harvest_campaign") or "")
    dest_ad_group = _sanitize_name(brand.get("harvest_ad_group") or "")

    # ID map'leri kur
    camp_id_map = {}   # {campaign_name_lower: campaign_id}
    ag_id_map = {}      # {(camp_lower, ag_lower): ad_group_id}
    kw_id_map = {}      # {(camp_lower, ag_lower, keyword_lower, match_lower): keyword_id}
    placement_adj_map = {}  # {(campaign_id, placement_api): mevcut_yuzde}
    for row in (report_rows or []):
        cid = _id_str(row.get("campaign_id"))
        cname = (row.get("campaign") or "").strip()
        if cid and cname:
            camp_id_map.setdefault(cname.lower(), cid)
        agid = _id_str(row.get("ad_group_id"))
        agname = (row.get("ad_group") or "").strip()
        if agid and cname and agname:
            ag_id_map.setdefault((cname.lower(), agname.lower()), agid)
        kwid = _id_str(row.get("keyword_id"))
        kwtext = (row.get("keyword") or "").strip()
        mt = (row.get("match_type") or "").strip()
        if kwid and cname and agname and kwtext and mt:
            kw_id_map.setdefault(
                (cname.lower(), agname.lower(), kwtext.lower(), mt.lower()), kwid)
        if row.get("entity") == "Bidding Adjustment" and cid and row.get("placement"):
            placement_adj_map[(cid, row["placement"])] = row.get("percentage") or 0

    # Hedef kampanya/ad group ID'sini ISIMDEN coz (ayri bir ID alani tutmuyoruz -
    # tek dogruluk kaynagi bulk_ids dosyasindan gelen isim<->ID esleme haritasi).
    # Marka ayarlarindaki isim, bulk_ids dosyasinda bulunan gercek bir kampanya/ad
    # group ile birebir eslesirse otomatik cozulur; eslesmezse harvest atlanir.
    dest_campaign_id = camp_id_map.get(dest_campaign.lower(), "") if dest_campaign else ""
    dest_ad_group_id = ag_id_map.get((dest_campaign.lower(), dest_ad_group.lower()), "") \
        if (dest_campaign and dest_ad_group) else ""
    harvest_ok = bool(dest_campaign_id and dest_ad_group_id)

    today = datetime.now().strftime("%Y%m%d")
    has_harvest = any(r["type"] in ("harvest", "harvest_pt") for r in recs)
    # will_setup: yeni kampanya olusturulmaya calisilmayacak artik (Amazon izin vermiyor)
    will_setup = False
    skipped_harvest = has_harvest and not harvest_ok

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    head_font = Font(bold=True, color="FFFFFF", size=10)
    head_fill = PatternFill("solid", fgColor="1F2937")

    # =============== ANA SHEET (Amazon canonical) ===============
    ws = wb.create_sheet("Sponsored Products Campaigns")
    ws.append(BULK_HEADERS)
    for cell in ws[1]:
        cell.font, cell.fill = head_font, head_fill
    ws.freeze_panes = "A2"

    counts = {"campaign_create": 0, "adgroup_create": 0,
              "keyword_create": 0, "keyword_update": 0,
              "pt_create": 0, "neg_kw": 0, "neg_pt": 0,
              "placement": 0, "sculpt": 0}
    skipped = {"no_camp_id": [], "no_ag_id": [], "no_harvest_dest": [], "no_kw_id": []}

    def get_camp_id(name):
        return camp_id_map.get((name or "").strip().lower(), "")

    def get_ag_id(camp, ag):
        return ag_id_map.get(((camp or "").strip().lower(),
                              (ag or "").strip().lower()), "")

    def get_kw_id(camp, ag, kw, mt):
        return kw_id_map.get(((camp or "").strip().lower(), (ag or "").strip().lower(),
                              (kw or "").strip().lower(), (mt or "").strip().lower()), "")

    # ---- Placement (Bidding Adjustment) ----
    # ONCEKI HATA (duzeltildi): eski surum reason metnindeki ilk sayiyi (genelde
    # ACOS yuzdesi) regex ile "artis yuzdesi" saniyordu - "carpani sifirla/dusur"
    # onerileri ters yonde +%68-136 ARTIS satiri uretiyordu. Simdi yon (up/zero)
    # ve adim (step_pct) analysis.py'den YAPISAL geliyor, metinden tahmin yok.
    # Mevcut carpan degeri de tahmin edilmiyor - bulk_ids dosyasindaki gercek
    # Bidding Adjustment satirlarindan (placement_adj_map) okunuyor; hic
    # adjustment yoksa Operation=Create, varsa Operation=Update kullanilir.
    for rec in recs:
        if rec["type"] != "placement":
            continue
        m = rec.get("metrics") or {}
        api_name = m.get("placement_api")
        direction = m.get("direction")
        step_pct = m.get("step_pct")
        src_camp = _sanitize_name(rec.get("campaign", ""))
        camp_id = get_camp_id(rec.get("campaign", ""))
        if not (src_camp and api_name and direction):
            continue
        if not camp_id:
            skipped["no_camp_id"].append(f"placement in {src_camp}")
            continue
        current_pct = placement_adj_map.get((camp_id, api_name))
        operation = "Update" if current_pct is not None else "Create"
        if direction == "up":
            new_pct = min(900, (current_pct or 0) + (step_pct or 25))
        else:  # 'zero'
            new_pct = 0
            if current_pct == 0:
                continue  # zaten 0 - yazacak bir sey yok
        r = _empty_row()
        r.update({
            "Product": "Sponsored Products",
            "Entity": "Bidding Adjustment",
            "Operation": operation,
            "Campaign ID": camp_id,
            "Campaign Name": src_camp,
            "Placement": api_name,
            "Percentage": int(new_pct),
        })
        ws.append([r[h] for h in BULK_HEADERS])
        counts["placement"] += 1

    # ---- Keyword Update (mevcut kelimede bid degisikligi) ----
    for rec in recs:
        if rec["type"] not in ("bid_down", "bid_up"):
            continue
        camp = _sanitize_name(rec.get("campaign", ""))
        ag = _sanitize_name(rec.get("ad_group", ""))
        kw = rec.get("keyword", "")
        mt = (rec.get("match_type") or "EXACT").lower()
        bid = rec.get("suggested_value")
        camp_id = get_camp_id(rec.get("campaign", ""))
        ag_id = get_ag_id(rec.get("campaign", ""), rec.get("ad_group", ""))
        keyword_id = get_kw_id(rec.get("campaign", ""), rec.get("ad_group", ""), kw, mt)
        if not (camp and ag and kw and bid):
            continue
        if not camp_id:
            skipped["no_camp_id"].append(f"bid update in {camp}")
            continue
        if not ag_id:
            skipped["no_ag_id"].append(f"bid update in {camp}/{ag}")
            continue
        if not keyword_id:
            # Amazon bu satiri Keyword ID (Read only) olmadan REDDEDIYOR
            # (gercek hata raporunda dogrulandi: "Missing value for column:
            # Keyword ID (Read only)"). Isimle eslesme yeterli degil - tahmin
            # etmek yerine atla.
            skipped["no_kw_id"].append(f"bid update '{kw}' in {camp}/{ag}")
            continue
        r = _empty_row()
        r.update({
            "Product": "Sponsored Products",
            "Entity": "Keyword",
            "Operation": "Update",
            "Campaign ID": camp_id,
            "Ad Group ID": ag_id,
            "Keyword ID (Read only)": keyword_id,
            "Campaign Name": camp,
            "Ad Group Name": ag,
            "State": "enabled",
            "Keyword Text": kw,
            "Match Type": mt,
            "Bid": bid,
        })
        ws.append([r[h] for h in BULK_HEADERS])
        counts["keyword_update"] += 1

    # ---- Keyword Create (harvest yeni kelime) ----
    for rec in recs:
        if rec["type"] != "harvest":
            continue
        kw = rec.get("keyword", "")
        bid = rec.get("suggested_value")
        if not (kw and bid):
            continue
            
        if harvest_ok:
            r = _empty_row()
            r.update({
                "Product": "Sponsored Products",
                "Entity": "Keyword",
                "Operation": "Create",
                "Campaign ID": dest_campaign_id,
                "Ad Group ID": dest_ad_group_id,
                "Campaign Name": dest_campaign,
                "Ad Group Name": dest_ad_group,
                "State": "enabled",
                "Keyword Text": kw,
                "Match Type": "exact",
                "Bid": bid,
            })
            ws.append([r[h] for h in BULK_HEADERS])
            counts["keyword_create"] += 1
        else:
            # SEARCH TERM ISOLATION: SKAG Kampanyasi Olustur
            skag_camp = f"SKAG - {kw.title()}"
            skag_ag = "Exact"
            
            # 1. Campaign Create
            r_camp = _empty_row()
            r_camp.update({
                "Product": "Sponsored Products",
                "Entity": "Campaign",
                "Operation": "Create",
                "Campaign Name": skag_camp,
                "State": "enabled",
                "Targeting Type": "MANUAL",  # Amazon buyuk harf bekler
                "Start Date": today,
                "Daily Budget": 10, # default $10
            })
            ws.append([r_camp[h] for h in BULK_HEADERS])
            counts["campaign_create"] += 1
            
            # 2. Ad Group Create
            r_ag = _empty_row()
            r_ag.update({
                "Product": "Sponsored Products",
                "Entity": "Ad Group",
                "Operation": "Create",
                "Campaign Name": skag_camp,
                "Ad Group Name": skag_ag,
                "State": "enabled",
                "Ad Group Default Bid": bid,
            })
            ws.append([r_ag[h] for h in BULK_HEADERS])
            counts["adgroup_create"] += 1
            
            # 3. Keyword Create
            r_kw = _empty_row()
            r_kw.update({
                "Product": "Sponsored Products",
                "Entity": "Keyword",
                "Operation": "Create",
                "Campaign Name": skag_camp,
                "Ad Group Name": skag_ag,
                "State": "enabled",
                "Keyword Text": kw,
                "Match Type": "exact",
                "Bid": bid,
            })
            ws.append([r_kw[h] for h in BULK_HEADERS])
            counts["keyword_create"] += 1
            
            # 4. Source Campaign'lerde Negatifle (Isolation)
            sources = [s.strip() for s in rec.get("campaign", "").split(",") if s.strip()]
            for src in sources:
                src_id = get_camp_id(src)
                if not src_id:
                    continue # Bulunamazsa atla
                r_neg = _empty_row()
                r_neg.update({
                    "Product": "Sponsored Products",
                    "Entity": "Campaign Negative Keyword",
                    "Operation": "Create",
                    "Campaign ID": src_id,
                    "Campaign Name": src,
                    "State": "enabled",
                    "Keyword Text": kw,
                    "Match Type": "negative exact",
                })
                ws.append([r_neg[h] for h in BULK_HEADERS])
                counts["neg_kw"] += 1

    # ---- Product Targeting Create (yeni ASIN hedefleri) ----
    for rec in recs:
        if rec["type"] != "harvest_pt":
            continue
        kw = (rec.get("keyword") or "").upper()
        bid = rec.get("suggested_value")
        if not (kw and bid):
            continue
        if not harvest_ok:
            skipped["no_harvest_dest"].append(f"PT ASIN: {kw}")
            continue
        r = _empty_row()
        r.update({
            "Product": "Sponsored Products",
            "Entity": "Product Targeting",
            "Operation": "Create",
            "Campaign ID": dest_campaign_id,
            "Ad Group ID": dest_ad_group_id,
            "Campaign Name": dest_campaign,
            "Ad Group Name": dest_ad_group,
            "State": "enabled",
            "Product Targeting Expression": f'asin="{kw}"',
            "Bid": bid,
        })
        ws.append([r[h] for h in BULK_HEADERS])
        counts["pt_create"] += 1

    # ---- Campaign Negative Keyword (0 siparis negatifleri) ----
    for rec in recs:
        if rec["type"] != "negative":
            continue
        mt = (rec.get("match_type") or "").upper()
        if "PRODUCT" in mt:
            continue  # ASIN neg ad group seviyesinde
        camp = _sanitize_name(rec.get("campaign", ""))
        kw = rec.get("keyword", "")
        camp_id = get_camp_id(rec.get("campaign", ""))
        if not (camp and kw):
            continue
        if not camp_id:
            skipped["no_camp_id"].append(f"neg kw '{kw}' in {camp}")
            continue
        r = _empty_row()
        r.update({
            "Product": "Sponsored Products",
            "Entity": "Campaign Negative Keyword",
            "Operation": "Create",
            "Campaign ID": camp_id,
            "Campaign Name": camp,
            "State": "enabled",
            "Keyword Text": kw,
            "Match Type": "negativeExact",
        })
        ws.append([r[h] for h in BULK_HEADERS])
        counts["neg_kw"] += 1

    # ---- Traffic sculpting negatifleri ----
    # KRITIK: sadece harvest_ok ise calisir. Aksi halde kazanan kelime hicbir
    # yere eklenmeden (Keyword/PT Create atlanir) kaynak kampanyada negatiflenir
    # -> para kazandiran trafik sessizce kesilir. Bu daha once BIR HATA idi.
    for rec in recs:
        if not harvest_ok:
            break
        if rec["type"] not in ("harvest", "harvest_pt"):
            continue
        kw = rec.get("keyword", "")
        camp_str = rec.get("campaign", "")
        src_camps_raw = [c.strip() for c in camp_str.split(",") if c.strip()]
        for src_raw in src_camps_raw:
            src = _sanitize_name(src_raw)
            src_id = get_camp_id(src_raw)
            if not src or not src_id:
                if src:
                    skipped["no_camp_id"].append(f"sculpt '{kw}' in {src}")
                continue
            if rec["type"] == "harvest_pt":
                r = _empty_row()
                r.update({
                    "Product": "Sponsored Products",
                    "Entity": "Campaign Negative Product Targeting",
                    "Operation": "Create",
                    "Campaign ID": src_id,
                    "Campaign Name": src,
                    "State": "enabled",
                    "Product Targeting Expression": f'asin="{kw.upper()}"',
                })
            else:
                r = _empty_row()
                r.update({
                    "Product": "Sponsored Products",
                    "Entity": "Campaign Negative Keyword",
                    "Operation": "Create",
                    "Campaign ID": src_id,
                    "Campaign Name": src,
                    "State": "enabled",
                    "Keyword Text": kw,
                    "Match Type": "negativeExact",
                })
            ws.append([r[h] for h in BULK_HEADERS])
            counts["sculpt"] += 1

    # =============== OKU_ONCE ===============
    ws0 = wb.create_sheet("OKU_ONCE")
    ws0.column_dimensions["A"].width = 4
    ws0.column_dimensions["B"].width = 95

    def write(row_i, text, style="body"):
        c = ws0.cell(row=row_i, column=2, value=text)
        if style == "title":
            c.font = Font(bold=True, size=13, color="F59E0B")
        elif style == "section":
            c.font = Font(bold=True, size=11, color="F59E0B")
        elif style == "warn":
            c.font = Font(bold=True, color="DC2626")
        elif style == "bold":
            c.font = Font(bold=True, size=11)
        else:
            c.font = Font(size=10.5)
        c.alignment = Alignment(wrap_text=True, vertical="center")

    r = 2
    write(r, f"AMAZON BULK UPLOAD - {brand_name}", "title"); r += 2
    write(r, "Bu dosya Amazon Ads Console'un Bulk Operations'ina DIREKT yuklenmek uzere hazirlandi.", "bold"); r += 2

    write(r, "1) NASIL YUKLENIR", "section"); r += 1
    for line in [
        "  a) advertising.amazon.com sitesine gir, hesabina giris yap",
        "  b) Sol menude 'Bulk operations' bolumune tikla",
        "  c) 'Upload' butonuna bas -> bu Excel dosyasini sec -> Upload",
        "  d) Amazon 10-30 sn icinde validate eder. 'Success' gorursen tamam.",
        "  e) Hata cikarsa 'Show errors' -> Excel indir -> PPC Asistan ekibine yolla.",
    ]:
        write(r, line); r += 1
    r += 1

    if skipped_harvest:
        write(r, "!!! HARVEST KELIMELERI BU DOSYAYA EKLENMEDI !!!", "warn"); r += 1
        write(r, "Sebep: Amazon 2024+ bulk validation Campaign ID istiyor.", ""); r += 1
        write(r, "Yeni kampanya olusturmak bulk'ta artik guvenli degil.", ""); r += 1
        r += 1
        write(r, "COZUM (5 dakika, tek seferlik):", "bold"); r += 1
        write(r, "  1. Amazon Ads Console -> Campaigns -> Create campaign", ""); r += 1
        write(r, "  2. Sponsored Products -> Manual targeting", ""); r += 1
        write(r, f"  3. Kampanya adi: (istedigin isim, ornek '{brand_name} Exact Kazananlar')", ""); r += 1
        write(r, "  4. Ad group ekle, urun ekle, bir tane exact kelime ekle (dummy)", ""); r += 1
        write(r, "  5. Kaydet -> kampanya ve ad group Amazon'da olusur", ""); r += 1
        write(r, "  6. Bulk operations -> 'Download spreadsheet' ile ID esleme dosyasini indir", ""); r += 1
        write(r, "     (Search Term/Targeting raporlarinda ID OLMAZ - sadece bu indirmede var)", ""); r += 1
        write(r, "  7. PPC Asistan'a bu dosyayi normal rapor gibi surukle-birak yukle", ""); r += 1
        write(r, "     (otomatik 'Bulk Operations' olarak tanir, Campaign/Ad Group ID kaydeder)", ""); r += 1
        write(r, "  8. Marka ayarindan 'Harvest hedef kampanya' ismini gir", ""); r += 1
        write(r, "  9. Yeni bulksheet indir - artik harvest kelimeler dahil olur", ""); r += 1
        r += 1

    write(r, f"{'3' if will_setup else '2'}) DOSYADA NELER VAR", "section"); r += 1
    for name, cnt, desc in [
        ("Yeni kampanya oluştur", counts["campaign_create"], "SP kampanya, Manual, down-only bid"),
        ("Yeni ad group oluştur", counts["adgroup_create"], "Kampanyaya bağlı, $1.00 varsayılan bid"),
        ("Yeni exact kelime (harvest)", counts["keyword_create"], "Kanıtlanmış kazanan kelimeler exact match"),
        ("Yeni ASIN hedefi", counts["pt_create"], "Rakip ürün sayfalarında reklam"),
        ("Mevcut kelime bid güncelleme", counts["keyword_update"], "ACOS'a göre bid düşür/artır"),
        ("Placement multiplier", counts["placement"], "TOS/Product Pages/Rest of Search çarpanı"),
        ("Negatif kelime (kanama durdur)", counts["neg_kw"], "0 sipariş getirenler"),
        ("Traffic sculpting negatifi", counts["sculpt"], "Harvest sonrası kaynak kampanyada otomatik"),
    ]:
        if cnt > 0:
            write(r, f"  {name}: {cnt} satır - {desc}"); r += 1
    r += 1

    write(r, f"{'4' if will_setup else '3'}) HATA CIKARSA CO2ZUM YOLLARI", "section"); r += 1
    for line in [
        "  * 'Parent campaign not found' -> Kampanya adi Amazon'daki gerceklerle ayni degil. Cift kontrol et.",
        "  * 'Parent ad group not found' -> Ad group ismi yanlis. Bosluk/case duyarli olabilir.",
        "  * 'Bid below minimum' -> Bid < $0.15 var. Amazon minimum kabul etmez.",
        "  * 'Duplicate entity' -> Ayni kelime zaten var. O satiri atla, digerleri gecmis olabilir.",
        "  * 'Invalid targeting expression' -> ASIN formati yanlis, 'asin=\"B0XXX\"' olmali.",
        "  * Baska hata -> 'Download error report' ile Excel indir, PPC Asistan'a yolla.",
    ]:
        write(r, line); r += 1
    r += 1

    write(r, f"{'5' if will_setup else '4'}) GENEL KURALLAR", "section"); r += 1
    for line in [
        "  * Bid degisikligi sonrasi 7-14 gun bekle, erken mudahale etme.",
        "  * Amazon attribution 2-3 gun gec - son 2-3 gunun verisi eksik gelir.",
        "  * Bir seferde max %25 bid degisimi - fazlasi algoritmayi konfuze eder.",
        "  * Onaylayip yukledin - 1 hafta sonra yeni rapor cek, dosyayi PPC Asistan'a yukle.",
    ]:
        write(r, line); r += 1

    # =============== Column genislikleri ===============
    from openpyxl.utils import get_column_letter
    for i, h in enumerate(BULK_HEADERS, 1):
        ws.column_dimensions[get_column_letter(i)].width = \
            min(30, max(12, len(h) + 2))

    # Amazon Bulk butun sheet'leri validate ediyor ve OKU_ONCE'ta "Invalid Headers"
    # hatasi verip TUM dosyayi reddediyor. Bu yuzden OKU_ONCE'i cikariyoruz.
    # Talimatlar zaten UI Wizard'da mevcut.
    wb.remove(ws0)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
