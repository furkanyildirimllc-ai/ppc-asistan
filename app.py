"""PPC Asistan - Amazon reklam raporu analiz araci.
Calistir: .venv/bin/uvicorn app:app --port 8642
"""
import io
import json
import re
import sqlite3
import zipfile
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill
from fastapi import FastAPI, UploadFile, HTTPException
from fastapi.responses import FileResponse, StreamingResponse, HTMLResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, field_validator

import parsers
import analysis
import ai_agent
import supervisor
import insights
import bulksheet
import launch as launch_mod
import benchmarks
import bulk_doctor
import verify as verify_mod
import chat as chat_mod
import market_intel
import brain
import amazon_ads

DB_PATH = Path(__file__).parent / "ppc.db"
app = FastAPI(title="PPC Asistan")

# Chrome uzantisi (content script + popup) backend'e erisebilsin diye.
# Uzanti origin'i chrome-extension://... olur; gelistirmede tumune izin veriyoruz.
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


def db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys=ON")
    conn.execute("PRAGMA journal_mode=WAL")
    return conn


def init_db():
    with db() as c:
        c.executescript("""
        CREATE TABLE IF NOT EXISTS brands(
            id INTEGER PRIMARY KEY, name TEXT UNIQUE NOT NULL,
            target_acos REAL DEFAULT 0.30,
            min_clicks_neg INTEGER DEFAULT 8,
            min_orders_harvest INTEGER DEFAULT 2,
            bid_change_cap REAL DEFAULT 0.25,
            sell_price REAL DEFAULT 0,
            cogs REAL DEFAULT 0,
            amazon_fee_pct REAL DEFAULT 0.15,
            fba_fee REAL DEFAULT 0,
            harvest_campaign TEXT DEFAULT '',
            harvest_ad_group TEXT DEFAULT '');
        CREATE TABLE IF NOT EXISTS uploads(
            id INTEGER PRIMARY KEY,
            brand_id INTEGER NOT NULL REFERENCES brands(id) ON DELETE CASCADE,
            filename TEXT, report_type TEXT, row_count INTEGER,
            uploaded_at TEXT);
        CREATE TABLE IF NOT EXISTS report_rows(
            id INTEGER PRIMARY KEY,
            brand_id INTEGER NOT NULL REFERENCES brands(id) ON DELETE CASCADE,
            report_type TEXT NOT NULL, data TEXT NOT NULL);
        CREATE TABLE IF NOT EXISTS recommendations(
            id INTEGER PRIMARY KEY,
            brand_id INTEGER NOT NULL REFERENCES brands(id) ON DELETE CASCADE,
            type TEXT, campaign TEXT, ad_group TEXT, keyword TEXT,
            match_type TEXT, current_value REAL, suggested_value REAL,
            reason TEXT, metrics TEXT,
            confidence_score INTEGER DEFAULT 0,
            is_auto_applied INTEGER DEFAULT 0,
            status TEXT DEFAULT 'pending');
        CREATE TABLE IF NOT EXISTS ai_strategies(
            id INTEGER PRIMARY KEY,
            brand_id INTEGER NOT NULL REFERENCES brands(id) ON DELETE CASCADE,
            created_at TEXT,
            strategy_json TEXT,
            review_json TEXT,
            approved INTEGER DEFAULT 0,
            safe_to_send INTEGER DEFAULT 0);
        CREATE TABLE IF NOT EXISTS rec_history(
            id INTEGER PRIMARY KEY,
            brand_id INTEGER NOT NULL REFERENCES brands(id) ON DELETE CASCADE,
            rec_snapshot TEXT,
            new_status TEXT,
            old_status TEXT,
            acted_at TEXT);
        CREATE TABLE IF NOT EXISTS chat_messages(
            id INTEGER PRIMARY KEY,
            brand_id INTEGER NOT NULL REFERENCES brands(id) ON DELETE CASCADE,
            role TEXT, content TEXT, created_at TEXT);
        CREATE TABLE IF NOT EXISTS products(
            id INTEGER PRIMARY KEY,
            brand_id INTEGER NOT NULL REFERENCES brands(id) ON DELETE CASCADE,
            name TEXT, asin TEXT, sell_price REAL DEFAULT 0,
            cogs REAL DEFAULT 0, amazon_fee_pct REAL DEFAULT 0.15,
            fba_fee REAL DEFAULT 0, share_pct REAL DEFAULT 0);
        """)
        # ALTER TABLE - eski DB'lere yeni kolonlari ekle (yoksa)
        cols = {r["name"] for r in c.execute("PRAGMA table_info(brands)")}
        for col, ddl in [
            ("sell_price", "ALTER TABLE brands ADD COLUMN sell_price REAL DEFAULT 0"),
            ("cogs", "ALTER TABLE brands ADD COLUMN cogs REAL DEFAULT 0"),
            ("amazon_fee_pct", "ALTER TABLE brands ADD COLUMN amazon_fee_pct REAL DEFAULT 0.15"),
            ("fba_fee", "ALTER TABLE brands ADD COLUMN fba_fee REAL DEFAULT 0"),
            ("harvest_campaign", "ALTER TABLE brands ADD COLUMN harvest_campaign TEXT DEFAULT ''"),
            ("harvest_ad_group", "ALTER TABLE brands ADD COLUMN harvest_ad_group TEXT DEFAULT ''"),
            # Rakip marka tespiti otomatik tahmindir; kullanici duzeltmesi burada saklanir
            ("competitor_brands", "ALTER TABLE brands ADD COLUMN competitor_brands TEXT DEFAULT ''"),
            ("not_brands", "ALTER TABLE brands ADD COLUMN not_brands TEXT DEFAULT ''"),
            # Lansman markalari ayri bolumde gosterilir: henuz reklam verisi
            # yoktur, optimizasyon ekranlari onlar icin anlamsizdir.
            ("kind", "ALTER TABLE brands ADD COLUMN kind TEXT DEFAULT 'active'"),
        ]:
            if col not in cols:
                c.execute(ddl)
                
        # ALTER TABLE recommendations
        rec_cols = {r["name"] for r in c.execute("PRAGMA table_info(recommendations)")}
        for col, ddl in [
            ("confidence_score", "ALTER TABLE recommendations ADD COLUMN confidence_score INTEGER DEFAULT 0"),
            ("is_auto_applied", "ALTER TABLE recommendations ADD COLUMN is_auto_applied INTEGER DEFAULT 0"),
        ]:
            if col not in rec_cols:
                c.execute(ddl)


init_db()


class BrandIn(BaseModel):
    name: str
    target_acos: float = 0.30
    min_clicks_neg: int = 8
    min_orders_harvest: int = 2
    bid_change_cap: float = 0.25
    sell_price: float = 0
    cogs: float = 0
    amazon_fee_pct: float = 0.15
    fba_fee: float = 0
    harvest_campaign: str = ""
    harvest_ad_group: str = ""
    competitor_brands: str = ""
    not_brands: str = ""
    kind: str = "active"   # active | launch


def _profit_calc_single(sp, cogs, fee_pct, fba):
    if sp <= 0:
        return None
    ref_fee = sp * fee_pct
    unit_profit = sp - cogs - ref_fee - fba
    be = unit_profit / sp if sp > 0 else 0
    return {
        "unit_profit_before_ads": round(unit_profit, 2),
        "break_even_acos_pct": round(be * 100, 1),
        "recommended_target_acos_pct": round(be * 0.7 * 100, 1),
        "margin_pct_before_ads": round(unit_profit / sp * 100, 1),
    }


def _profit_calc(b, products=None):
    """Break-even ve karli hedef ACOS. Coklu urun varsa agirlikli ortalama."""
    if products:
        # Agirlikli ortalama - share_pct kullan (yoksa esit paylas)
        total_share = sum((p.get("share_pct") or 0) for p in products)
        if total_share <= 0:
            weights = [1 / len(products)] * len(products)
        else:
            weights = [(p.get("share_pct") or 0) / total_share for p in products]
        agg_profit = 0
        agg_price = 0
        margins = []
        for p, w in zip(products, weights):
            sp = p.get("sell_price") or 0
            if sp <= 0:
                continue
            cogs = p.get("cogs") or 0
            fee = sp * (p.get("amazon_fee_pct") or 0.15)
            fba = p.get("fba_fee") or 0
            up = sp - cogs - fee - fba
            agg_profit += up * w
            agg_price += sp * w
            margins.append({"name": p.get("name") or p.get("asin") or "?",
                            "share_pct": round(w * 100, 1),
                            "unit_profit": round(up, 2),
                            "break_even_acos_pct": round(up / sp * 100, 1) if sp > 0 else 0})
        if agg_price <= 0:
            return None
        be = agg_profit / agg_price
        return {
            "unit_profit_before_ads": round(agg_profit, 2),
            "break_even_acos_pct": round(be * 100, 1),
            "recommended_target_acos_pct": round(be * 0.7 * 100, 1),
            "margin_pct_before_ads": round(agg_profit / agg_price * 100, 1),
            "product_count": len(products),
            "per_product": margins,
            "mode": "weighted",
        }
    # Tek urun (eski)
    sp = b.get("sell_price") or 0
    if sp <= 0:
        return None
    res = _profit_calc_single(sp, b.get("cogs") or 0,
                               b.get("amazon_fee_pct") or 0.15,
                               b.get("fba_fee") or 0)
    if res:
        res["mode"] = "single"
    return res


_PAGE_CSS = """
body{font:15px/1.6 system-ui,-apple-system,sans-serif;background:#0f172a;
color:#e2e8f0;max-width:760px;margin:0 auto;padding:40px 24px}
h1{font-size:22px;margin:0 0 6px} h2{font-size:16px;margin:26px 0 6px;color:#38bdf8}
code{background:#1e293b;padding:2px 6px;border-radius:5px;font-size:13px;
word-break:break-all} .muted{color:#94a3b8;font-size:13px}
.box{background:#1e293b;border:1px solid #334155;border-radius:10px;padding:16px;margin:16px 0}
"""


@app.get("/privacy")
def privacy_notice():
    """Gizlilik bildirimi.

    Amazon Login with Amazon (LWA) guvenlik profili olustururken
    'Consent Privacy Notice URL' alani herkese acik bir HTTPS adres ister;
    localhost kabul edilmez. Bu sayfa o amaca hizmet eder.
    """
    html = f"""<!doctype html><html lang="tr"><head><meta charset="utf-8">
<title>Gizlilik Bildirimi - PPC Asistan</title><style>{_PAGE_CSS}</style></head><body>
<h1>Gizlilik Bildirimi</h1>
<p class="muted">PPC Asistan - Amazon reklam raporu analiz araci</p>

<h2>Bu uygulama nedir</h2>
<p>PPC Asistan, tek bir saticinin kendi Amazon reklam hesabini yonetmek icin
kullandigi <b>kisisel bir aractir</b>. Ucuncu taraflara hizmet sunmaz, kullanici
kaydi almaz, baska kullanicilarin verisini islemez.</p>

<h2>Hangi veriler islenir</h2>
<ul>
<li>Hesap sahibinin kendi Amazon reklam raporlari (kampanya, arama terimi,
hedefleme, yerlesim) ve Brand Analytics verileri</li>
<li>Amazon Advertising API kullanildiginda: yalnizca hesap sahibinin kendi
reklam hesabina ait kampanya ve teklif verileri</li>
</ul>
<p>Alici kisisel verisi, odeme bilgisi, adres veya iletisim bilgisi
<b>islenmez ve saklanmaz</b>.</p>

<h2>Veriler nerede tutulur</h2>
<p>Veriler yalnizca hesap sahibinin kendi ortaminda (yerel makine veya kendi
sunucusu) bir SQLite veritabaninda tutulur. Ucuncu taraflarla paylasilmaz,
satilmaz, reklam amaciyla kullanilmaz.</p>

<h2>Erisim yetkileri</h2>
<p>Amazon API erisim anahtarlari yalnizca sunucu ortam degiskenlerinde tutulur,
kaynak koda yazilmaz ve surum kontrolune dahil edilmez. Yetki istendigi anda
iptal edilebilir.</p>

<h2>Saklama ve silme</h2>
<p>Veriler uygulama icinden marka bazinda istenildigi zaman tamamen silinebilir.
Uygulama kaldirildiginda tum veri de silinmis olur.</p>

<h2>Iletisim</h2>
<p>Bu arac hakkinda soru icin hesap sahibiyle iletisime gecin.</p>
<p class="muted">Son guncelleme: {datetime.now():%Y-%m-%d}</p>
</body></html>"""
    return HTMLResponse(html)


@app.get("/callback")
def lwa_callback(code: str = "", error: str = "", error_description: str = ""):
    """LWA yetkilendirme donusu.

    Amazon buraya ?code=... ile doner. Kod kisa omurludur; sayfada gosterilir
    ki refresh token'a cevirmek icin kopyalanabilsin.
    """
    if error:
        body = (f'<div class="box"><b>Yetkilendirme reddedildi</b><br>'
                f'<code>{error}</code><p class="muted">{error_description}</p></div>')
    elif code:
        body = (f'<div class="box"><b>Authorization code alindi</b>'
                f'<p><code>{code}</code></p>'
                f'<p class="muted">Bu kod birkac dakika gecerlidir. '
                f'ADS_API_KURULUM.md icindeki curl komutuyla hemen '
                f'refresh token\'a cevir.</p></div>')
    else:
        body = ('<div class="box">Bu adres Amazon Login with Amazon '
                'yonlendirmesi icindir. Dogrudan acildiginda gosterilecek '
                'bir sey yoktur.</div>')
    html = f"""<!doctype html><html lang="tr"><head><meta charset="utf-8">
<title>LWA Callback - PPC Asistan</title><style>{_PAGE_CSS}</style></head><body>
<h1>Amazon yetkilendirme</h1>{body}</body></html>"""
    return HTMLResponse(html)


@app.get("/")
def index():
    return FileResponse(
        Path(__file__).parent / "static" / "index.html",
        headers={"Cache-Control": "no-cache, no-store, must-revalidate",
                 "Pragma": "no-cache", "Expires": "0"})


@app.get("/api/brands")
def list_brands():
    with db() as c:
        brands = [dict(r) for r in c.execute("SELECT * FROM brands ORDER BY name")]
        for b in brands:
            b["uploads"] = [dict(r) for r in c.execute(
                "SELECT filename, report_type, row_count, uploaded_at FROM uploads "
                "WHERE brand_id=? ORDER BY id DESC LIMIT 10", (b["id"],))]
            # Veri tazeligi: hangi rapor tipi ne zaman ve kac satir yuklendi.
            # "Eskiyi yeni sanma" sorununu gorunur kilar.
            b["data_freshness"] = [dict(r) for r in c.execute(
                "SELECT rr.report_type, COUNT(*) rows, "
                "  (SELECT MAX(uploaded_at) FROM uploads u "
                "   WHERE u.brand_id=rr.brand_id AND u.report_type=rr.report_type) last_upload "
                "FROM report_rows rr WHERE rr.brand_id=? "
                "GROUP BY rr.report_type ORDER BY rr.report_type", (b["id"],))]
            counts = c.execute(
                "SELECT status, COUNT(*) n FROM recommendations WHERE brand_id=? "
                "GROUP BY status", (b["id"],)).fetchall()
            b["rec_counts"] = {r["status"]: r["n"] for r in counts}
            prods = [dict(r) for r in c.execute(
                "SELECT id,name,asin,sell_price,cogs,amazon_fee_pct,fba_fee,share_pct "
                "FROM products WHERE brand_id=? ORDER BY id", (b["id"],))]
            b["products"] = prods
            b["profit"] = _profit_calc(b, prods if prods else None)
    return brands


@app.post("/api/brands")
def create_brand(body: BrandIn):
    with db() as c:
        try:
            cur = c.execute(
                "INSERT INTO brands(name,target_acos,min_clicks_neg,"
                "min_orders_harvest,bid_change_cap,sell_price,cogs,"
                "amazon_fee_pct,fba_fee,harvest_campaign,harvest_ad_group,"
                "competitor_brands,not_brands) "
                "VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (body.name.strip(), body.target_acos, body.min_clicks_neg,
                 body.min_orders_harvest, body.bid_change_cap,
                 body.sell_price, body.cogs, body.amazon_fee_pct, body.fba_fee,
                 body.harvest_campaign.strip(), body.harvest_ad_group.strip(),
                 body.competitor_brands.strip(), body.not_brands.strip()))
        except sqlite3.IntegrityError:
            raise HTTPException(400, "Bu isimde marka zaten var")
        return {"id": cur.lastrowid}


@app.put("/api/brands/{brand_id}")
def update_brand(brand_id: int, body: BrandIn):
    with db() as c:
        c.execute(
            "UPDATE brands SET name=?,target_acos=?,min_clicks_neg=?,"
            "min_orders_harvest=?,bid_change_cap=?,sell_price=?,cogs=?,"
            "amazon_fee_pct=?,fba_fee=?,harvest_campaign=?,harvest_ad_group=?,"
            "competitor_brands=?,not_brands=?,kind=? "
            "WHERE id=?",
            (body.name.strip(), body.target_acos, body.min_clicks_neg,
             body.min_orders_harvest, body.bid_change_cap,
             body.sell_price, body.cogs, body.amazon_fee_pct, body.fba_fee,
             body.harvest_campaign.strip(), body.harvest_ad_group.strip(),
             body.competitor_brands.strip(), body.not_brands.strip(),
             (body.kind or "active").strip(),
             brand_id))
    _regenerate(brand_id)
    return {"ok": True}


@app.delete("/api/brands/{brand_id}")
def delete_brand(brand_id: int):
    with db() as c:
        c.execute("DELETE FROM brands WHERE id=?", (brand_id,))
    return {"ok": True}


@app.get("/api/ads-api/status")
def ads_api_status():
    """Amazon Ads API baglantisi kurulu mu, calisiyor mu?"""
    return amazon_ads.check()


@app.get("/api/brands/{brand_id}/discovery-status")
def discovery_status(brand_id: int):
    """FAZ 0 sonucu: ne oldu, simdi ne yapmali?"""
    with db() as c:
        if not c.execute("SELECT 1 FROM brands WHERE id=?", (brand_id,)).fetchone():
            raise HTTPException(404, "Marka bulunamadi")
        rows = _load_rows(c, brand_id, "targeting") or _load_rows(c, brand_id, "search_term")
    import benchmarks
    sonuc = benchmarks.diagnose_discovery(rows)
    # Ayni markada birden fazla urun varsa her birini AYRI goster.
    sonuc["products"] = benchmarks.products_in(rows)
    return sonuc


@app.get("/api/product-status")
def product_status(asin: str = "", brand_id: int | None = None):
    """Bir URUN icin: olculmus CPC var mi, hangi faz gerekli?

    Uzanti ASIN girilir girilmez bunu sorar. Amac: kullanicinin "ben hangi
    fazdayim, verim var mi" diye tahmin etmesini bitirmek.
    """
    import benchmarks
    a = (asin or "").strip().upper()
    if not a:
        return {"asin": "", "phase": "bilinmiyor", "headline": "ASIN girilmedi"}

    rows = []
    marka_adi = None
    with db() as c:
        if brand_id:
            r = c.execute("SELECT name FROM brands WHERE id=?", (brand_id,)).fetchone()
            marka_adi = r["name"] if r else None
            rows = _load_rows(c, brand_id, "targeting")
        else:
            # Marka secilmediyse ASIN'i TUM markalarda ara - ama hangi markada
            # bulundugunu SOYLE (sessizce baska markanin verisini kullanma).
            for b in c.execute("SELECT id,name FROM brands"):
                rs = _load_rows(c, b["id"], "targeting")
                if any(benchmarks.asin_of_row(x) == a for x in rs):
                    rows, marka_adi, brand_id = rs, b["name"], b["id"]
                    break

    urun = next((u for u in benchmarks.products_in(rows) if u["asin"] == a), None)
    if not urun:
        return {"asin": a, "brand_id": brand_id, "brand_name": marka_adi,
                "phase": "faz0", "have_data": False,
                "headline": "Bu ürün için ölçülmüş veri yok",
                "detail": "Önce Faz 0 (CPC keşfi) çalıştır — 3 gün, ~$135.",
                "action": "Faz 0 dosyasını indir ve yükle."}

    yeter = urun["enough_for_cpc"]
    return {
        "asin": a, "brand_id": brand_id, "brand_name": marka_adi,
        "phase": "faz1" if yeter else "faz0_devam",
        "have_data": True, "clicks": urun["clicks"], "cpc": urun["cpc"],
        "cvr_pct": urun["cvr_pct"], "orders": urun["orders"],
        "headline": (f"Ölçüldü: CPC ${urun['cpc']} ({urun['clicks']:.0f} tıklama)"
                     if yeter else
                     f"Veri zayıf: {urun['clicks']:.0f} tıklama (en az 15 gerekir)"),
        "detail": (f"Bu ürünün kendi verisi kullanılacak. CVR %{urun['cvr_pct']}."
                   if yeter else "Keşfi biraz daha sürdür ya da temkinli devam et."),
        "action": ("Faz 1 planını üretebilirsin." if yeter
                   else "Faz 0'ı sürdür; bid hesabı zayıf veriyle yapılır."),
    }


@app.get("/api/brands/{brand_id}/data-inventory")
def data_inventory(brand_id: int):
    """Bu markada NE VAR: rapor tipi, satir sayisi, son yukleme tarihi.
    Neyi silecegini gormeden silme."""
    with db() as c:
        if not c.execute("SELECT 1 FROM brands WHERE id=?", (brand_id,)).fetchone():
            raise HTTPException(404, "Marka bulunamadi")
        rows = [dict(r) for r in c.execute(
            "SELECT rr.report_type, COUNT(*) rows, "
            "  (SELECT MAX(uploaded_at) FROM uploads u WHERE u.brand_id=rr.brand_id "
            "   AND u.report_type=rr.report_type) last_upload "
            "FROM report_rows rr WHERE rr.brand_id=? GROUP BY rr.report_type",
            (brand_id,))]
        recs = c.execute("SELECT COUNT(*) n FROM recommendations WHERE brand_id=?",
                         (brand_id,)).fetchone()["n"]
        ai = c.execute("SELECT COUNT(*) n FROM ai_strategies WHERE brand_id=?",
                       (brand_id,)).fetchone()["n"]
    return {"brand_id": brand_id, "reports": rows,
            "recommendations": recs, "ai_strategies": ai}


class ResetIn(BaseModel):
    # Bos birakilirsa TUM rapor tipleri silinir.
    report_types: list[str] = []
    clear_recommendations: bool = True
    clear_ai: bool = False


@app.post("/api/brands/{brand_id}/reset")
def reset_brand_data(brand_id: int, body: ResetIn):
    """Markanin verisini sifirlar. Marka kaydi ve ayarlari KORUNUR.

    Haftalik yeni rapor yuklemeden once eskiyi temizlemek icin. Silme
    islemi yalnizca bu brand_id'yi etkiler - baska marka etkilenmez.
    """
    with db() as c:
        if not c.execute("SELECT 1 FROM brands WHERE id=?", (brand_id,)).fetchone():
            raise HTTPException(404, "Marka bulunamadi")
        deleted = {}
        if body.report_types:
            for rt in body.report_types:
                n = c.execute("DELETE FROM report_rows WHERE brand_id=? AND report_type=?",
                              (brand_id, rt)).rowcount
                c.execute("DELETE FROM uploads WHERE brand_id=? AND report_type=?",
                          (brand_id, rt))
                deleted[rt] = n
        else:
            n = c.execute("DELETE FROM report_rows WHERE brand_id=?", (brand_id,)).rowcount
            c.execute("DELETE FROM uploads WHERE brand_id=?", (brand_id,))
            deleted["hepsi"] = n
        if body.clear_recommendations:
            deleted["recommendations"] = c.execute(
                "DELETE FROM recommendations WHERE brand_id=?", (brand_id,)).rowcount
            c.execute("DELETE FROM rec_history WHERE brand_id=?", (brand_id,))
        if body.clear_ai:
            deleted["ai_strategies"] = c.execute(
                "DELETE FROM ai_strategies WHERE brand_id=?", (brand_id,)).rowcount
    return {"ok": True, "deleted": deleted}


def _load_rows(c, brand_id, rtype):
    return [json.loads(r["data"]) for r in c.execute(
        "SELECT data FROM report_rows WHERE brand_id=? AND report_type=?",
        (brand_id, rtype))]


def _regenerate(brand_id):
    """Marka verisinden onerileri yeniden uret (onaylanmis/reddedilmisler korunur)."""
    with db() as c:
        brand = c.execute("SELECT * FROM brands WHERE id=?", (brand_id,)).fetchone()
        if not brand:
            return
        sts = _load_rows(c, brand_id, "search_term")
        tgs = _load_rows(c, brand_id, "targeting")
        plc = _load_rows(c, brand_id, "placement")
        camps = _load_rows(c, brand_id, "campaign")
        recs = analysis.run_all(dict(brand), sts, tgs, plc, camps)
        # islenmis (approved/rejected) onerileri tekrar gosterme
        done = {(r["type"], r["campaign"], r["keyword"], r["match_type"])
                for r in c.execute(
                    "SELECT type,campaign,keyword,match_type FROM recommendations "
                    "WHERE brand_id=? AND status!='pending'", (brand_id,))}
        c.execute("DELETE FROM recommendations WHERE brand_id=? AND status='pending'",
                  (brand_id,))
        for r in recs:
            if (r["type"], r["campaign"], r["keyword"], r["match_type"]) in done:
                continue
            
            auto_apply = r.get("auto_apply", False)
            status = 'approved' if auto_apply else 'pending'
            
            c.execute(
                "INSERT INTO recommendations(brand_id,type,campaign,ad_group,keyword,"
                "match_type,current_value,suggested_value,reason,metrics,"
                "confidence_score,is_auto_applied,status) "
                "VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)",
                (brand_id, r["type"], r["campaign"], r["ad_group"], r["keyword"],
                 r["match_type"], r["current_value"], r["suggested_value"],
                 r["reason"], json.dumps(r["metrics"]), 
                 r.get("confidence", 0), 1 if auto_apply else 0, status))


@app.post("/api/brands/{brand_id}/upload")
async def upload(brand_id: int, files: list[UploadFile]):
    results = []
    with db() as c:
        if not c.execute("SELECT 1 FROM brands WHERE id=?", (brand_id,)).fetchone():
            raise HTTPException(404, "Marka bulunamadi")
        for f in files:
            content = await f.read()
            # Guvenlik: 100MB'tan buyuk dosyalari reddet (OOM koruması)
            if len(content) > 100 * 1024 * 1024:
                results.append({"file": f.filename, "ok": False,
                                "error": "Dosya çok büyük (max 100MB)"})
                continue
            try:
                rtype, rows = parsers.parse(f.filename, content)
            except Exception as e:
                results.append({"file": f.filename, "ok": False, "error": str(e)})
                continue
            if rows:
                # ayni tip eski veriyi degistir
                c.execute("DELETE FROM report_rows WHERE brand_id=? AND report_type=?",
                          (brand_id, rtype))
                c.executemany(
                    "INSERT INTO report_rows(brand_id,report_type,data) VALUES(?,?,?)",
                    [(brand_id, rtype, json.dumps(r)) for r in rows])
            c.execute(
                "INSERT INTO uploads(brand_id,filename,report_type,row_count,"
                "uploaded_at) VALUES(?,?,?,?,?)",
                (brand_id, f.filename, rtype, len(rows),
                 datetime.now().isoformat(timespec="seconds")))
            results.append({"file": f.filename, "ok": True,
                            "type": parsers.REPORT_LABELS.get(rtype, rtype),
                            "rows": len(rows)})
    _regenerate(brand_id)
    return results


@app.get("/api/brands/{brand_id}/recommendations")
def get_recs(brand_id: int, status: str = "pending"):
    with db() as c:
        rows = c.execute(
            "SELECT * FROM recommendations WHERE brand_id=? AND status=? "
            "ORDER BY type, id", (brand_id, status)).fetchall()
    out = []
    for r in rows:
        d = dict(r)
        d["metrics"] = json.loads(d["metrics"] or "{}")
        out.append(d)
    return out


class RecAction(BaseModel):
    ids: list[int]
    status: str  # approved | rejected | pending


@app.post("/api/brands/{brand_id}/recommendations/action")
def rec_action(brand_id: int, body: RecAction):
    if body.status not in ("approved", "rejected", "pending"):
        raise HTTPException(400, "Gecersiz durum")
    with db() as c:
        # history icin snapshot
        rows = c.execute(
            "SELECT id,type,campaign,ad_group,keyword,match_type,current_value,"
            "suggested_value,reason,metrics,status FROM recommendations "
            f"WHERE brand_id=? AND id IN ({','.join('?'*len(body.ids))})",
            (brand_id, *body.ids)).fetchall()
        now = datetime.now().isoformat(timespec="seconds")
        for r in rows:
            c.execute(
                "INSERT INTO rec_history(brand_id,rec_snapshot,new_status,"
                "old_status,acted_at) VALUES(?,?,?,?,?)",
                (brand_id, json.dumps(dict(r)), body.status, r["status"], now))
        c.executemany(
            "UPDATE recommendations SET status=? WHERE id=? AND brand_id=?",
            [(body.status, i, brand_id) for i in body.ids])
    return {"ok": True}


@app.get("/api/brands/{brand_id}/history")
def get_history(brand_id: int, limit: int = 50):
    with db() as c:
        rows = c.execute(
            "SELECT * FROM rec_history WHERE brand_id=? ORDER BY id DESC LIMIT ?",
            (brand_id, limit)).fetchall()
    out = []
    for r in rows:
        d = dict(r)
        d["rec_snapshot"] = json.loads(d["rec_snapshot"])
        out.append(d)
    return out


class UndoBody(BaseModel):
    history_id: int


@app.post("/api/brands/{brand_id}/history/undo")
def undo_action(brand_id: int, body: UndoBody):
    with db() as c:
        h = c.execute(
            "SELECT * FROM rec_history WHERE id=? AND brand_id=?",
            (body.history_id, brand_id)).fetchone()
        if not h:
            raise HTTPException(404, "Geri alma kaydi bulunamadi")
        snap = json.loads(h["rec_snapshot"])
        c.execute(
            "UPDATE recommendations SET status=? WHERE id=? AND brand_id=?",
            (h["old_status"], snap["id"], brand_id))
        c.execute("DELETE FROM rec_history WHERE id=?", (h["id"],))
    return {"ok": True, "restored_to": h["old_status"]}


class BidEdit(BaseModel):
    suggested_value: float


@app.put("/api/recommendations/{rec_id}/bid")
def edit_bid(rec_id: int, body: BidEdit):
    with db() as c:
        c.execute("UPDATE recommendations SET suggested_value=? WHERE id=?",
                  (round(body.suggested_value, 2), rec_id))
    return {"ok": True}


@app.get("/api/brands/{brand_id}/summary")
def summary(brand_id: int):
    with db() as c:
        camps = _load_rows(c, brand_id, "campaign")
        sts = _load_rows(c, brand_id, "search_term")
    src = camps if camps else sts
    spend = sum(r["spend"] for r in src)
    sales = sum(r["sales"] for r in src)
    orders = sum(r["orders"] for r in src)
    clicks = sum(r["clicks"] for r in src)
    top = sorted(camps, key=lambda r: -r["spend"])[:15] if camps else []
    return {
        "spend": round(spend, 2), "sales": round(sales, 2),
        "orders": int(orders), "clicks": int(clicks),
        "acos": round(spend / sales * 100, 1) if sales else None,
        "source": "campaign" if camps else ("search_term" if sts else None),
        "top_campaigns": top,
    }


TYPE_SHEETS = [
    ("harvest", "Yeni Kelimeler"),
    ("harvest_pt", "Yeni Urun Hedefleri"),
    ("negative", "Negatifler"),
    ("bid_down", "Bid Dusur"),
    ("bid_up", "Bid Artir"),
    ("placement", "Placement Ayarlari"),
]


@app.get("/api/brands/{brand_id}/export")
def export(brand_id: int, status: str = "approved"):
    with db() as c:
        brand = c.execute("SELECT name FROM brands WHERE id=?", (brand_id,)).fetchone()
        if not brand:
            raise HTTPException(404, "Marka bulunamadi")
        rows = c.execute(
            "SELECT * FROM recommendations WHERE brand_id=? AND status=? ORDER BY type",
            (brand_id, status)).fetchall()
    if not rows:
        raise HTTPException(400, "Aktarilacak onayli oneri yok")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="1F2937")
    # OKU_ONCE sheet - net rehber
    ws0 = wb.create_sheet("OKU_ONCE")
    from openpyxl.styles import Alignment
    lines = [
        (f"MARKA: {brand['name']}", True),
        (f"Uretilme: {datetime.now():%Y-%m-%d %H:%M}", False),
        ("", False),
        ("BU DOSYA NEDIR?", True),
        ("PPC asistan tarafindan onerilen ve senin onayladigin degisikliklerin listesi.", False),
        ("Her sheet bir kategori. Sirayla uygulanmasi tavsiye edilir:", False),
        ("", False),
        ("SIRA 1 - NEGATIFLER (kanamayi durdur)", True),
        ("Amazon Ads Console -> ilgili kampanya -> Negative targeting -> ekle.", False),
        ("Kelimeler icin 'Negative Exact', ASIN'ler icin 'Negative Product Targeting'.", False),
        ("DIKKAT: Negative phrase EKLEME - istemedigin varyasyonlari da keser.", False),
        ("", False),
        ("SIRA 2 - YENI KELIMELER (Kelime Avcisi / Harvest)", True),
        ("Onerilen kelimeleri Amazon'da EXACT kampanyaya ekle.", False),
        ("YOKSA yeni bir 'MarkaAdi - Exact - Kazananlar' kampanyasi ac.", False),
        ("ONEMLI: Ayni kelimeyi kaynak kampanyada NEGATIVE EXACT olarak ekle!", False),
        ("Bu 'traffic sculpting' - yoksa iki kampanyan birbirine acik artirmaya girer.", False),
        ("", False),
        ("SIRA 3 - YENI URUN HEDEFLERI (ASIN)", True),
        ("Product Targeting kampanyana ASIN'leri ekle.", False),
        ("Kaynak Auto kampanyada bu ASIN'leri negative product yap.", False),
        ("", False),
        ("SIRA 4 - BID DEGISIKLIKLERI", True),
        ("Amazon Ads -> kampanya -> Targeting sekmesi -> kelimeyi bul -> bid guncelle.", False),
        ("'Mevcut CPC' senin gercek bid'in degil - odedigin ortalama tik ucreti.", False),
        ("Bid genelde CPC'nin biraz uzerinde. Onerilen bid = hedef ACOS icin ideal.", False),
        ("", False),
        ("SIRA 5 - PLACEMENT AYARLARI", True),
        ("Amazon Ads -> kampanya -> Settings -> Adjust bids by placement.", False),
        ("Top of search / Product pages / Rest of search icin carpan guncelle.", False),
        ("", False),
        ("GENEL KURALLAR", True),
        ("- Bid degisikligi sonrasi 7-14 gun bekle. Attribution 2-3 gun eksik.", False),
        ("- Tek seferde max %25 bid degisikligi - fazlasi algoritmayi konfuze eder.", False),
        ("- Onayladigin her degisikligi Amazon'da uyguladiginda tik at (kendin icin).", False),
        ("", False),
        ("BULKSHEET INDIRDIYSEM NE OLACAK?", True),
        ("Bulksheet .xlsx dosyasi Amazon 'Bulk Operations' sayfasina direkt yuklenir.", False),
        ("Menudeki '📦 Bulksheet' butonu ile indir - saatlik is 5 dakikaya iner.", False),
    ]
    for i, (line, bold) in enumerate(lines, 1):
        cell = ws0.cell(row=i, column=1, value=line)
        if bold:
            cell.font = Font(bold=True, size=12, color="F59E0B")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws0.column_dimensions["A"].width = 95

    by_type = {}
    for r in rows:
        by_type.setdefault(r["type"], []).append(r)
    for rtype, sheet_name in TYPE_SHEETS:
        items = by_type.get(rtype)
        if not items:
            continue
        ws = wb.create_sheet(sheet_name)
        headers = ["Kampanya", "Ad Group", "Kelime/Hedef", "Match Type",
                   "Mevcut CPC", "Onerilen Bid", "Tiklama", "Harcama",
                   "Satis", "Siparis", "ACOS %", "Aciklama"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font, cell.fill = head_font, head_fill
        for it in items:
            m = json.loads(it["metrics"] or "{}")
            ws.append([it["campaign"], it["ad_group"], it["keyword"],
                       it["match_type"], it["current_value"],
                       it["suggested_value"], m.get("clicks"), m.get("spend"),
                       m.get("sales"), m.get("orders"), m.get("acos"),
                       it["reason"]])
        widths = [34, 24, 34, 16, 11, 13, 9, 10, 10, 9, 9, 60]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"{brand['name']}_ppc_aksiyon_{datetime.now():%Y%m%d}.xlsx"
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument"
        ".spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'})


# ---------- AI Strateji + Denetci ----------

@app.post("/api/brands/{brand_id}/ai-strategy")
def ai_strategy(brand_id: int):
    """AI stratejisi uret + denetciden gecir. Sonucu DB'ye kaydet."""
    with db() as c:
        brand_row = c.execute("SELECT * FROM brands WHERE id=?", (brand_id,)).fetchone()
        if not brand_row:
            raise HTTPException(404, "Marka bulunamadi")
        brand = dict(brand_row)
        sts = _load_rows(c, brand_id, "search_term")
        tgs = _load_rows(c, brand_id, "targeting")
        plc = _load_rows(c, brand_id, "placement")
        camps = _load_rows(c, brand_id, "campaign")
        rec_rows = c.execute(
            "SELECT type,campaign,ad_group,keyword,match_type,current_value,"
            "suggested_value,reason,metrics FROM recommendations "
            "WHERE brand_id=? AND status='pending'", (brand_id,)).fetchall()
    if not (sts or tgs):
        raise HTTPException(400, "Once rapor yukleyin")
    det_recs = []
    for r in rec_rows:
        d = dict(r)
        d["metrics"] = json.loads(d["metrics"] or "{}")
        det_recs.append(d)

    try:
        payload = ai_agent.build_input(brand, sts, tgs, plc, camps, det_recs)
        strategy = ai_agent.generate_strategy(payload)
    except Exception as e:
        raise HTTPException(500, f"AI strateji hatasi: {e}")

    try:
        review = supervisor.review(strategy, det_recs, brand)
    except Exception as e:
        review = {
            "approved": False, "risk_level": "high",
            "summary": f"Denetci hata verdi: {e}",
            "issues": [{"severity": "critical", "location": "supervisor",
                        "problem": str(e), "fix": "Manuel incele"}],
            "safe_to_send_to_user": False,
        }

    with db() as c:
        cur = c.execute(
            "INSERT INTO ai_strategies(brand_id,created_at,strategy_json,"
            "review_json,approved,safe_to_send) VALUES(?,?,?,?,?,?)",
            (brand_id, datetime.now().isoformat(timespec="seconds"),
             json.dumps(strategy, ensure_ascii=False),
             json.dumps(review, ensure_ascii=False),
             1 if review.get("approved") else 0,
             1 if review.get("safe_to_send_to_user") else 0))
        sid = cur.lastrowid
    return {"id": sid, "strategy": strategy, "review": review}


@app.get("/api/brands/{brand_id}/ai-strategy/latest")
def ai_strategy_latest(brand_id: int):
    with db() as c:
        row = c.execute(
            "SELECT * FROM ai_strategies WHERE brand_id=? "
            "ORDER BY id DESC LIMIT 1", (brand_id,)).fetchone()
    if not row:
        return {"exists": False}
    return {
        "exists": True,
        "id": row["id"],
        "created_at": row["created_at"],
        "strategy": json.loads(row["strategy_json"]),
        "review": json.loads(row["review_json"]),
        "approved": bool(row["approved"]),
        "safe_to_send": bool(row["safe_to_send"]),
    }


@app.get("/api/brands/{brand_id}/ai-strategy/history")
def ai_strategy_history(brand_id: int):
    with db() as c:
        rows = c.execute(
            "SELECT id,created_at,approved,safe_to_send FROM ai_strategies "
            "WHERE brand_id=? ORDER BY id DESC LIMIT 20", (brand_id,)).fetchall()
    return [dict(r) for r in rows]


# ---------- Insights (dashboard) ----------

@app.get("/api/brands/{brand_id}/bulk-readiness")
def bulk_readiness(brand_id: int):
    """Bulk indirmeye hazir miyiz kontrol et."""
    with db() as c:
        brand_row = c.execute("SELECT * FROM brands WHERE id=?", (brand_id,)).fetchone()
        if not brand_row:
            raise HTTPException(404, "Marka bulunamadi")
        all_rows = []
        for rtype in ("search_term", "targeting", "campaign", "placement", "bulk_ids"):
            all_rows.extend(_load_rows(c, brand_id, rtype))
        approved_cnt = c.execute(
            "SELECT COUNT(*) c FROM recommendations WHERE brand_id=? AND status='approved'",
            (brand_id,)).fetchone()["c"]
    has_ids = any(r.get("campaign_id") for r in all_rows)
    id_coverage = 0
    if all_rows:
        with_id = sum(1 for r in all_rows if r.get("campaign_id"))
        id_coverage = round(with_id / len(all_rows) * 100, 1)
    return {
        "approved_count": approved_cnt,
        "has_campaign_ids": has_ids,
        "id_coverage_pct": id_coverage,
        "harvest_campaign_set": bool(brand_row["harvest_campaign"]),
        "ready": has_ids and approved_cnt > 0,
        "message": (
            "Hazir" if has_ids and approved_cnt > 0 else
            "Onay yok - once oneri onayla" if not approved_cnt else
            "Campaign ID yok - Amazon Ads Console > Bulk Operations > Download "
            "spreadsheet ile ID dosyasini indirip PPC Asistan'a yukle "
            "(normal performans raporlarinda ID bulunmaz, kac kere indirsen fark etmez)"
        ),
    }


@app.get("/api/brands/{brand_id}/campaign-ad-groups")
def campaign_ad_groups(brand_id: int):
    """Bulk Operations ID dosyasindan gercek kampanya/ad-group isim listesi.

    Harvest hedefi secerken kullanici elle yazip yazim hatasi yapmasin diye -
    burada donen her (kampanya, ad group) cifti garanti ID'si cozulebilir
    olandir (ayni report_rows kaynagindan geliyor, bulksheet.py'nin ID
    haritasiyla birebir tutarli)."""
    with db() as c:
        if not c.execute("SELECT 1 FROM brands WHERE id=?", (brand_id,)).fetchone():
            raise HTTPException(404, "Marka bulunamadi")
        bulk_ids = _load_rows(c, brand_id, "bulk_ids")
    pairs = set()
    for row in bulk_ids:
        camp = (row.get("campaign") or "").strip()
        ag = (row.get("ad_group") or "").strip()
        if camp and ag and row.get("campaign_id") and row.get("ad_group_id"):
            pairs.add((camp, ag))
    by_camp = {}
    for camp, ag in sorted(pairs):
        by_camp.setdefault(camp, []).append(ag)
    return [{"campaign": c, "ad_groups": sorted(set(ags))}
            for c, ags in sorted(by_camp.items())]


@app.get("/api/brands/{brand_id}/insights")
def get_insights(brand_id: int):
    with db() as c:
        brand_row = c.execute("SELECT * FROM brands WHERE id=?", (brand_id,)).fetchone()
        if not brand_row:
            raise HTTPException(404, "Marka bulunamadi")
        brand = dict(brand_row)
        sts = _load_rows(c, brand_id, "search_term")
        tgs = _load_rows(c, brand_id, "targeting")
        plc = _load_rows(c, brand_id, "placement")
        camps = _load_rows(c, brand_id, "campaign")
        uploads = [dict(r) for r in c.execute(
            "SELECT filename,report_type,row_count,uploaded_at FROM uploads "
            "WHERE brand_id=? ORDER BY id DESC LIMIT 20", (brand_id,))]
        rec_rows = c.execute(
            "SELECT type,campaign,keyword,match_type,current_value,suggested_value "
            "FROM recommendations WHERE brand_id=? AND status='pending'",
            (brand_id,)).fetchall()
    if not (sts or tgs):
        return {"empty": True}
    det_recs = [dict(r) for r in rec_rows]
    with db() as c:
        prods = [dict(r) for r in c.execute(
            "SELECT * FROM products WHERE brand_id=?", (brand_id,))]
    data = insights.dashboard(brand, sts, tgs, plc, camps, det_recs, uploads)
    data["profit"] = _profit_calc(brand, prods if prods else None)
    data["products"] = prods
    return data


# ---------- Gunluk is listesi ("bugun ne yapmaliyim?") ----------

@app.get("/api/brands/{brand_id}/today")
def today(brand_id: int):
    """Tum motorlari okuyup tek oncelikli is listesi verir."""
    ctx = _market_context(brand_id)
    brand = ctx["brand"]
    profit = _profit_calc(brand, ctx["products"] or None)

    opp = None
    if ctx["queries"]:
        opp = market_intel.analyze(
            ctx["queries"], search_terms=ctx["search_terms"], catalog=ctx["catalog"],
            basket=ctx["basket"], brand_name=brand["name"], profit=profit,
            known_brands=_split_list(brand.get("competitor_brands")),
            not_brands=_split_list(brand.get("not_brands")))
        opp["period"] = ctx["period"]

    with db() as c:
        rec_rows = [dict(r) for r in c.execute(
            "SELECT type,campaign,keyword,current_value,suggested_value,metrics "
            "FROM recommendations WHERE brand_id=? AND status='pending'", (brand_id,))]
    for r in rec_rows:
        try:
            r["metrics"] = json.loads(r.get("metrics") or "{}")
        except Exception:
            r["metrics"] = {}

    return brain.today(recs=rec_rows, opp=opp, brand=brand, has={
        "ba_query": bool(ctx["queries"]),
        "ba_catalog": bool(ctx["catalog"]),
        "economics": bool(profit),
    })


# ---------- Firsat Radari (Brand Analytics) ----------

def _split_list(s):
    return [p.strip().lower() for p in str(s or "").replace("\n", ",").split(",")
            if p.strip()]


def _market_context(brand_id):
    """Firsat analizi icin gereken tum veriyi tek yerde toplar."""
    with db() as c:
        row = c.execute("SELECT * FROM brands WHERE id=?", (brand_id,)).fetchone()
        if not row:
            raise HTTPException(404, "Marka bulunamadi")
        brand = dict(row)
        # Ceyreklik rapor daha genis veri tabanidir; yoksa aylik ile calis
        queries = _load_rows(c, brand_id, "ba_search_query")
        period = "quarterly"
        if not queries:
            queries = _load_rows(c, brand_id, "ba_search_query_month")
            period = "monthly"
        catalog = _load_rows(c, brand_id, "ba_catalog")
        basket = _load_rows(c, brand_id, "ba_market_basket")
        sts = _load_rows(c, brand_id, "search_term")
        prods = [dict(r) for r in c.execute(
            "SELECT * FROM products WHERE brand_id=?", (brand_id,))]
    return {"brand": brand, "queries": queries, "period": period,
            "catalog": catalog, "basket": basket, "search_terms": sts,
            "products": prods}


@app.get("/api/brands/{brand_id}/opportunities")
def get_opportunities(brand_id: int, min_volume: int = 200,
                      min_relevance: float = 0.34, limit: int = 50):
    ctx = _market_context(brand_id)
    if not ctx["queries"]:
        return {
            "empty": True,
            "message": ("Brand Analytics arama terimi raporu yuklenmemis. "
                        "Seller Central > Marka > Marka Analizi > Arama Terimi "
                        "Performansi (Marka Gorunumu) raporunu indirip normal "
                        "rapor gibi yukleyin."),
        }
    brand = ctx["brand"]
    profit = _profit_calc(brand, ctx["products"] or None)
    res = market_intel.analyze(
        ctx["queries"], search_terms=ctx["search_terms"], catalog=ctx["catalog"],
        basket=ctx["basket"], brand_name=brand["name"], profit=profit,
        min_volume=min_volume, min_relevance=min_relevance, limit_per_bucket=limit,
        known_brands=_split_list(brand.get("competitor_brands")),
        not_brands=_split_list(brand.get("not_brands")))
    res["period"] = ctx["period"]
    res["profit"] = profit
    res["has_ad_data"] = bool(ctx["search_terms"])
    res["brand_lists"] = {
        "competitor_brands": _split_list(brand.get("competitor_brands")),
        "not_brands": _split_list(brand.get("not_brands")),
    }
    return res


class BrandListIn(BaseModel):
    add: list[str] = []          # rakip marka olarak isaretle
    remove: list[str] = []       # "bu marka degil" olarak isaretle


@app.post("/api/brands/{brand_id}/competitor-brands")
def edit_competitor_brands(brand_id: int, body: BrandListIn):
    """Rakip marka listesini duzenler.

    Otomatik tespit guvenilir degil (hacim/sorgu sayisi markayi jenerikten
    ayiramiyor), bu yuzden kullanici duzeltmesi asil mekanizmadir.
    """
    with db() as c:
        row = c.execute("SELECT competitor_brands,not_brands FROM brands WHERE id=?",
                        (brand_id,)).fetchone()
        if not row:
            raise HTTPException(404, "Marka bulunamadi")
        known = set(_split_list(row["competitor_brands"]))
        nots = set(_split_list(row["not_brands"]))
        for t in body.add:
            t = t.strip().lower()
            if t:
                known.add(t)
                nots.discard(t)
        for t in body.remove:
            t = t.strip().lower()
            if t:
                nots.add(t)
                known.discard(t)
        c.execute("UPDATE brands SET competitor_brands=?,not_brands=? WHERE id=?",
                  (",".join(sorted(known)), ",".join(sorted(nots)), brand_id))
    return {"ok": True, "competitor_brands": sorted(known), "not_brands": sorted(nots)}


class OppExport(BaseModel):
    queries: list[str] = []
    bucket: str = "WHITESPACE"
    match: str = "exact"
    asin: str = ""
    sku: str = ""
    tiers: int = 3


@app.post("/api/brands/{brand_id}/opportunities/plan")
def opportunities_plan(brand_id: int, body: OppExport):
    """Secilen firsat kelimelerinden kampanya plani uretir (indirmeden once onizleme)."""
    ctx = _market_context(brand_id)
    if not ctx["queries"]:
        raise HTTPException(400, "Brand Analytics raporu yuklenmemis")
    brand = ctx["brand"]
    res = market_intel.analyze(
        ctx["queries"], search_terms=ctx["search_terms"], catalog=ctx["catalog"],
        basket=ctx["basket"], brand_name=brand["name"],
        profit=_profit_calc(brand, ctx["products"] or None),
        limit_per_bucket=500,
        known_brands=_split_list(brand.get("competitor_brands")),
        not_brands=_split_list(brand.get("not_brands")))
    pool = res["buckets"].get(body.bucket.upper(), [])
    if body.queries:
        want = {q.strip().lower() for q in body.queries}
        pool = [o for o in pool if o["query"] in want]
    if not pool:
        raise HTTPException(400, "Secilen kelimeler bu kovada bulunamadi")

    asin = body.asin.strip().upper()
    if not asin and ctx["catalog"]:
        # En cok satan ASIN'i varsayilan yap
        asin = max(ctx["catalog"], key=lambda c: c.get("purchases") or 0).get("asin", "")
    plan = market_intel.build_campaign_plan(
        pool, asin=asin, sku=body.sku.strip(), period=ctx["period"],
        match=body.match.lower(), tiers=max(1, min(body.tiers, 6)),
        campaign_prefix=f"{body.bucket.upper()}",
        negatives=[n["query"] for n in res.get("negatives", [])])
    if not plan:
        raise HTTPException(400, "Gecerli bid hesaplanamadi - marka ekonomisini kontrol edin")
    return plan


@app.post("/api/brands/{brand_id}/opportunities/export")
def opportunities_export(brand_id: int, body: OppExport):
    """Secilen firsatlari Amazon'a yuklenebilir bulksheet olarak indirir."""
    plan = opportunities_plan(brand_id, body)
    buf = launch_mod.build_bulksheet(plan)
    with db() as c:
        name = c.execute("SELECT name FROM brands WHERE id=?", (brand_id,)).fetchone()[0]
    fname = f"{name}_firsat_{body.bucket.lower()}_{datetime.now():%Y%m%d}.xlsx"
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument"
        ".spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'})


# ---------- Amazon Bulksheet Export ----------

@app.get("/api/brands/{brand_id}/export-bulksheet")
def export_bulksheet(brand_id: int):
    with db() as c:
        brand = c.execute("SELECT * FROM brands WHERE id=?", (brand_id,)).fetchone()
        if not brand:
            raise HTTPException(404, "Marka bulunamadi")
        rows = c.execute(
            "SELECT * FROM recommendations WHERE brand_id=? AND status='approved'",
            (brand_id,)).fetchall()
    if not rows:
        raise HTTPException(400, "Aktarilacak onayli oneri yok")
    recs = []
    for r in rows:
        d = dict(r)
        d["metrics"] = json.loads(d["metrics"] or "{}")
        recs.append(d)
    # Raw report rows'i ID map icin gec
    with db() as c:
        all_rows = []
        for rtype in ("search_term", "targeting", "campaign", "placement", "bulk_ids"):
            all_rows.extend(_load_rows(c, brand_id, rtype))
    # Pre-flight: ID map var mi kontrol et
    has_ids = any(r.get("campaign_id") for r in all_rows)
    if not has_ids:
        raise HTTPException(400,
            "Campaign ID/Ad Group ID bulunamadi. Normal performans raporlari "
            "(Search Term, Targeting, Campaign, Placement) bu ID'leri HICBIR ZAMAN "
            "icermez - kac kere yeniden indirirseniz indirin degismez. "
            "COZUM: Amazon Ads Console > Bulk operations > 'Download spreadsheet' "
            "ile ID eslemesi iceren dosyayi indirin, PPC Asistan'a normal rapor "
            "gibi surukleyip yukleyin (otomatik taninir), sonra Bulksheet'i "
            "tekrar indirin.")
    buf = bulksheet.build(recs, brand["name"], dict(brand), report_rows=all_rows)
    fname = f"{brand['name']}_amazon_bulksheet_{datetime.now():%Y%m%d}.xlsx"
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument"
        ".spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'})


# ---------- AI Chat ----------

class ChatIn(BaseModel):
    message: str


@app.post("/api/brands/{brand_id}/chat")
def chat_send(brand_id: int, body: ChatIn):
    with db() as c:
        brand_row = c.execute("SELECT * FROM brands WHERE id=?", (brand_id,)).fetchone()
        if not brand_row:
            raise HTTPException(404, "Marka bulunamadi")
        brand = dict(brand_row)
        sts = _load_rows(c, brand_id, "search_term")
        tgs = _load_rows(c, brand_id, "targeting")
        camps = _load_rows(c, brand_id, "campaign")
        history = [dict(r) for r in c.execute(
            "SELECT role,content FROM chat_messages WHERE brand_id=? "
            "ORDER BY id ASC LIMIT 40", (brand_id,))]
        now = datetime.now().isoformat(timespec="seconds")
        c.execute(
            "INSERT INTO chat_messages(brand_id,role,content,created_at) "
            "VALUES(?,?,?,?)", (brand_id, "user", body.message, now))
    with db() as c:
        plc = _load_rows(c, brand_id, "placement")
    try:
        reply = chat_mod.reply(brand, sts, tgs, camps, history, body.message,
                                placements=plc)
    except Exception as e:
        raise HTTPException(500, f"Chat hatasi: {e}")
    with db() as c:
        c.execute(
            "INSERT INTO chat_messages(brand_id,role,content,created_at) "
            "VALUES(?,?,?,?)", (brand_id, "assistant", reply,
                                datetime.now().isoformat(timespec="seconds")))
    return {"reply": reply}


@app.get("/api/brands/{brand_id}/data-summary")
def data_summary(brand_id: int):
    """Chat'e 'elimdeki veri' ozetini vermek icin."""
    with db() as c:
        sts = _load_rows(c, brand_id, "search_term")
        tgs = _load_rows(c, brand_id, "targeting")
        camps = _load_rows(c, brand_id, "campaign")
        brand = c.execute("SELECT * FROM brands WHERE id=?", (brand_id,)).fetchone()
    if not brand:
        raise HTTPException(404, "Marka bulunamadi")
    return chat_mod.data_summary(dict(brand), sts, tgs, camps)


@app.get("/api/brands/{brand_id}/chat")
def chat_history(brand_id: int):
    with db() as c:
        rows = c.execute(
            "SELECT role,content,created_at FROM chat_messages WHERE brand_id=? "
            "ORDER BY id ASC LIMIT 80", (brand_id,)).fetchall()
    return [dict(r) for r in rows]


@app.delete("/api/brands/{brand_id}/chat")
def chat_clear(brand_id: int):
    with db() as c:
        c.execute("DELETE FROM chat_messages WHERE brand_id=?", (brand_id,))
    return {"ok": True}


# ---------- Products (coklu urun kar hesabi) ----------

class ProductIn(BaseModel):
    name: str = ""
    asin: str = ""
    sell_price: float = 0
    cogs: float = 0
    amazon_fee_pct: float = 0.15
    fba_fee: float = 0
    share_pct: float = 0


@app.get("/api/brands/{brand_id}/products")
def get_products(brand_id: int):
    with db() as c:
        rows = c.execute(
            "SELECT * FROM products WHERE brand_id=? ORDER BY id",
            (brand_id,)).fetchall()
    return [dict(r) for r in rows]


@app.post("/api/brands/{brand_id}/products")
def add_product(brand_id: int, body: ProductIn):
    with db() as c:
        cur = c.execute(
            "INSERT INTO products(brand_id,name,asin,sell_price,cogs,"
            "amazon_fee_pct,fba_fee,share_pct) VALUES(?,?,?,?,?,?,?,?)",
            (brand_id, body.name.strip(), body.asin.strip().upper(),
             body.sell_price, body.cogs, body.amazon_fee_pct,
             body.fba_fee, body.share_pct))
    return {"id": cur.lastrowid}


@app.put("/api/products/{pid}")
def update_product(pid: int, body: ProductIn):
    with db() as c:
        c.execute(
            "UPDATE products SET name=?,asin=?,sell_price=?,cogs=?,"
            "amazon_fee_pct=?,fba_fee=?,share_pct=? WHERE id=?",
            (body.name.strip(), body.asin.strip().upper(),
             body.sell_price, body.cogs, body.amazon_fee_pct,
             body.fba_fee, body.share_pct, pid))
    return {"ok": True}


@app.delete("/api/products/{pid}")
def delete_product(pid: int):
    with db() as c:
        c.execute("DELETE FROM products WHERE id=?", (pid,))
    return {"ok": True}


# ======================= LAUNCH (sifir urun PPC) =======================
# Chrome uzantisi buraya baglanir: urunu tani -> keyword bul -> kampanya
# plani -> bulk sheet. Mevcut "marka/rapor" akisindan bagimsizdir.

def _loose_int(v):
    """Scraping'den '1,234' / 1234.0 / '' gibi degerler gelir; 422 yerine coerce."""
    if v is None or v == "":
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return int(v)
    digits = "".join(ch for ch in str(v) if ch.isdigit())
    return int(digits) if digits else None


def _loose_float(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    cleaned = "".join(ch for ch in str(v).replace(",", ".") if ch.isdigit() or ch == ".")
    try:
        return float(cleaned)
    except ValueError:
        return None


def _loose_list(v):
    """None veya tek string gelirse listeye cevir."""
    if v is None:
        return []
    if isinstance(v, str):
        return [v] if v.strip() else []
    return v


class CompetitorIn(BaseModel):
    asin: str | None = None
    title: str | None = None
    price: float | None = None
    rating: float | None = None
    review_count: int | None = None
    bsr_rank: int | None = None
    bullets: list[str] = []
    description: str | None = None
    is_best_seller: bool = False
    is_amazon_choice: bool = False
    # Otomatik kesiften gelen sinyaller (uzanti doldurur)
    bought_past_month: int | None = None   # satis hizi
    keyword_overlap: int | None = None     # kac ana kelimede siralaniyor
    avg_rank: float | None = None          # o kelimelerde ortalama sirasi

    _c_ints = field_validator("review_count", "bsr_rank", "bought_past_month",
                              "keyword_overlap", mode="before")(_loose_int)
    _c_floats = field_validator("price", "rating", "avg_rank",
                                mode="before")(_loose_float)
    _c_lists = field_validator("bullets", mode="before")(_loose_list)


class LaunchProductIn(BaseModel):
    title: str
    asin: str | None = None
    sku: str | None = None
    price: float | None = None
    brand: str | None = None
    cogs: float | None = None
    fba_fee: float | None = None
    fee_pct: float | None = 0.15
    bullets: list[str] = []
    description: str | None = None
    rating: float | None = None
    review_count: int | None = None
    bsr: dict | None = None
    competitors: list[CompetitorIn] = []
    search_suggestions: list[str] = []
    catalog_products: list[dict] = []
    use_ai: bool = True
    bid_strategy: str = "profit"   # profit | balanced | aggressive
    measured_cpc: float | None = None  # kendi raporundan gercek ortalama CPC
    brand_id: int | None = None        # varsa markanin olculmus verisi kullanilir
    assumed_cvr: float | None = None   # veri yoksa kullanicinin beklentisi

    _p_ints = field_validator("review_count", mode="before")(_loose_int)
    _p_floats = field_validator(
        "price", "cogs", "fba_fee", "fee_pct", "rating", "measured_cpc",
        mode="before")(_loose_float)
    _p_lists = field_validator(
        "bullets", "search_suggestions", "catalog_products", "competitors",
        mode="before")(_loose_list)

    @field_validator("bsr", mode="before")
    @classmethod
    def _p_bsr(cls, v):
        # Scraping bazen "#1,234 in Beauty" gibi duz string dondurur.
        if v is None or isinstance(v, dict):
            return v
        return {"raw": str(v)}


@app.post("/api/launch/analyze")
def launch_analyze(body: LaunchProductIn):
    """Urun + rakiplerden keyword + kampanya plani uretir (bulk sheet DEGIL)."""
    product = {
        "title": body.title, "asin": (body.asin or "").strip().upper(),
        "sku": body.sku, "price": body.price, "brand": body.brand,
        "cogs": body.cogs, "fba_fee": body.fba_fee, "fee_pct": body.fee_pct,
        "bullets": body.bullets, "description": body.description,
        "rating": body.rating, "review_count": body.review_count,
        "bsr": body.bsr,
        "search_suggestions": body.search_suggestions,
        "catalog_products": body.catalog_products
    }
    competitors = [c.model_dump() for c in body.competitors]
    try:
        # Marka verildiyse KENDI olculmus CPC/CVR'i esas alinir.
        # MARKA IZOLASYONU: veriler SADECE bu brand_id'den cekilir.
        report_rows = ba_rows = None
        brand_name = None
        if body.brand_id:
            with db() as c:
                row = c.execute("SELECT name FROM brands WHERE id=?",
                                (body.brand_id,)).fetchone()
                brand_name = row["name"] if row else None
                rs = c.execute(
                    "SELECT data FROM report_rows WHERE brand_id=? AND report_type='targeting'",
                    (body.brand_id,)).fetchall()
                bs = c.execute(
                    "SELECT data FROM report_rows WHERE brand_id=? AND "
                    "report_type IN ('ba_search_query','ba_search_query_month')",
                    (body.brand_id,)).fetchall()
            report_rows = [json.loads(r["data"]) for r in rs] or None
            ba_rows = [json.loads(r["data"]) for r in bs] or None
        plan = launch_mod.build_plan(product, competitors, use_ai=body.use_ai,
                                     bid_strategy=body.bid_strategy,
                                     measured_cpc=body.measured_cpc,
                                     report_rows=report_rows, ba_rows=ba_rows,
                                     brand_id=body.brand_id, brand_name=brand_name,
                                     assumed_cvr=body.assumed_cvr)
    except Exception as e:
        raise HTTPException(500, f"Plan uretilemedi: {e}")
    return plan


EXT_DIR = Path(__file__).parent / "extension"


@app.get("/api/extension/files")
def extension_files():
    """Uzanti klasorundeki dosyalarin listesi (arayuzde tek tek indirmek icin)."""
    if not EXT_DIR.is_dir():
        raise HTTPException(404, "extension/ klasoru bulunamadi")
    files = sorted(p for p in EXT_DIR.rglob("*") if p.is_file()
                   and not p.name.startswith("."))
    return {"files": [{"name": str(p.relative_to(EXT_DIR)), "bytes": p.stat().st_size}
                      for p in files]}


@app.get("/api/extension/file/{name:path}")
def extension_file(name: str):
    """Tek dosyayi indir. Path traversal'a karsi klasor disina cikilamaz."""
    if not EXT_DIR.is_dir():
        raise HTTPException(404, "extension/ klasoru bulunamadi")
    target = (EXT_DIR / name).resolve()
    if not str(target).startswith(str(EXT_DIR.resolve()) + "/") or not target.is_file():
        raise HTTPException(404, "Dosya bulunamadi")
    return FileResponse(target, media_type="text/plain", filename=target.name)


@app.get("/api/extension/download")
def extension_download():
    """Uzantiyi tek .zip olarak indir -> ac -> chrome://extensions 'Load unpacked'."""
    if not EXT_DIR.is_dir():
        raise HTTPException(404, "extension/ klasoru bulunamadi")
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        for p in sorted(EXT_DIR.rglob("*")):
            if p.is_file() and not p.name.startswith("."):
                z.write(p, Path("ppc-launch-extension") / p.relative_to(EXT_DIR))
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/zip",
        headers={"Content-Disposition":
                 'attachment; filename="ppc-launch-extension.zip"'})


@app.get("/api/brands/{brand_id}/known-products")
def known_products(brand_id: int):
    """Bu markada BILDIGIMIZ urunler: ASIN + baslik + fiyat + olculmus CPC.

    Toplu lansmanda kullanicinin her urunu elle yazmasini onler. Kaynaklar:
      - Brand Analytics katalog raporu (baslik + fiyat)
      - Reklam raporundaki kampanya adlari (daha once reklam verilen ASIN'ler)
      - Kayitli urunler tablosu
    """
    import benchmarks
    with db() as c:
        if not c.execute("SELECT 1 FROM brands WHERE id=?", (brand_id,)).fetchone():
            raise HTTPException(404, "Marka bulunamadi")
        katalog = _load_rows(c, brand_id, "ba_catalog")
        hedef = _load_rows(c, brand_id, "targeting")
        reklamli = _load_rows(c, brand_id, "advertised_product")
        kayitli = [dict(r) for r in c.execute(
            "SELECT asin,name,sell_price,cogs,fba_fee FROM products WHERE brand_id=?",
            (brand_id,))]

    urunler = {}
    for k in katalog:
        a = str(k.get("asin") or "").strip().upper()
        if len(a) == 10:
            urunler[a] = {"asin": a, "title": k.get("title") or "",
                          "price": k.get("price"), "source": "katalog"}
    for p_ in kayitli:
        a = str(p_.get("asin") or "").strip().upper()
        if len(a) == 10:
            u = urunler.setdefault(a, {"asin": a, "title": "", "source": "kayitli"})
            u.update({"title": u.get("title") or p_.get("name") or "",
                      "price": u.get("price") or p_.get("sell_price"),
                      "cogs": p_.get("cogs"), "fba_fee": p_.get("fba_fee")})
    # Advertised Product raporu: ASIN <-> SKU eslesmesi + urun bazinda
    # performans. SKU'yu saglayan TEK kaynak bu.
    ap = {}
    for r in reklamli:
        a = str(r.get("asin") or "").strip().upper()
        if len(a) != 10:
            continue
        g = ap.setdefault(a, {"sku": "", "clicks": 0, "spend": 0.0,
                              "orders": 0, "sales": 0.0})
        if r.get("sku") and not g["sku"]:
            g["sku"] = r["sku"]
        g["clicks"] += r.get("clicks", 0) or 0
        g["spend"] += r.get("spend", 0) or 0.0
        g["orders"] += r.get("orders", 0) or 0
        g["sales"] += r.get("sales", 0) or 0.0
    for a, g in ap.items():
        u = urunler.setdefault(a, {"asin": a, "title": "", "source": "reklam"})
        if g["sku"]:
            u["sku"] = g["sku"]
        if g["clicks"]:
            u.setdefault("clicks", g["clicks"])
            u.setdefault("cpc", round(g["spend"] / g["clicks"], 2))
            u.setdefault("cvr_pct", round(100 * g["orders"] / g["clicks"], 2))
            u.setdefault("measured", g["clicks"] >= benchmarks.MIN_CLICKS_CPC)

    # Reklam verilmis ASIN'ler + olculmus performans
    for m in benchmarks.products_in(hedef):
        u = urunler.setdefault(m["asin"], {"asin": m["asin"], "title": "",
                                           "source": "reklam"})
        u.update({"clicks": m["clicks"], "cpc": m["cpc"], "cvr_pct": m["cvr_pct"],
                  "measured": m["enough_for_cpc"]})
    return {"brand_id": brand_id, "products": sorted(
        urunler.values(), key=lambda x: (-(x.get("clicks") or 0), x["asin"]))}


class BatchLaunchIn(BaseModel):
    """Coklu urun lansmani: her urun icin ayri plan, TEK dosya."""
    products: list[LaunchProductIn] = []
    discovery: bool = True          # True -> Faz 0, False -> Faz 1
    brand_id: int | None = None
    bid_strategy: str = "profit"
    use_ai: bool = True


@app.post("/api/launch/batch")
def launch_batch(body: BatchLaunchIn):
    """6-8 urunu tek seferde planla, TEK bulksheet dondur.

    Her urun kendi ASIN'iyle etiketlendigi icin kampanyalar karismaz;
    olculmus veri de urun bazinda ayrisir.
    """
    if not body.products:
        raise HTTPException(400, "Urun listesi bos")
    if len(body.products) > 20:
        raise HTTPException(400, "Tek seferde en fazla 20 urun")

    report_rows = ba_rows = None
    brand_name = None
    if body.brand_id:
        with db() as c:
            r = c.execute("SELECT name FROM brands WHERE id=?",
                          (body.brand_id,)).fetchone()
            brand_name = r["name"] if r else None
            report_rows = _load_rows(c, body.brand_id, "targeting") or None
            bs = c.execute(
                "SELECT data FROM report_rows WHERE brand_id=? AND "
                "report_type IN ('ba_search_query','ba_search_query_month')",
                (body.brand_id,)).fetchall()
            ba_rows = [json.loads(x["data"]) for x in bs] or None

    planlar, hatalar = [], []
    for u in body.products:
        try:
            urun = {
                "title": u.title, "asin": (u.asin or "").strip().upper(),
                "sku": u.sku, "price": u.price, "brand": u.brand,
                "cogs": u.cogs, "fba_fee": u.fba_fee, "fee_pct": u.fee_pct,
                "bullets": u.bullets, "description": u.description,
                "rating": u.rating, "review_count": u.review_count,
                "bsr": u.bsr, "search_suggestions": u.search_suggestions,
                "catalog_products": u.catalog_products,
            }
            planlar.append(launch_mod.build_plan(
                urun, [c.model_dump() for c in u.competitors],
                use_ai=body.use_ai, bid_strategy=body.bid_strategy,
                report_rows=report_rows, ba_rows=ba_rows,
                brand_id=body.brand_id, brand_name=brand_name))
        except Exception as e:
            hatalar.append({"asin": u.asin, "title": u.title[:40], "error": str(e)})

    if not planlar:
        raise HTTPException(500, f"Hicbir plan uretilemedi: {hatalar[:2]}")
    return {"plans": planlar, "errors": hatalar}


# ---------------------------------------------------------------- bulk doktor
# Amazon konsol raporlarinda Campaign ID YOKTUR; guncelleme dosyasi ancak
# Bulk Operations indirmesinden uretilebilir. Bu iki uc onu yapar.

def _sayi(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _marka_hedefleri(brand_id):
    """Markanin OLCULMUS CVR'i ve hedef ACOS'u. Varsayim uretmez -
    olcum yoksa cagiran tarafa None doner ve kullaniciya sorulur."""
    with db() as c:
        rows = _load_rows(c, brand_id, "targeting")
        brand = c.execute("SELECT * FROM brands WHERE id=?",
                          (brand_id,)).fetchone()
    if brand is None:
        raise HTTPException(404, "Marka bulunamadi")
    b = dict(brand)
    bench = benchmarks.resolve(rows=rows, brand_id=brand_id,
                               brand_name=b.get("name"))
    acct = bench.get("account") or {}
    # Hedef ACOS oncelik sirasi:
    #   1) cagirinin verdigi deger (endpoint parametresi)
    #   2) break-even  - fiyat/maliyet girilmisse hesaplanir
    #   3) markanin kayitli hedef ACOS'u
    # Ucu de yoksa VARSAYIM URETILMEZ; kullaniciya sorulur.
    be = None
    fiyat = _sayi(b.get("sell_price"))
    if fiyat > 0:
        econ = launch_mod.break_even(
            fiyat, _sayi(b.get("cogs")),
            _sayi(b.get("amazon_fee_pct")) or 0.15, _sayi(b.get("fba_fee")))
        be = econ.get("break_even_acos_pct")
    if not be:
        kayitli = _sayi(b.get("target_acos"))
        if kayitli > 0:
            # target_acos oran olarak tutulur (0.30 = %30)
            be = kayitli * 100 if kayitli <= 1 else kayitli
    return {"cvr": acct.get("cvr"), "cvr_clicks": acct.get("clicks"),
            "break_even_acos_pct": be, "cpc": acct.get("cpc"),
            "brand_name": b.get("name"),
            "acos_source": ("break-even (fiyat/maliyetten)" if fiyat > 0
                            else "markanin kayitli hedef ACOS'u")}


@app.post("/api/brands/{brand_id}/bulk-doctor")
async def bulk_doctor_teshis(brand_id: int, file: UploadFile,
                             target_acos: float = None):
    """Bulk Operations dosyasini teshis et. Dosya URETMEZ - once gosterir."""
    ham = await file.read()
    try:
        bulk = bulk_doctor.read_bulk(ham)
    except ValueError as e:
        raise HTTPException(400, str(e))

    h = _marka_hedefleri(brand_id)
    if not h["cvr"]:
        raise HTTPException(400,
            "Bu marka icin olculmus donusum orani yok. Once Targeting "
            "raporunu yukle - varsayimla teklif degistirmek riskli.")
    hedef = target_acos or h["break_even_acos_pct"]
    if not hedef:
        raise HTTPException(400,
            "Hedef ACOS belirlenemedi. Marka ayarlarinda fiyat/maliyet ya da "
            "hedef ACOS gir, veya bu ekrandan hedef ACOS yaz.")

    d = bulk_doctor.diagnose(bulk, hedef, h["cvr"],
                             fallback_bid=h.get("cpc") or 2.00)
    u = bulk_doctor.utilization(bulk)
    return {
        "campaigns_live": d["campaigns_live"],
        "target_acos_pct": round(hedef, 1),
        "measured_cvr_pct": round(h["cvr"] * 100, 2),
        "cvr_clicks": h["cvr_clicks"],
        "utilization": u,
        "actions": d["actions"],
        "untouched": d["notes"],
    }


@app.get("/api/brands/{brand_id}/competitiveness")
def rekabet_gucu(brand_id: int, acos_ceiling: float = 1.00):
    """TEKLIFIM PAZARI KARSILIYOR MU?

    Kullanicinin en buyuk endisesi: teklif dusuk kalir, gosterim gelmez,
    ciro tutmaz. Bu uc onu OLCER - tahmin degil, olculmus veriyle.

    Ekonomik tavan (tiklama basina ciro x kabul edilen ACOS) pazar CPC'sini
    karsilamiyorsa, o match type'ta rekabet edilemez. Cozum ya sepeti/CVR'i
    buyutmek ya da daha yuksek ACOS'u bilerek kabul etmektir.
    """
    with db() as c:
        rows = _load_rows(c, brand_id, "targeting")
        brand = c.execute("SELECT * FROM brands WHERE id=?", (brand_id,)).fetchone()
    if brand is None:
        raise HTTPException(404, "Marka bulunamadi")
    if not rows:
        raise HTTPException(400, "Targeting raporu yok - once veri yukle.")
    bench = benchmarks.resolve(rows=rows, brand_name=dict(brand).get("name"))
    acct = bench.get("account") or {}
    out = []
    for k in ("exact", "phrase", "broad", "auto", "pt"):
        cvr = (bench.get("cvr") or {}).get(k)
        cpc = (bench.get("cpc") or {}).get(k)
        aov = (bench.get("aov") or {}).get(k) or acct.get("aov")
        if not cvr or not cpc or not aov:
            continue
        rpc = round(aov * cvr, 2)                 # tiklama basina ciro
        tavan = benchmarks.economic_ceiling(aov, cvr, acos_ceiling)
        oran = (tavan / cpc) if cpc else 0
        out.append({
            "match": k,
            "revenue_per_click": rpc,
            "ceiling": tavan,
            "market_cpc": round(cpc, 2),
            "ratio": round(oran, 2),
            "status": ("rahat" if oran >= 1.15 else
                       "sinirda" if oran >= 0.95 else "yetmez"),
            # Pazari karsilamak icin kabul edilmesi gereken ACOS
            "acos_to_compete_pct": round(cpc / rpc * 100) if rpc else None,
        })
    yetmez = [o for o in out if o["status"] == "yetmez"]
    return {
        "brand": dict(brand).get("name"),
        "acos_ceiling_pct": round(acos_ceiling * 100),
        "rows": out,
        "uncompetitive": len(yetmez),
        "verdict": ("Mevcut ekonomiyle pazarda rekabet edebilirsin."
                    if not yetmez else
                    f"{len(yetmez)} match type'ta teklif pazari karsilamiyor - "
                    f"gosterim alamazsin. Ya daha yuksek ACOS kabul et, ya "
                    f"sepeti/donusumu buyut."),
    }


@app.post("/api/brands/{brand_id}/bulk-doctor/verify")
async def bulk_doctor_dogrula(brand_id: int, file: UploadFile,
                              acos_ceiling: float = 1.00):
    """Yukleme SONRASI dogrulama: hesap gercekten istedigimiz halde mi?

    "106/109 basarili" demek "hesap dogru" demek degildir. Bu uc, yazilan
    degere degil, hesabin SON HALINE bakar."""
    ham = await file.read()
    try:
        bulk = bulk_doctor.read_bulk(ham)
    except ValueError as e:
        raise HTTPException(400, str(e))
    with db() as c:
        rows = _load_rows(c, brand_id, "targeting")
        brand = c.execute("SELECT * FROM brands WHERE id=?", (brand_id,)).fetchone()
    if brand is None:
        raise HTTPException(404, "Marka bulunamadi")
    ad = dict(brand).get("name")
    tavan = verify_mod.ceilings_for(rows, ad, acos_ceiling) if rows else None
    sonuc = verify_mod.audit(bulk, ceilings=tavan)
    sonuc["ceilings"] = tavan
    sonuc["acos_ceiling_pct"] = round(acos_ceiling * 100)
    return sonuc


@app.post("/api/brands/{brand_id}/bulk-doctor/file")
async def bulk_doctor_dosya(brand_id: int, file: UploadFile,
                            target_acos: float = None):
    """Teshisi uygulayan Update dosyasini uretir."""
    ham = await file.read()
    try:
        bulk = bulk_doctor.read_bulk(ham)
    except ValueError as e:
        raise HTTPException(400, str(e))

    h = _marka_hedefleri(brand_id)
    if not h["cvr"]:
        raise HTTPException(400, "Olculmus donusum orani yok - once "
                                 "Targeting raporunu yukle.")
    hedef = target_acos or h["break_even_acos_pct"]
    if not hedef:
        raise HTTPException(400, "Hedef ACOS belirlenemedi.")

    d = bulk_doctor.diagnose(bulk, hedef, h["cvr"],
                             fallback_bid=h.get("cpc") or 2.00)
    if not d["actions"]:
        raise HTTPException(400, "Duzeltilecek bir sey bulunamadi - "
                                 "kampanyalar hedefle uyumlu gorunuyor.")
    veri, sayac = bulk_doctor.build_update(bulk, d["actions"])
    ad = re.sub(r"[^A-Za-z0-9]+", "-",
                (h.get("brand_name") or f"marka{brand_id}")).strip("-")
    return StreamingResponse(
        io.BytesIO(veri),
        media_type="application/vnd.openxmlformats-officedocument."
                   "spreadsheetml.sheet",
        headers={"Content-Disposition":
                 f'attachment; filename="{ad}-DUZELTME.xlsx"',
                 "X-Changes": json.dumps(sayac)})


@app.post("/api/launch/batch-bulksheet")
def launch_batch_bulksheet(body: dict):
    """Coklu planlari TEK bulksheet'e birlestirir."""
    planlar = body.get("plans") or []
    if not planlar:
        raise HTTPException(400, "Plan listesi bos")
    try:
        buf, ozet = launch_mod.build_batch_bulksheet(
            planlar, discovery=bool(body.get("discovery", True)))
    except Exception as e:
        raise HTTPException(500, f"Toplu bulksheet uretilemedi: {e}")
    if not buf or not ozet:
        # Faz 0 istenmis ama urunlerin olculmus CPC'si zaten var -> kesif
        # gerekmiyor. Bos dosya vermek yerine ne yapmasi gerektigini soyle.
        raise HTTPException(
            400, "Uretilecek kampanya yok. Bu urunlerin olculmus CPC'si zaten "
                 "var ise Faz 0 gerekmez - discovery=false ile Faz 1 iste.")
    faz = "FAZ0" if body.get("discovery", True) else "FAZ1"
    marka = (planlar[0].get("product", {}).get("brand") or "launch")
    safe = "".join(ch for ch in str(marka)[:20] if ch.isalnum() or ch in " -_").strip() or "launch"
    fname = f"{safe}_{faz}_TOPLU_{len(ozet)}urun_{datetime.now():%Y%m%d}.xlsx"
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument"
        ".spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"',
                 "X-Product-Count": str(len(ozet))})


@app.post("/api/launch/discovery-bulksheet")
def launch_discovery_bulksheet(plan: dict):
    """FAZ 0 kesif kampanyasi (gercek CPC olcumu icin)."""
    try:
        buf = launch_mod.build_discovery_campaign(plan)
    except Exception as e:
        raise HTTPException(500, f"Kesif kampanyasi uretilemedi: {e}")
    if not buf:
        raise HTTPException(400, "Bu plan icin kesif fazi gerekmiyor "
                                 "(olculmus CPC zaten var) ya da kelime yok.")
    title = (plan.get("product", {}).get("brand")
             or plan.get("product", {}).get("title", "launch"))
    safe = "".join(ch for ch in title[:24] if ch.isalnum() or ch in " -_").strip() or "launch"
    fname = f"{safe}_FAZ0_cpc_kesif_{datetime.now():%Y%m%d}.xlsx"
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument"
        ".spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'})


@app.post("/api/launch/bulksheet")
def launch_bulksheet(plan: dict):
    """Analyze'den donen plani (veya elle duzenlenmis halini) bulk sheet yapar."""
    if not plan.get("campaigns"):
        raise HTTPException(400, "Plan bos - once /api/launch/analyze cagir.")
    try:
        buf = launch_mod.build_bulksheet(plan)
    except Exception as e:
        raise HTTPException(500, f"Bulk sheet uretilemedi: {e}")
    title = (plan.get("product", {}).get("brand")
             or plan.get("product", {}).get("title", "launch"))
    safe = "".join(ch for ch in title[:24] if ch.isalnum() or ch in " -_").strip() or "launch"
    fname = f"{safe}_launch_bulksheet_{datetime.now():%Y%m%d}.xlsx"
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument"
        ".spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'})
