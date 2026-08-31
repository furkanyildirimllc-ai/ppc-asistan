"""Amazon Ads rapor dosyalarini (csv/xlsx) okur, tipini otomatik tanir, normalize eder."""
import csv
import io
import re

import openpyxl

# Rapor tipini ayirt eden imza kolonlari
SIGNATURES = [
    # TUM LISTELEMELER RAPORU (Seller Central > Envanter > Raporlar)
    # SIFIR reklam gecmisi olan marka icin TEK GEREKLI DOSYA. Icinde
    # SKU + ASIN + baslik + fiyat + stok var; reklam raporu gerektirmez.
    # Diger tum raporlar reklam calismis olmasini sart kosar - yeni markada
    # hicbiri yoktur. Sekmeyle ayrilmis .txt olarak iner.
    ("listings", {"seller-sku", "asin1"}),
    ("listings", {"seller-sku", "item-name"}),
    ("bulk_ids", {"Entity", "Operation", "Campaign ID"}),
    # Advertised Product raporu: ASIN <-> SKU eslesmesini tasiyan TEK reklam
    # raporu. Diger raporlarda SKU yoktur; bulksheet icin SKU zorunludur.
    ("advertised_product", {"Advertised SKU"}),
    ("advertised_product", {"Advertised ASIN", "Impressions"}),
    ("search_term_is", {"Customer Search Term", "Search Term Impression Rank"}),
    ("search_term", {"Customer Search Term", "Match Type"}),
    ("targeting", {"Targeting", "Top-of-search Impression Share"}),
    ("placement", {"Placement", "Bidding strategy"}),
    ("campaign", {"Budget Amount", "Targeting Type"}),
    # Brand Analytics (Seller Central > Marka Analizi) raporlari.
    # Bunlar reklam raporu DEGIL - pazarin tamamini gosterir: her arama
    # teriminde toplam talep + senin marka payin. Reklam raporunda olmayan
    # "rakipler ne kadar aliyor" bilgisi burada.
    ("ba_search_query", {"Search Query", "Search Query Volume"}),
    ("ba_catalog", {"ASIN Title", "Impressions: Impressions"}),
    ("ba_market_basket", {"#1 Purchase Combination: ASIN"}),
    # Brand Analytics "Top Search Terms": her sorguda EN COK TIKLANAN
    # markalar ve urunler + tiklama paylari. Rakip istihbaratinin en
    # dogrudan kaynagi - kimin kazandigini ACIKCA soyler.
    ("ba_top_terms", {"Search Frequency Rank", "Top Clicked Brand #1"}),
]

REPORT_LABELS = {
    "listings": "Tum Listelemeler Raporu (SKU + ASIN + baslik + fiyat)",
    "search_term": "Search Term Raporu",
    "search_term_is": "Search Term Impression Share Raporu",
    "targeting": "Targeting Raporu",
    "placement": "Placement Raporu",
    "campaign": "Kampanya Raporu",
    "bulk_ids": "Bulk Operations (Campaign/Ad Group ID eslemesi)",
    "advertised_product": "Advertised Product Raporu (ASIN + SKU)",
    "ba_search_query": "Brand Analytics - Arama Terimi Performansi (Ceyrek)",
    "ba_search_query_month": "Brand Analytics - Arama Terimi Performansi (Ay)",
    "ba_catalog": "Brand Analytics - Katalog Performansi",
    "ba_market_basket": "Brand Analytics - Sepet Analizi",
    "ba_top_terms": "Brand Analytics - En Cok Aranan Terimler (rakip istihbarati)",
}

ASIN_RE = re.compile(r"^b0[a-z0-9]{8}$", re.IGNORECASE)


def _num(v):
    """'$1,234.56', '12.5%', '-', None -> float"""
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("$", "").replace(",", "").replace("%", "")
    if s in ("", "-", "None"):
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def _pct(v):
    """Yuzde degerini orana cevirir. CSV'de '39.7%' -> 0.397, xlsx'te 0.397 zaten oran.

    Mantik: xlsx'te ACOS bazen 0.397 (oran) bazen 39.7 (yuzde) olarak gelir.
    Esik 1.0: 1'den buyukse yuzde kabul et (39.7 -> 0.397), kucukse oran kabul et.
    Eski esik 5.0'ti — bu %1-5 arasi ACOS degerlerini yanlis yorumluyordu!
    """
    if isinstance(v, str) and "%" in v:
        return _num(v) / 100.0
    n = _num(v)
    return n / 100.0 if n > 1.0 else n


def _read_sheet(ws):
    it = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h is not None else "" for h in next(it, ())]
    out = []
    for row in it:
        if row is None or all(v is None for v in row):
            continue
        out.append(dict(zip(headers, row)))
    return headers, out


META_RE = re.compile(r'([^,=]+)=\["([^"]*)"\]')


def _parse_meta_line(cells):
    """Brand Analytics CSV'lerinin ilk satiri veri degil, filtre ozetidir:
    'Reporting Range=["Quarterly"],Select year=["2026"],Select quarter=["2"]'
    Bunu dict'e cevirir. Normal rapor satiriysa None doner."""
    line = ",".join(str(c) for c in cells)
    if "=[" not in line:
        return None
    pairs = META_RE.findall(line)
    return {k.strip(): v.strip() for k, v in pairs} if pairs else None


def read_rows(filename, content: bytes):
    """Dosyayi header listesi + dict satirlari + meta olarak dondurur.

    Amazon'un 'Bulk operations > Download spreadsheet' dosyasi COK SEKMELIDIR
    (Portfolios, Sponsored Products Campaigns, Sponsored Brands Campaigns, ...)
    ve openpyxl'in 'aktif' sekmesi bizim istedigimiz veri olmayabilir (ornegin
    Portfolios sekmesi acik kalmis olabilir). Once aktif sekmeyi dene, taninmazsa
    diger sekmeleri sirayla tara ve ilk taninan sekmeyi kullan.
    """
    ad = filename.lower()
    if ad.endswith((".csv", ".txt", ".tsv")):
        text = content.decode("utf-8-sig", errors="replace")
        # Amazon "Tum Listelemeler" raporunu SEKMEYLE ayrilmis .txt verir;
        # reklam raporlari virgullu .csv. Ayiraci ilk satirdan tespit et -
        # uzantiya guvenmek yanlis, kullanici dosyayi yeniden adlandirabilir.
        ilk = text.split("\n", 1)[0]
        ayirac = "\t" if ilk.count("\t") > ilk.count(",") else ","
        reader = csv.reader(io.StringIO(text), delimiter=ayirac)
        rows = [r for r in reader if any(c.strip() for c in r)]
        if not rows:
            return [], [], {}
        meta = _parse_meta_line(rows[0])
        if meta:
            rows = rows[1:]
            if not rows:
                return [], [], meta
        headers = [h.strip() for h in rows[0]]
        return headers, [dict(zip(headers, r)) for r in rows[1:]], (meta or {})
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    headers, out = _read_sheet(wb.active)
    meta = _parse_meta_line(headers)
    if meta and out:
        # Ilk satir meta ise gercek basliklar ikinci satirda
        headers = [str(h).strip() if h is not None else "" for h in out[0].values()]
        out = [dict(zip(headers, r.values())) for r in out[1:]]
    if detect_type(headers) is None:
        for name in wb.sheetnames:
            if wb[name] is wb.active:
                continue
            h2, r2 = _read_sheet(wb[name])
            if detect_type(h2) is not None:
                headers, out = h2, r2
                break
    wb.close()
    return headers, out, (meta or {})


def detect_type(headers):
    hs = set(headers)
    for rtype, sig in SIGNATURES:
        if sig <= hs:
            return rtype
    # Fallback: case-insensitive eslestirme (Amazon bazen kolon adlarini degistirir)
    hs_lower = {h.lower() for h in headers}
    for rtype, sig in SIGNATURES:
        if {s.lower() for s in sig} <= hs_lower:
            return rtype
    # ---- ESNEK GERI DUSUS ----
    # Amazon rapor kolonlarini zaman zaman degistiriyor; katı imza eslesmesi
    # tutmayinca kullanici hicbir sey yukleyemiyor. Ayirt edici tek kolon +
    # metrik kolonlari varsa tipi yine de belirle.
    metrik = {"impressions", "clicks", "spend"}
    if metrik & hs_lower:
        if "customer search term" in hs_lower:
            return "search_term"
        if "targeting" in hs_lower:
            return "targeting"
        if "placement" in hs_lower:
            return "placement"
    return None


def _sayi(v):
    """Metni sayiya cevirir; olmuyorsa 0.0. Rapor alanlari bos gelebilir."""
    try:
        return float(str(v).replace("$", "").replace(",", "").strip() or 0)
    except (TypeError, ValueError):
        return 0.0


def _norm_listing(r):
    """Tum Listelemeler satirini normalize eder.

    Amazon bu raporu sekmeyle ayrilmis, kucuk harfli-tireli kolonlarla verir.
    Bize lazim olan: SKU (reklam icin ZORUNLU), ASIN, baslik, fiyat, stok.
    """
    def al(*adlar):
        for a in adlar:
            for k in r:
                if str(k).strip().lower() == a:
                    return r[k]
        return None

    asin = str(al("asin1", "asin", "product-id") or "").strip().upper()
    return {
        "sku": str(al("seller-sku", "sku") or "").strip(),
        "asin": asin if len(asin) == 10 else "",
        "title": str(al("item-name", "title") or "").strip(),
        "price": _sayi(al("price")),
        # FBA urunlerde bu alan BOS gelir - stok Amazon deposundadir ve bu
        # raporda gorunmez. Bosu 0 yapmak "stok yok" yanilgisi uretir.
        "quantity": (_sayi(al("quantity"))
                     if str(al("quantity") or "").strip() != "" else None),
        "status": str(al("status") or "").strip().lower(),
        "fulfillment": str(al("fulfillment-channel") or "").strip(),
    }


def _yuzde(v):
    """'6.230%' -> 0.0623. Amazon bu raporda yuzdeleri metin olarak verir."""
    t = str(v or "").replace("%", "").replace(",", ".").strip()
    try:
        return round(float(t) / 100.0, 5)
    except (TypeError, ValueError):
        return None


def _norm_top_terms(r, meta=None):
    """Top Search Terms satiri: bir sorguda kim kazaniyor?

    frequency_rank : dusuk = cok araniyor (1 = en cok aranan)
    brands         : en cok tiklanan 3 marka - RAKIPLER
    products       : ASIN + tiklama payi - kim ne kadar aliyor
    """
    def al(*adlar):
        for a in adlar:
            for k in r:
                if str(k).strip().lower() == a.lower():
                    return r[k]
        return None

    urunler = []
    for i in (1, 2, 3):
        a = str(al(f"Top Clicked Product #{i}: ASIN") or "").strip().upper()
        if len(a) != 10:
            continue
        urunler.append({
            "asin": a,
            "title": str(al(f"Top Clicked Product #{i}: Product Title") or "")[:160],
            "click_share": _num(al(f"Top Clicked Product #{i}: Click Share")),
            "conversion_share": _num(
                al(f"Top Clicked Product #{i}: Conversion Share")),
        })
    markalar = [str(al(f"Top Clicked Brand{s} #{i}") or "").strip()
                for i, s in ((1, ""), (2, "s"), (3, "s"))]
    return {
        "term": str(al("Search Term") or "").strip().lower(),
        "frequency_rank": _num(al("Search Frequency Rank")),
        "brands": [b for b in markalar if b],
        "categories": [str(al(f"Top Clicked Category #{i}") or "").strip()
                       for i in (1, 2, 3)],
        "products": urunler,
        "period": (meta or {}).get("Reporting Range", ""),
    }


def _norm_search_term_is(r):
    """Search Term Impression Share satiri.

    Iki kritik alan:
      rank  : o sorguda kacinci sirada gorunuyorsun (1 = en ust)
      share : gorunebilecegin gosterimlerin yuzde kacini aldin
    Ikisi birlikte "bu kelimede buyume alanim var mi" sorusunu cevaplar.
    """
    def al(*adlar):
        for a in adlar:
            for k in r:
                if str(k).strip().lower().rstrip() == a:
                    return r[k]
        return None

    return {
        "term": str(al("customer search term") or "").strip().lower(),
        "targeting": str(al("targeting") or "").strip(),
        "match_type": str(al("match type") or "").strip().upper(),
        "campaign": str(al("campaign name") or "").strip(),
        "ad_group": str(al("ad group name") or "").strip(),
        "rank": _num(al("search term impression rank")),
        "impression_share": _yuzde(al("search term impression share")),
        "impressions": _num(al("impressions")),
        "clicks": _num(al("clicks")),
        "spend": _num(al("spend")),
        "orders": _num(al("7 day total orders (#)", "7 day total orders")),
        "sales": _num(al("7 day total sales", "7 day total sales ")),
    }


def parse(filename, content: bytes):
    """-> (report_type, normalized_rows)"""
    headers, rows, meta = read_rows(filename, content)
    rtype = detect_type(headers)
    if rtype is None:
        raise ValueError(
            f"Rapor tipi taninamadi. Kolonlar: {', '.join(headers[:8])}...")
    if rtype == "listings":
        # Yalnizca ASIN'i ve SKU'su olan AKTIF urunler ise yarar.
        temiz = []
        for r in rows:
            n = _norm_listing(r)
            if not n["asin"] or not n["sku"]:
                continue
            if n["status"] and "inactive" in n["status"]:
                continue
            temiz.append(n)
        if not temiz:
            raise ValueError(
                "Listeleme raporunda kullanilabilir urun bulunamadi. "
                "ASIN ve SKU tasiyan aktif urun gerekiyor.")
        return rtype, temiz
    if rtype == "ba_search_query":
        # Ayni kolonlarla hem aylik hem ceyreklik dosya gelir. Ayri tiplerde
        # sakla ki biri digerinin ustune yazmasin (trend karsilastirmasi icin
        # ikisi de lazim).
        period = (meta.get("Reporting Range") or "").lower()
        if period.startswith("month"):
            rtype = "ba_search_query_month"
        return rtype, [_norm_ba_query(r, meta) for r in rows if r.get("Search Query")]
    if rtype == "ba_catalog":
        return rtype, [_norm_ba_catalog(r, meta) for r in rows if r.get("ASIN")]
    if rtype == "ba_top_terms":
        return rtype, [_norm_top_terms(r, meta) for r in rows
                       if r.get("Search Term")]
    if rtype == "ba_market_basket":
        return rtype, [_norm_ba_basket(r) for r in rows if r.get("ASIN")]
    if rtype == "search_term_is":
        # HATA GECMISI: bu tip icin DAL YOKTU - dosya taniniyordu ama
        # `return rtype, []` ile bos donuyordu. Kullanici "0 satir" gordu.
        # Bu rapor gosterim PAYINI tasir: bir sorguda kac kez gorunebilecekken
        # kac kez gorunmussun. "Kelime kazaniyor ama daha fazla alabilir miyim"
        # sorusunun TEK cevabi burasi.
        return rtype, [_norm_search_term_is(r) for r in rows
                       if r.get("Customer Search Term")]
    if rtype == "search_term":
        return rtype, [_norm_search_term(r) for r in rows]
    if rtype == "targeting":
        return rtype, [_norm_targeting(r) for r in rows]
    if rtype == "campaign":
        return rtype, [_norm_campaign(r) for r in rows]
    if rtype == "placement":
        return rtype, [_norm_placement(r) for r in rows]
    if rtype == "bulk_ids":
        return rtype, [_norm_bulk_ids(r) for r in rows]
    if rtype == "advertised_product":
        return rtype, [_norm_advertised(r) for r in rows]
    # impression share: simdilik saklamiyoruz
    return rtype, []


def _norm_ba_query(r, meta):
    """Brand Analytics 'Search Query Performance' satiri.

    ONEMLI: Bu dosyada 'Share %' kolonlari zaten 0-100 arasi yuzde olarak gelir
    (18.54 = %18.54). _pct() kullanma - 1'den kucuk gercek yuzdeleri (orn %0.9)
    orana cevirip bozar.
    """
    vol = _num(r.get("Search Query Volume"))
    clicks_total = _num(r.get("Clicks: Total Count"))
    pur_total = _num(r.get("Purchases: Total Count"))
    return {
        "query": str(r.get("Search Query") or "").strip().lower(),
        "score": _num(r.get("Search Query Score")),
        "volume": vol,
        "period": (meta.get("Reporting Range") or "").lower(),
        "imp_total": _num(r.get("Impressions: Total Count")),
        "imp_brand": _num(r.get("Impressions: Brand Count")),
        "imp_share": _num(r.get("Impressions: Brand Share %")),
        "clicks_total": clicks_total,
        "click_rate": _num(r.get("Clicks: Click Rate %")),
        "clicks_brand": _num(r.get("Clicks: Brand Count")),
        "click_share": _num(r.get("Clicks: Brand Share %")),
        "cart_total": _num(r.get("Cart Adds: Total Count")),
        "cart_brand": _num(r.get("Cart Adds: Brand Count")),
        "cart_share": _num(r.get("Cart Adds: Brand Share %")),
        "pur_total": pur_total,
        "pur_brand": _num(r.get("Purchases: Brand Count")),
        "pur_share": _num(r.get("Purchases: Brand Share %")),
        # Fiyat karsilastirmasi: pazar ne fiyata satiyor vs sen ne fiyata
        "market_price": _num(r.get("Purchases: Price (Median)")),
        "brand_price": _num(r.get("Purchases: Brand Price (Median)")),
        "click_price": _num(r.get("Clicks: Price (Median)")),
        "brand_click_price": _num(r.get("Clicks: Brand Price (Median)")),
        # Pazarin tiklamadan satisa donusum orani - bid hesabinin temeli
        "market_cvr": round(pur_total / clicks_total * 100, 2) if clicks_total else 0.0,
        "is_asin": bool(ASIN_RE.match(str(r.get("Search Query") or "").strip())),
    }


def _norm_ba_catalog(r, meta):
    """Brand Analytics 'Search Catalog Performance' - ASIN bazli huni."""
    clicks = _num(r.get("Clicks: Clicks"))
    pur = _num(r.get("Purchases: Purchases"))
    return {
        "asin": str(r.get("ASIN") or "").strip().upper(),
        "title": str(r.get("ASIN Title") or "").strip(),
        "category": str(r.get("Category") or "").strip(),
        "impressions": _num(r.get("Impressions: Impressions")),
        "clicks": clicks,
        "ctr": _num(r.get("Clicks: Click Rate (CTR)")),
        "cart_adds": _num(r.get("Cart Adds: Cart Adds")),
        "purchases": pur,
        "search_sales": _num(r.get("Purchases: Search Traffic Sales")),
        "cvr": _num(r.get("Purchases: Conversion Rate %")),
        "price": _num(r.get("Purchases: Price (Median)")
                      or r.get("Impressions: Price (Median)")),
        "rating": _num(r.get("Impressions: Rating (Median)")),
        "click_to_purchase": round(pur / clicks * 100, 2) if clicks else 0.0,
    }


def _norm_ba_basket(r):
    """Brand Analytics 'Market Basket' - birlikte alinan urunler.
    Product targeting ve bundle kampanyalari icin kullanilir."""
    combos = []
    for i in (1, 2, 3):
        asin = str(r.get(f"#{i} Purchase Combination: ASIN") or "").strip().upper()
        if not asin:
            continue
        combos.append({
            "asin": asin,
            "title": str(r.get(f"#{i} Purchase Combination: Product Title") or "").strip(),
            "pct": _num(r.get(f"#{i} Purchase Combination: Combination %")),
        })
    return {
        "asin": str(r.get("ASIN") or "").strip().upper(),
        "title": str(r.get("Product Title") or "").strip(),
        "brand": str(r.get("Brand Name") or "").strip(),
        "orders": _num(r.get("Number of Orders")),
        "combos": combos,
    }


def _base_metrics(r):
    return {
        "impressions": _num(r.get("Impressions")),
        "clicks": _num(r.get("Clicks")),
        "spend": _num(r.get("Spend")),
        "sales": _num(r.get("7 Day Total Sales ") or r.get("7 Day Total Sales")),
        "orders": _num(r.get("7 Day Total Orders (#)")),
        "cpc": _num(r.get("Cost Per Click (CPC)")),
        "acos": _pct(r.get("Total Advertising Cost of Sales (ACOS) ")
                     or r.get("Total Advertising Cost of Sales (ACOS)")),
    }


def _id(v):
    """Amazon ID kolonlarini normalize et (float/int/str)."""
    if v is None or v == "":
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _ids(r):
    """Amazon rapor kolonlarindan Campaign/AdGroup/Portfolio/Keyword/Target ID'lerini yakala."""
    return {
        "campaign_id": _id(r.get("Campaign ID")),
        "ad_group_id": _id(r.get("Ad Group ID")),
        "portfolio_id": _id(r.get("Portfolio ID")),
        "keyword_id": _id(r.get("Keyword ID") or r.get("Keyword or Product Targeting ID")),
        "targeting_id": _id(r.get("Product Targeting ID")),
    }


def _norm_search_term(r):
    d = _base_metrics(r)
    d.update(_ids(r))
    d.update({
        "campaign": str(r.get("Campaign Name") or "").strip(),
        "ad_group": str(r.get("Ad Group Name") or "").strip(),
        "targeting": str(r.get("Targeting") or "").strip(),
        "match_type": str(r.get("Match Type") or "").strip().upper(),
        "term": str(r.get("Customer Search Term") or "").strip().lower(),
    })
    d["is_asin"] = bool(ASIN_RE.match(d["term"]))
    return d


def _norm_targeting(r):
    d = _base_metrics(r)
    d.update(_ids(r))
    d.update({
        "campaign": str(r.get("Campaign Name") or "").strip(),
        "ad_group": str(r.get("Ad Group Name") or "").strip(),
        "targeting": str(r.get("Targeting") or "").strip(),
        "match_type": str(r.get("Match Type") or "").strip().upper(),
        "tos_is": _pct(r.get("Top-of-search Impression Share")),
    })
    return d


def _norm_placement(r):
    d = _base_metrics(r)
    d.update(_ids(r))
    d.update({
        "campaign": str(r.get("Campaign Name") or "").strip(),
        "placement": str(r.get("Placement") or "").strip(),
        "bidding_strategy": str(r.get("Bidding strategy") or "").strip(),
    })
    return d


def _norm_advertised(r):
    """Advertised Product raporu: hangi ASIN hangi SKU ile reklam veriliyor.

    Bu rapor, bulksheet icin zorunlu olan SKU'yu saglayan tek kaynaktir.
    Ayrica urun bazinda performansi kampanya adindan cikarmaya gerek
    kalmadan dogrudan verir.
    """
    def _f(*adlar):
        for a in adlar:
            if r.get(a) not in (None, ""):
                return r.get(a)
        return None
    return {
        "campaign": str(_f("Campaign Name", "Campaign") or "").strip(),
        "ad_group": str(_f("Ad Group Name", "Ad Group") or "").strip(),
        "asin": str(_f("Advertised ASIN", "ASIN") or "").strip().upper(),
        "sku": str(_f("Advertised SKU", "SKU") or "").strip(),
        "impressions": _num(_f("Impressions")),
        "clicks": _num(_f("Clicks")),
        "spend": _num(_f("Spend", "Spend(USD)", "Cost")),
        "sales": _num(_f("7 Day Total Sales", "7 Day Total Sales (USD)", "Sales")),
        "orders": _num(_f("7 Day Total Orders (#)", "Orders", "7 Day Total Orders")),
    }


def _norm_bulk_ids(r):
    """Amazon 'Bulk Operations' indirmesinden Campaign/Ad Group/Keyword ID
    eslemesini cikarir. Bu dosyada metrik yok - sadece isim<->ID eslemesi icin
    kullanilir.

    ONEMLI: Amazon ust seviye satirlarda (Campaign, Ad Group) ismi 'Campaign
    Name'/'Ad Group Name' kolonuna yazar, ama alt seviye satirlarda (Keyword,
    Product Targeting) bu kolonlar BOS birakilir - isim sadece 'Campaign Name
    (Informational only)' / 'Ad Group Name (Informational only)' kolonlarinda
    bulunur. Ikisini de dener, hangisi doluysa onu kullanir."""
    d = _ids(r)
    campaign = r.get("Campaign Name") or r.get("Campaign Name (Informational only)")
    ad_group = r.get("Ad Group Name") or r.get("Ad Group Name (Informational only)")
    d.update({
        "campaign": str(campaign or "").strip(),
        "ad_group": str(ad_group or "").strip(),
        "entity": str(r.get("Entity") or "").strip(),
        "keyword": str(r.get("Keyword Text") or "").strip().lower(),
        "match_type": str(r.get("Match Type") or "").strip().upper(),
        # Entity == 'Bidding Adjustment' satirlari icin: mevcut placement
        # carpanini okumak icin (tahmin etmek yerine gercek degerden Update
        # yapabilmek). Diger entity turlerinde bos kalir.
        "placement": str(r.get("Placement") or "").strip(),
        "percentage": _num(r.get("Percentage")),
        "impressions": 0, "clicks": 0, "spend": 0, "sales": 0,
        "orders": 0, "cpc": 0, "acos": 0,
    })
    return d


def _norm_campaign(r):
    d = _base_metrics(r)
    d.update(_ids(r))
    d.update({
        "campaign": str(r.get("Campaign Name") or "").strip(),
        "status": str(r.get("Status") or "").strip(),
        "budget": _num(r.get("Budget Amount")),
        "targeting_type": str(r.get("Targeting Type") or "").strip(),
        "bidding_strategy": str(r.get("Bidding strategy") or "").strip(),
    })
    return d
