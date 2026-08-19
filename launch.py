"""Sifir urun (launch) PPC asistani.

Yeni listelenen bir urun icin:
  1) urunu tani (baslik/ASIN/fiyat/rakipler)  -> extension veya gorsel yukleme saglar
  2) keyword bul (rakip basliklarindan + AI genisletme)
  3) kampanya plani kur (Auto + Manual Broad/Phrase/Exact + ASIN targeting)
  4) Amazon canonical bulk sheet uret (Create operasyonlari)

Bu modul mevcut bulksheet.BULK_HEADERS formatini yeniden kullanir; cikti
Seller Central > Bulk Operations'a dogrudan yuklenebilir.
"""
import concurrent.futures
import io
import re
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill

import config
from bulksheet import BULK_HEADERS, _sanitize_name
import competitor_intel
import benchmarks

try:
    import keepa_engine
except Exception:  # pragma: no cover
    keepa_engine = None


# ------------------------------------------------------------------ keywords
_STOP = set("""
a an and or the for with of to in on at by from your you our this that these those
new best top hot sale premium quality pack set of size pcs pack piece pieces
amazon fba prime free shipping buy set kit pack x large small medium
ve ile icin bir bu su cok en daha the a
""".split())

_NOISE_RE = re.compile(r"[^a-z0-9\s\-'|]")
_BOUNDARY_RE = re.compile(r"[,;:/()\[\]{}!?.–—]|\s-\s")
# Olcu/adet token'lari keyword olamaz: 320, 500mg, 12oz, 2pack
_MEASURE_RE = re.compile(r"^\d+([a-z]{1,4})?$")

# Baslikta yoksa bosa harcamayi kesen jenerik negatifler.
# (tetikleyici token, negatif terimler)
_NEG_RULES = [
    ({"women", "woman", "womens", "female", "her"}, ["men", "mens", "man", "male", "for men"]),
    ({"men", "mens", "man", "male", "his"}, ["women", "womens", "woman", "female", "for women"]),
    ({"dog", "cat", "pet", "puppy", "kitten"}, ["human"]),
    ({"kids", "kid", "baby", "toddler", "child", "children"}, ["adult"]),
]
# Urun basliginda gecmiyorsa her zaman negatiflenecek jenerik gurultu
_NEG_ALWAYS = ["dog", "cat", "pet", "kids", "baby", "sample", "travel size",
               "wholesale", "bulk", "refill", "diy", "used", "gift card"]


def _tokens(text):
    text = _NOISE_RE.sub(" ", (text or "").lower())
    return [t for t in text.split()
            if t and t not in _STOP and len(t) > 2 and not _MEASURE_RE.match(t)]


def _segments(text):
    """Metni n-gram uretilebilir parcalara ayirir.

    Kritik: stopword'ler ve noktalama SILINMEZ, SINIR olur. Yoksa
    "shampoo for thin, fine hair" -> "shampoo thin" gibi kimsenin
    aratmadigi sahte komsuluklar uretilir.
    """
    low = (text or "").lower()
    low = _BOUNDARY_RE.sub("|", low)          # noktalama = sinir
    low = _NOISE_RE.sub(" ", low.replace("|", " | "))
    segs, cur = [], []
    for w in low.split():
        if w == "|" or w in _STOP or len(w) <= 2 or _MEASURE_RE.match(w):
            if cur:
                segs.append(cur)
                cur = []
            continue
        cur.append(w)
    if cur:
        segs.append(cur)
    return segs


def _ngrams(tokens, n):
    return [" ".join(tokens[i:i + n]) for i in range(len(tokens) - n + 1)]


def _head_tokens(title, competitors):
    """Urunun kategori 'kok' kelimeleri: kendi basliginda gecen ve rakip
    basliklarinda da siklikla gecen terimler (ornek: shampoo, hair).
    Cok kelimeli bir keyword bunlardan en az birini icermeliyse gercek bir
    arama sorgusudur; icermiyorsa icerik/ingredient artigidir."""
    own = set(_tokens(title))
    if not own:
        return set()
    def _from_title():
        """Rakip yokken: kategori nounu BASLIK ICINDE tekrar eder
        (shampoo x2, hair x3); icerik kelimeleri (malva, redensyl) bir kez
        gecer. Hicbiri tekrar etmiyorsa ilk parcanin son kelimesi head'dir."""
        tf = {}
        for seg in _segments(title):
            for t in seg:
                tf[t] = tf.get(t, 0) + 1
        rep = {t for t, c in tf.items() if c >= 2}
        if rep:
            return rep
        segs = _segments(title)
        return {segs[0][-1]} if segs and segs[0] else own

    comp_titles = [c.get("title") for c in (competitors or []) if c.get("title")]
    if not comp_titles:
        return _from_title()
    df = {}
    for ct in comp_titles:
        for t in set(_tokens(ct)):
            if t in own:
                df[t] = df.get(t, 0) + 1
    if not df:
        return _from_title()
    # Mutlak degil GORELI esik: en cok gecen terime yakin olanlar kategori
    # nounudur (hair, shampoo). Icerik/ozellik kelimeleri (biotin, volumizing)
    # rakiplerde gecse de daha seyrektir ve elenir.
    top = max(df.values())
    thresh = max(2, -(-top * 7 // 10)) if top >= 2 else 1
    heads = {t for t, c in df.items() if c >= thresh}
    if not heads:
        heads = set(sorted(df, key=df.get, reverse=True)[:3])
    return set(sorted(heads, key=df.get, reverse=True)[:5])


def heuristic_keywords(title, competitors, max_kw=40, search_suggestions=None):
    """Baslik + rakip metinlerinden (baslik, bullets, description) frekansa gore aday keyword'ler."""
    freq = {}
    sources = []
    if title:
        sources.append((title, 1.0))
        
    for c in (competitors or []):
        if c.get("title"):
            sources.append((c["title"], 1.0))
        for b in c.get("bullets", []):
            if b:
                sources.append((b, 0.8))  # bullets have lower weight than title
        if c.get("description"):
            sources.append((c["description"], 0.5))  # description lowest weight

    for src, src_weight in sources:
        for seg in _segments(src):
            for n in (1, 2, 3):
                for g in _ngrams(seg, n):
                    w = (1.0 if n == 1 else 1.6 if n == 2 else 1.4) * src_weight
                    freq[g] = freq.get(g, 0) + w


    # Arama onerileri (search_suggestions) varsa agirligini cok artir
    if search_suggestions:
        sug_lower = [s.lower() for s in search_suggestions]
        for g in freq:
            if g in sug_lower:
                freq[g] *= 2.5

    # kendi basligindaki terimlere bonus + alakasiz filtreleme
    own = set(_tokens(title))
    
    # Alakasiz keyword'leri filtrele: keyword'un en az 1 kelimesi
    # urunun kendi basligindaki terimlerle ortusmeli VEYA
    # search_suggestions'da yer almali
    sug_set = set()
    if search_suggestions:
        for s in search_suggestions:
            for t in _tokens(s):
                sug_set.add(t)
    
    # ST STRICT CATEGORY RELEVANCE CHECK
    # Root product tokens (excluding generic size/brand/stop words)
    generic_words = {'the','and','for','with','set','pack','size','new','best','top','sale','prime','men','women','kids','pack','pcs','oz','ml','gram','kg','large','small','medium','black','white','blue','red'}
    root_tokens = set(t for t in own if t not in generic_words and len(t) > 2)

    heads = _head_tokens(title, competitors)

    def _is_relevant(kw_str):
        """Keyword gercek bir arama sorgusu gibi mi?"""
        kw_words = set(kw_str.split())
        if root_tokens and not (kw_words & root_tokens):
            return False
        # Cok kelimeli terim mutlaka bir kategori kok kelimesi icermeli.
        # "hair growth" / "thickening shampoo" gecer;
        # "caffeine malva" / "redensyl biotin" elenir.
        if len(kw_words) > 1 and heads and not (kw_words & heads):
            return False
        return True

    ranked = sorted(freq.items(),
                    key=lambda kv: (kv[1] + (0.5 if set(kv[0].split()) & own else 0)),
                    reverse=True)
    out, seen = [], set()
    for kw, score in ranked:
        if kw in seen:
            continue
        if not _is_relevant(kw):
            continue
        seen.add(kw)
        out.append(kw)
        if len(out) >= max_kw:
            break
    return out


def baseline_negatives(title, extra=None):
    """AI kapaliyken bile lansmanin negatifsiz gitmemesi icin taban liste.
    Urunun kendi basliginda gecen hicbir terim negatiflenmez."""
    own = set(_tokens(title))
    own_raw = set((title or "").lower().split())
    out = []

    def _add(term):
        t = term.strip().lower()
        if not t or t in out:
            return
        if any(w in own or w in own_raw for w in t.split()):
            return          # kendi urunumuzu bloklama
        out.append(t)

    for trigger, negs in _NEG_RULES:
        if own & trigger:
            for n in negs:
                _add(n)
    for n in _NEG_ALWAYS:
        _add(n)
    for n in (extra or []):
        _add(n)
    return out


def ai_strategy(title, competitors, price=None, economics=None, model=None, product_data=None, keyword_analysis=None, competitor_intel_data=None, market_assessment=None):
    """Claude ile tam launch stratejisi: keyword'ler + negatifler + gerekce +
    2 haftalik aksiyon plani. Anahtar yoksa None."""
    if not config.ANTHROPIC_API_KEY:
        return None
    try:
        from anthropic import Anthropic
    except Exception:
        return None
        
    product_data = product_data or {}
    comp_titles = "\n".join(f"- {c.get('title','')} (Fiyat: {c.get('price', '-')})" for c in (competitors or [])[:14])
    
    intel_text = ""
    if competitor_intel_data:
        intel_text += f"\nREKABET DURUMU:\n{competitor_intel_data.get('market_summary', '')}\n"
    if market_assessment:
        intel_text += f"FIRSAT SKORU: {market_assessment.get('opportunity_score')}/100\n"
        intel_text += f"Önerilen Agresiflik: {market_assessment.get('recommended_aggression')}\n"
    if keyword_analysis and keyword_analysis.get('high_priority'):
        hp_kws = [k['keyword'] for k in keyword_analysis['high_priority'][:10]]
        intel_text += f"YÜKSEK ÖNCELİKLİ KELİMELER (Rakipler + Arama Önerilerinden):\n{', '.join(hp_kws)}\n"
        
    try:
        import expert_knowledge
        expert_context = expert_knowledge.EXPERT_KNOWLEDGE
    except ImportError:
        expert_context = ""

    econ_line = ""
    if economics:
        econ_line = (f"\nEKONOMI: birim kar ${economics.get('unit_profit_before_ads')}, "
                     f"break-even ACOS %{economics.get('break_even_acos_pct')}, "
                     f"onerilen hedef ACOS %{economics.get('recommended_target_acos_pct')}. "
                     f"Bid onerilerini bu marja gore mantikli tut.")
                     
    prompt = f"""Sen Amazon PPC uzmanisin. Yeni listelenen ("sifir") bir urun icin
Sponsored Products LAUNCH stratejisi kur.

UZMANLIK BİLGİSİ (Referans Al):
{expert_context[:2000] if expert_context else '(yok)'}

URUN: {title}
FIYAT: {price or 'bilinmiyor'}{econ_line}
{intel_text}
RAKIPLER:
{comp_titles or '(yok)'}

Sadece gecerli JSON dondur, baska hicbir sey yazma:
{{
  "exact": ["5-10 en yuksek donusum, spesifik alici-niyeti terim"],
  "phrase": ["8-15 orta genislik terim"],
  "broad": ["10-20 kesif terim"],
  "brand_defense": ["varsa marka/rakip terimleri, yoksa []"],
  "negatives": ["auto/broad'da bosa para yakacak alakasiz terimler (exact negatif)"],
  "rationale": "2-3 cumle Turkce: bu urun icin launch stratejisinin OZU",
  "expert_reasoning": {{
    "keyword_strategy": "Kelimeler neden seçildi...",
    "bid_strategy": "Bid seviyeleri neden bu şekilde...",
    "campaign_structure": "Kampanya yapısı neden böyle...",
    "risk_assessment": "Potansiyel riskler...",
    "competitive_edge": "Rekabet avantajımız..."
  }},
  "action_plan": [
    "Hafta 1: ...",
    "Hafta 2: ..."
  ],
  "launch_phases": {{
    "week_1": {{"focus": "...", "actions": ["..."], "budget_pct": 30}},
    "week_2": {{"focus": "...", "actions": ["..."], "budget_pct": 25}},
    "week_3_4": {{"focus": "...", "actions": ["..."], "budget_pct": 25}},
    "week_5_plus": {{"focus": "...", "actions": ["..."], "budget_pct": 20}}
  }}
}}
KRİTİK UYARI: Yalnızca doğrudan bu ürünle ALAKALI terimleri kullan. Örneğin ürün Şampuan (shampoo) ise deodorant, body wash, soap, lotion gibi FARKLI KATEGORİ kelimelerini KESİNLİKLE ele ve ekleme. Ürün başlığı ({title}) temel kategoridir.
Keyword'leri urun dilinde (genelde Ingilizce) yaz. Alakasiz cok-genel tek kelimeleri ve baska kategori kelimelerini ele."""
    import json
    client = Anthropic(api_key=config.ANTHROPIC_API_KEY)
    last_err = None
    for attempt in range(2):
        try:
            resp = client.messages.create(
                # LAUNCH_MODEL dusunen bir model (opus-5); 20sn'lik eski
                # timeout ve 2000 token her denemede yetmiyordu, AI bu yuzden
                # kredi olsa bile hicbir zaman devreye girmiyordu.
                model=model or config.LAUNCH_MODEL,
                max_tokens=config.MAX_LAUNCH_TOKENS,
                timeout=config.LAUNCH_AI_TIMEOUT,
                system="Sadece istenen JSON'u dondur. Aciklama, markdown fence veya ek metin yazma.",
                messages=[{"role": "user", "content": prompt}],
            )
            text = "".join(b.text for b in resp.content if getattr(b, "type", "") == "text")
            text = text.strip().removeprefix("```json").removeprefix("```").removesuffix("```").strip()
            m = re.search(r"\{.*\}", text, re.S)
            if not m:
                last_err = f"JSON bulunamadi (yanit: {text[:80]!r})"
                continue
            raw = m.group(0)
            try:
                data = json.loads(raw)
            except Exception:
                import json_repair
                data = json_repair.loads(raw)
            clean = {}
            for k in ("exact", "phrase", "broad", "brand_defense", "negatives"):
                vals = data.get(k) or []
                clean[k] = [str(v).strip().lower() for v in vals if str(v).strip()]
            clean["rationale"] = str(data.get("rationale") or "").strip()
            clean["action_plan"] = [str(x).strip() for x in (data.get("action_plan") or []) if str(x).strip()]
            clean["expert_reasoning"] = data.get("expert_reasoning") or {}
            clean["launch_phases"] = data.get("launch_phases") or {}
            if clean["exact"] or clean["broad"]:
                return clean
            last_err = "bos keyword listesi"
        except Exception as e:
            last_err = str(e)
    print(f"ai_strategy basarisiz (2 deneme): {last_err}")
    globals()["_LAST_AI_ERROR"] = str(last_err)
    return None


def break_even(price, cogs=0, fee_pct=0.15, fba_fee=0):
    """Basit break-even ACOS + onerilen hedef ACOS."""
    p = float(price or 0)
    if p <= 0:
        return None
    unit_profit = p - float(cogs or 0) - p * float(fee_pct or 0.15) - float(fba_fee or 0)
    be = unit_profit / p if p > 0 else 0
    return {
        "unit_profit_before_ads": round(unit_profit, 2),
        "break_even_acos_pct": round(be * 100, 1),
        "recommended_target_acos_pct": round(be * 0.7 * 100, 1),
        "margin_pct_before_ads": round(be * 100, 1),
    }


# ------------------------------------------------------------------ bids/budget



def discovery_plan(price, econ, bench, keywords=None, market_price=None):
    """FAZ 0 - FIYAT KESFI: gercek CPC'yi 1-2 gunde ve ucuza satin alir.

    Neden bu ise yarar (uc gercege dayanir):

    1) Amazon SP IKINCI-FIYAT acik artirmasidir. Teklifin $5 olsa da pazar
       $2.60'ta temizleniyorsa ~$2.60 odersin. Yuksek teklif "cok para
       odemek" demek DEGILDIR; sadece acik artirmaya girmeyi garanti eder.
       Riski BUTCE sinirlar, teklif degil.

    2) CPC olcmek UCUZDUR, CVR olcmek PAHALI. Her tiklama bir CPC olcumudur
       (~20 tik -> +-%9 dogruluk). CVR ise nadir olaydir; %30 dogruluk icin
       ~100 tiklama gerekir. Bu yuzden once SADECE CPC satin alinir.

    3) Bilinmeyen CPC, plandaki en buyuk hata kaynagidir: bid'i de butceyi de
       tahmini ACOS'u da o belirler. Onu 1-2 gunde kapatmak, iki hafta yanlis
       varsayimla harcamaktan cok daha ucuzdur.

    Sonuc: yuksek teklif + kucuk butce = birkac saatte gercek CPC.
    """
    p = float(price or 0)
    if p <= 0:
        return {}
    be = float((econ or {}).get("break_even_acos_pct") or 40.0) / 100.0
    rev = float((bench.get("account") or {}).get("aov") or 0) or p
    cvr = bench["cvr"].get("phrase") or bench["cvr"].get("exact") or 0.06

    # KESIF TEKLIFI - odenebilirlikten TURETILMEZ.
    # Onceki surumde "odenebilir bid x 3" kullaniliyordu; bu mantik hatasiydi.
    # Odenebilir bid dusukse (dusuk CVR varsayimi yuzunden) kesif teklifi de
    # dusuk cikiyor, acik artirma kazanilamiyor ve HICBIR SEY olculemiyordu -
    # yani kesif fazi kendi amacini bosa cikariyordu.
    #
    # Dogrusu: teklif "kazanacak kadar yuksek" olmali. Ikinci-fiyat oldugu
    # icin bu tutari genelde ODEMEZSIN. Riski butce sinirlar.
    # Fiyatin ~%12'si, cogu kategoride pazar CPC'sinin belirgin ustunde kalir.
    afford = rev * cvr * be
    probe = round(max(3.00, min(9.00, p * 0.12)), 2)

    # CPC icin ~20 tiklama yeter (+-%9). Butce en kotu durumu sinirlar:
    # gercekten teklif kadar odersen bile gunluk kayip ~bu kadardir.
    # (Amazon gunluk butceyi %25'e kadar asabilir; ustte pay birakildi.)
    target_clicks = 20
    days = 3          # kampanya baslamasi + rapor gecikmesi icin pay
    per_day = int(max(20, min(80, round(probe * target_clicks / days))))

    return {
        "phase": "0 - Fiyat Keşfi",
        "purpose": "Gerçek CPC'yi ölçmek. Tek amaç bu; satış beklentisi yok.",
        "days": days,
        "probe_bid": probe,
        "affordable_bid": round(afford, 2),
        "budget_per_day": per_day,
        # Dosyada exact + auto var; auto butcesi exact'in yarisi.
# Amazon gunluk butceyi %25'e kadar asabilir -> 1.25 pay.
        "max_total_spend": round(per_day * 1.5 * 1.25 * days),
        "safety_net": (
            "Dosyada EXACT'in yanına küçük bütçeli bir AUTO kampanya da var. "
            "Yeni ASIN'de exact keyword'ler gösterim almayabilir (Amazon'un o "
            "kelimeler için alaka geçmişi yoktur); auto'da yerleşimi Amazon "
            "seçer ve ilk gösterimi almanın en güvenilir yoludur. Exact hiç "
            "veri vermezse bile CPC auto'dan ölçülür."),
        "target_clicks": target_clicks,
        "keywords": list(keywords or [])[:10],
        "match_type": "EXACT",
        "why_high_bid": (
            f"Amazon ikinci-fiyat açık artırması kullanır: ${probe} teklif etsen de "
            f"pazar neyde temizleniyorsa onu ödersin. Bu teklif kârlılıktan "
            f"türetilmedi — kârlı bid ${afford:.2f} olurdu ve o tutarla açık "
            f"artırmayı kazanamayıp hiçbir şey ölçemezdin. Amaç kazanmak; "
            f"riski günlük ${per_day} bütçe sınırlar."),
        "why_exact_only": (
            "Sadece EXACT match kullanılır. Yüksek teklif + broad match, "
            "alakasız pahalı trafik çeker. Exact'te ne için ödediğini bilirsin."),
        "budget_warning": (
            "Amazon günlük bütçeyi %25'e kadar aşabilir (ay içinde dengeler). "
            f"Gerçek üst sınırı ~${round(per_day*1.25)}/gün kabul et."),
        "why_cheap": (
            "CPC ölçmek ucuzdur: her tıklama bir ölçümdür, ~20 tıklama ±%9 "
            "doğruluk verir. CVR ölçmek pahalıdır (~100 tıklama), o yüzden "
            "bu fazda CVR'yi ölçmeye çalışmıyoruz."),
        "success_criteria": (
            f"{target_clicks} tıklamaya ulaşmak. Sipariş gelmemesi BAŞARISIZLIK "
            f"DEĞİLDİR — bu fazın amacı satış değil, fiyat bilgisi."),
        "stop_rule": f"{target_clicks} tıklama veya {days} gün — hangisi önce gelirse.",
        # "Veri gelmezse ne yapacagim?" sorusunun cevabi plana gomulu olmali.
        # Her gun icin tek bir kontrol ve tek bir eylem.
        "escalation": [
            {"when": "1. gün sonunda gösterim = 0",
             "diagnosis": "Bu bir teklif sorunu DEĞİL — reklam hiç yayınlanmıyor.",
             "action": "Buy Box, stok, kampanya durumu ve reddedilen keyword'leri "
                       "kontrol et. Teklifi artırmak bu durumu çözmez."},
            {"when": "2. gün sonunda gösterim var ama tıklama < 5",
             "diagnosis": "Yayındasın ama ya çok az gösterim alıyorsun ya da "
                          "listing tıklama almıyor.",
             "action": f"Gösterim düşükse teklifi ${round(probe*1.4,2)}'a çıkar "
                       f"(bütçeyi değil). Gösterim yüksek ama tıklama yoksa "
                       f"sorun görsel/fiyat/yorum — teklif artırma."},
            {"when": "3. gün sonunda tıklama 6-14 arası",
             "diagnosis": "Ölçüm zayıf ama kullanılabilir (±%11-16).",
             "action": "Raporu yükle; araç bu veriyi 'ZAYIF' etiketiyle kullanır. "
                       "Faz 1'e temkinli (Kârlı) strateji ile başla."},
            {"when": "Bütçe her gün erken bitiyor",
             "diagnosis": "Teklif yeterli, hacim var — sadece bütçe kısıtlıyor.",
             "action": "Bütçeyi artır, teklifi değil. Bu iyi haber: pazar seni "
                       "kabul ediyor."},
        ],
        "no_data_promise": (
            "Hiç veri gelmezse araç uydurmaz: 'ölçüm yok' der ve yukarıdaki "
            "teşhis adımlarını gösterir. Varsayımla lansman kurmaya zorlamaz."),
        "next_step": (
            "Seller Central > Reports > Targeting raporunu indir, uygulamada "
            "bu markaya yükle. Araç gerçek CPC'yi okuyup Faz 1 planını "
            "varsayımsız hesaplar."),
    }


def bid_feasibility_v2(price, econ, profit_bids, bench):
    """Karli bid pazarin gercek CPC'sini karsiliyor mu?

    Karsilamiyorsa bu bir bid ayari sorunu degil, fiyat/maliyet sorunudur.
    Karsilastirma artik tahmini degil OLCULMUS CPC ile yapilir.
    """
    if not profit_bids:
        return {}
    market = bench["cpc"].get("exact") or 0
    afford = profit_bids.get("exact") or 0

    if not market:
        # CPC olculmemis. "Karsiliyor" demek yanlis olur - kiyaslanacak sayi yok.
        return {
            "status": "unknown",
            "headline": "Pazar CPC'si bilinmiyor — karşılaştırma yapılamıyor.",
            "market_cpc_estimate": None,
            "affordable_bid": round(afford, 2),
            "ratio_pct": None,
            "break_even_acos_pct": (econ or {}).get("break_even_acos_pct"),
            "advice": ["Aşağıdaki Ölç-Düzelt planıyla 3 günde gerçek CPC'yi ölç.",
                       "Ya da bildiğin bir CPC varsa elle gir."],
            "note": bench["cpc_source"],
        }

    ratio = (afford / market) if market > 0 else 1.0

    if ratio >= 0.85:
        status = "ok"
        head = "Kârlı bid pazarın gerçek CPC'sini karşılıyor."
        advice = []
    elif ratio >= 0.55:
        status = "tight"
        head = "Kârlı bid pazar CPC'sinin altında — gösterim almakta zorlanabilirsin."
        advice = ["Uzun kuyruk kelimelerle başla; oralarda CPC daha düşük.",
                  "Listing dönüşümünü yükselt: her CVR puanı ödenebilir bid'i büyütür.",
                  "Dengeli stratejiyle kontrollü zarara razı olabilirsin."]
    else:
        status = "blocked"
        head = ("Bu fiyat/maliyet yapısıyla pazarın CPC'sine kârlı giremezsin.")
        advice = ["Satış fiyatını yükselt ya da COGS/FBA maliyetini düşür.",
                  "Daha dar, daha ucuz nişlere odaklan (long tail + ASIN targeting).",
                  "Bilinçli pazar payı yatırımı yapacaksan zarar yazacağını kabul et."]

    return {
        "status": status, "headline": head,
        "market_cpc_estimate": round(market, 2),
        "affordable_bid": round(afford, 2),
        "ratio_pct": round(ratio * 100),
        "break_even_acos_pct": (econ or {}).get("break_even_acos_pct"),
        "advice": advice,
        "note": bench["cpc_source"],
    }


# Bid stratejileri: OLCULMUS pazar CPC'si ile OLCULMUS ekonomi arasinda
# nerede duracagini secersin.
#   profit     -> ekonomiye sadik kal (hedef ACOS'u tuttur, break-even'i asma)
#   balanced   -> arada dur; kontrollu zarar, daha fazla gosterim
#   aggressive -> pazari karsila; bilincli zarar, maksimum hacim ve organik ivme
# market_weight = pazar CPC'sine ne kadar yaklasilacagi (0 = hic, 1 = tamamen)
BID_STRATEGIES = {
    "profit":     {"label": "Karli", "market_weight": 0.00},
    "balanced":   {"label": "Dengeli", "market_weight": 0.55},
    "aggressive": {"label": "Pazar Payi", "market_weight": 1.00},
}


def suggest_bids_v2(price, econ, bench, strategy="profit"):
    """Bid'i OLCULMUS referanslardan hesaplar (bench = benchmarks.resolve()).

        odenebilir_bid = fiyat x beklenen_CVR x hedef_ACOS
        pazar_bid      = o match type'in olculmus CPC'si
        secilen        = strateji agirligina gore ikisi arasinda

    Onceki surumde CVR ve pazar CPC'si uydurma sabitlerdi; ikisi de
    olculmus veriyle degistirildi.
    """
    p = float(price or 0)
    if p <= 0:
        return {k: 0.30 for k in bench["cvr"]}
    # ACOS ciro tabanlidir. Coklu adet siparisler yuzunden siparis basina ciro
    # (AOV) birim fiyattan yuksek olabilir; olculmusse onu kullan, yoksa fiyat.
    rev = float((bench.get("account") or {}).get("aov") or 0) or p

    be = float((econ or {}).get("break_even_acos_pct") or 40.0) / 100.0
    target = float((econ or {}).get("recommended_target_acos_pct") or be * 70) / 100.0
    w = BID_STRATEGIES.get(strategy, BID_STRATEGIES["profit"])["market_weight"]

    out = {}
    for key, cvr in bench["cvr"].items():
        if cvr is None:
            # CVR olculmemis ve varsayilmamis: bid hesaplanamaz. Uydurma sayi
            # uretmektense bos birak; cagiran taraf kullaniciya sorar.
            out[key] = None
            continue
        afford = rev * cvr * target        # hedef ACOS'u tutturan bid
        ceiling = rev * cvr * be           # break-even bid (zarar siniri)
        market = bench["cpc"].get(key)

        if strategy == "profit" or market is None:
            # CPC olculmemisse pazara capa atilamaz; ekonomiye sadik kal.
            bid = min(afford, ceiling)
        else:
            bid = afford + (market - afford) * w
            if w >= 0.9:
                bid = max(bid, market)     # pazar payi: pazari karsila
        out[key] = round(max(0.15, bid), 2)
    return out


def bid_outlook_v2(price, econ, bids, bench):
    """Secilen bid'lerle gercekte ne olur: tahmini ACOS, gosterim sansi, karlilik.
    Tum sayilar bench'teki OLCULMUS CVR/CPC uzerinden hesaplanir."""
    p = float(price or 0)
    if p <= 0 or not bids:
        return {}
    be = float((econ or {}).get("break_even_acos_pct") or 0) / 100.0
    rev = float((bench.get("account") or {}).get("aov") or 0) or p
    rows = {}
    for key, bid in bids.items():
        if bid is None:
            rows[key] = {"bid": None, "note": "CVR verisi yok - hesaplanamadi"}
            continue
        cvr = bench["cvr"].get(key) or 0
        market = bench["cpc"].get(key) or 0
        acos = bid / (rev * cvr) if (rev and cvr) else 0
        ratio = bid / market if market else 1
        rows[key] = {
            "bid": bid,
            "expected_cvr_pct": round(cvr * 100, 2),
            "market_cpc": round(market, 2),
            "expected_acos_pct": round(acos * 100, 1),
            "vs_market_pct": round(ratio * 100) if market else None,
            "impression_odds": (None if not market else
                                "iyi" if ratio >= 0.95 else
                                "orta" if ratio >= 0.75 else "dusuk"),
            "profitable": (acos <= be) if be else None,
            "cvr_source": bench["cvr_source"].get(key),
        }
    return {
        "per_match": rows,
        "break_even_acos_pct": round(be * 100, 1),
        "revenue_per_order": round(rev, 2),
        "cpc_source": bench["cpc_source"],
        "ramp_pct": round(bench["ramp"] * 100),
        "calibration": bench["calibration_note"],
        "account": bench.get("account"),
    }


# Amazon saglik/tibbi iddia iceren terimleri reddedebilir ya da reklami
# durdurabilir. Kelime dosyaya girmeden once kullaniciya soylenir; boylece
# yukleme raporundan hata ayiklamak zorunda kalmaz.
_RISKLI_TERIMLER = {
    # tedavi/tibbi iddia
    "cure", "treat", "treatment", "heal", "therapy", "therapeutic",
    "clinical", "clinically", "medical", "medicine", "prescription",
    "doctor", "dermatologist", "fda", "approved", "proven",
    # hastalik adlari
    "alopecia", "psoriasis", "eczema", "dandruff-cure", "baldness",
    "hair-loss-cure", "regrow", "regrowth", "restore",
    # abartili vaat
    "guaranteed", "miracle", "instant", "permanent", "100%",
    "best", "#1", "number one",
}


def keyword_risks(keywords):
    """Reddedilme riski tasiyan kelimeleri isaretler.

    Doner: [{keyword, terms, severity}] - severity: yuksek | orta
    Kelime SILINMEZ; karar kullanicinindir. Amac surprizi onlemek.
    """
    out = []
    for kw in keywords or []:
        k = str(kw).lower()
        kelimeler = set(re.sub(r"[^a-z0-9%#]+", " ", k).split())
        bulunan = sorted(kelimeler & _RISKLI_TERIMLER)
        # "hair loss" gibi ifadeler sorun degil; "cure"/"clinically" gibi
        # iddialar sorundur. Iki ve uzeri isaret = yuksek risk.
        if not bulunan:
            continue
        out.append({
            "keyword": kw,
            "terms": bulunan,
            "severity": "yuksek" if len(bulunan) > 1 or bulunan[0] in {
                "cure", "clinically", "clinical", "fda", "guaranteed",
                "miracle", "permanent", "prescription", "medical"} else "orta",
        })
    return out


def bid_explanation(price, econ, bids, bench):
    """Bid'lerin nereden geldigini acikla - kara kutu olmasin."""
    if not econ:
        return {}
    return {
        "formula": "max_cpc = siparis_basina_ciro x beklenen_CVR x hedef_ACOS",
        "price": round(float(price or 0), 2),
        "revenue_per_order": round(
            float((bench.get("account") or {}).get("aov") or 0) or float(price or 0), 2),
        "target_acos_pct": econ.get("recommended_target_acos_pct"),
        "break_even_acos_pct": econ.get("break_even_acos_pct"),
        "cvr_used_pct": {k: (round(v * 100, 2) if v is not None else None)
                         for k, v in bench["cvr"].items()},
        "cvr_source": bench.get("cvr_basis"),
        "cpc_source": bench.get("cpc_source"),
        "launch_ramp_pct": round(bench.get("ramp", 0) * 100),
        "bids": bids,
    }


def keyword_signals(keywords, search_suggestions=None, competitors=None,
                    head_tokens=None):
    """Her kelime icin elimizdeki GERCEK sinyalleri toplar.

    Onceden bid ayrimi sadece kelime sayisina bakiyordu; oysa veri var:
      - autocomplete sirasi: Amazon en cok aranani basa koyar (talep vekili)
      - rakip basliklarinda gecme sikligi: ticari olarak kanitlanmis terim
      - urunun ana token'lariyla ortusme: alaka
    """
    sugg = [str(s).lower() for s in (search_suggestions or [])]
    sugg_rank = {s: i for i, s in enumerate(sugg)}
    comp_titles = " || ".join(str(c.get("title") or "").lower()
                              for c in (competitors or []))
    head = set(head_tokens or [])

    out = {}
    for kw in keywords or []:
        k = str(kw).lower().strip()
        words = k.split()
        # 1) Talep: autocomplete'te var mi, kacinci sirada
        rank = sugg_rank.get(k)
        if rank is None and len(words) >= 2:
            # Tam eslesme yoksa oneriyi ICEREN cok kelimeli obek de sayilir.
            # NOT: tek kelime icin bunu yapmak yanlisti - "loss" kelimesi
            # "hair loss shampoo" onerisine takilip en yuksek talep skorunu
            # aliyordu. Tek kelimelik parca, aranan terim degildir.
            for s, i in sugg_rank.items():
                if f" {k} " in f" {s} ":
                    rank = i
                    break
        demand = 0.0 if rank is None else max(0.0, 1.0 - rank / 20.0)

        # 2) Ticari kanit: rakip basliklarinda kac kez geciyor
        comp_hits = comp_titles.count(k) if k else 0

        # 3) Alaka: ana token ortusmesi
        overlap = sum(1 for w in words if w in head)

        out[kw] = {
            "words": len(words),
            "demand": round(demand, 3),
            "in_suggestions": rank is not None,
            "suggestion_rank": rank,
            "competitor_hits": comp_hits,
            "head_overlap": overlap,
        }
    return out


def keyword_bid(kw, base_bid, match_key, head_tokens=None, signals=None):
    """Kelime bazli bid: kampanyanin taban bid'ini kelimenin niteligine gore
    yukari/asagi oynatir. Tum kelimelere ayni bid vermek, talebi ve alakayi
    yok saymak demektir.

    Carpanlar (birlestirilir, sonuc 0.6x - 1.6x arasinda kirpilir):
      talep (autocomplete)      : 1.00 -> 1.30
      rakip basliklarinda gecme : 1.00 -> 1.15
      ana token ortusmesi       : 0.80 -> 1.15
      uzunluk                   : long tail 0.90, tek kelime 0.85
    """
    s = (signals or {}).get(kw) or {}
    words = s.get("words") or len(str(kw).split())
    f = 1.0

    # Talep: cok aranan kelime daha degerlidir, daha yuksek teklif hak eder.
    f *= 1.0 + 0.30 * float(s.get("demand") or 0.0)

    # Rakipler bu terimi basliklarinda kullaniyorsa ticari olarak kanitlidir.
    hits = int(s.get("competitor_hits") or 0)
    if hits >= 2:
        f *= 1.15
    elif hits == 1:
        f *= 1.07

    # Alaka: urunun kok kelimeleriyle ortusme
    ov = s.get("head_overlap")
    if ov is None and head_tokens:
        ov = sum(1 for w in str(kw).lower().split() if w in head_tokens)
    if ov is not None:
        if ov >= 2:
            f *= 1.15
        elif ov == 0:
            f *= 0.80

    # Uzunluk: long tail ucuz alinir, tek kelime genis ve riskli
    if words >= 4:
        f *= 0.90
    elif words == 1:
        f *= 0.85

    f = max(0.60, min(1.60, f))
    return round(max(0.15, base_bid * f), 2)


def suggest_budgets(price, bids=None, clicks_per_day=None):
    """Butce = hedeflenen gunluk tiklama x o kampanyanin bid'i.

    Sabit $15 yerine bid'den turetilir; bid dusukse butce de dusuk olur ve
    para bosa beklemede kalmaz.
    """
    if not bids:
        p = float(price or 0)
        lo = 10 if p < 25 else 15
        return {"auto": lo, "broad": lo, "phrase": lo, "exact": lo + 5, "pt": lo}

    # Kampanya basina gunluk hedef tiklama: ogrenme icin yeterli veri toplayacak
    # ama tek gunde butceyi yakmayacak seviye.
    target_clicks = clicks_per_day or {
        "auto": 12, "broad": 12, "phrase": 10, "exact": 14, "pt": 8,
    }
    out = {}
    for key, clicks in target_clicks.items():
        b = bids.get(key)
        if b is None:
            continue
        raw = b * clicks
        # Amazon minimumu $1; cok kucuk butce gun icinde hic gosterim almaz.
        out[key] = max(5, int(round(raw)))
    return out


def _campaign_prefix(brand, title, asin):
    """Kampanya adi onekini uretir: 'Marka | Urun (ASIN)'.

    Sadece marka yazmak iki sorun cikariyordu:
      1) Ayni markanin ikinci urununu lansmanlarken kampanya adlari BIREBIR
         ayni oluyordu - Amazon ayni isimde ikinci kampanyayi reddeder.
      2) Panelde hangi kampanyanin hangi urune ait oldugu anlasilmiyordu.
    """
    b = _sanitize_name(brand or "") or ""
    t = _sanitize_name(title or "") or ""

    # Basliktan marka kelimelerini dus; kalan kisim urunu tanimlar.
    # Noktalama normalize edilmeli: marka "Natural.clinic", baslikta
    # "Natural Clinic" olarak geciyor - duz karsilastirma bunu kaciriyordu.
    def _norm(s):
        return re.sub(r"[^a-z0-9]+", " ", str(s).lower()).split()

    btoks = set(_norm(b))
    words, seen = [], set()
    for w in t.split():
        lw = re.sub(r"[^a-z0-9]+", "", w.lower())
        if not lw or lw in btoks or lw in seen:
            continue
        # Olcu/adet artiklarini atla (500ml, 2pack, 16oz)
        if re.fullmatch(r"\d+([a-z]{1,4})?", lw):
            continue
        seen.add(lw)
        words.append(w)
        if len(words) >= 3:          # 3 kelime urunu ayirt etmeye yeter
            break

    prod = " ".join(words)[:32].strip()
    parts = [p for p in (b[:24], prod) if p]
    label = " | ".join(parts) if parts else "Launch"
    # ASIN benzersizligi garanti eder (ayni urunun varyantlari icin sart).
    if asin:
        label = f"{label} ({asin})"
    return _sanitize_name(label)[:90] or "Launch"


# ------------------------------------------------------------------ plan
def build_plan(product, competitors=None, use_ai=True, model=None,
               bid_strategy="profit", measured_cpc=None, report_rows=None,
               ba_rows=None, brand_id=None, brand_name=None, assumed_cvr=None):
    """product: {title, asin, sku, price, brand, cogs, fba_fee, fee_pct}. -> plan dict."""
    title = product.get("title") or ""
    asin = (product.get("asin") or "").strip()
    sku = (product.get("sku") or "").strip() or asin  # SKU yoksa ASIN'i placeholder yap
    # Autocomplete/infer missing financial values
    price = product.get("price")
    try:
        price = float(price) if price is not None else 0
    except (ValueError, TypeError):
        price = 0

    competitors = competitors or []
    search_suggestions = product.get("search_suggestions") or []
    if price <= 0:
        comp_prices = [float(c.get("price") or 0) for c in competitors if c.get("price") and float(c.get("price") or 0) > 0]
        if comp_prices:
            price = round(sum(comp_prices) / len(comp_prices), 2)
        else:
            price = 29.99
        product["price"] = price

    if not product.get("cogs"):
        product["cogs"] = round(price * 0.25, 2)

    if not product.get("fba_fee"):
        product["fba_fee"] = round(max(3.50, price * 0.18 + 1.50), 2)

    if not product.get("fee_pct"):
        product["fee_pct"] = 0.15

    econ = break_even(price, product.get("cogs"),
                      product.get("fee_pct", 0.15), product.get("fba_fee"))

    # Competitor intelligence computations
    intel_data = competitor_intel.analyze_competitors(competitors, product)
    kw_analysis = competitor_intel.reverse_engineer_keywords(title, competitors, search_suggestions)
    market_assess = competitor_intel.assess_market_opportunity(product, competitors, product.get("bsr"))

    with concurrent.futures.ThreadPoolExecutor(max_workers=2) as executor:
        heur_future = executor.submit(heuristic_keywords, title, competitors, search_suggestions=search_suggestions)
        ai_future = executor.submit(
            ai_strategy, title, competitors, price, econ, 
            model=model, product_data=product,
            keyword_analysis=kw_analysis,
            competitor_intel_data=intel_data,
            market_assessment=market_assess
        ) if use_ai else None
        
        heur = heur_future.result()
        ai = ai_future.result() if ai_future else None

    negatives, rationale, action_plan, expert_reasoning, launch_phases = [], "", [], {}, {}
    # Exact'a tek kelimelik jenerik head term koyma: lansmanda en pahali,
    # en alakasiz trafigi getirir. Onlar broad'da arastirilir.
    heur_exact = [k for k in heur if len(k.split()) > 1]
    if ai:
        exact = ai.get("exact") or heur_exact[:6]
        phrase = ai.get("phrase") or heur[:12]
        broad = ai.get("broad") or heur[:20]
        negatives = ai.get("negatives") or []
        rationale = ai.get("rationale") or ""
        action_plan = ai.get("action_plan") or []
        expert_reasoning = ai.get("expert_reasoning") or {}
        launch_phases = ai.get("launch_phases") or {}
        source = "ai"
    else:
        exact = heur_exact[:6]
        phrase = heur[:12]
        broad = heur[:20]
        source = "heuristic"

    # AI de tek kelimelik head term onerebiliyor; ayni kural her kaynaga uygulanir.
    exact = [k for k in (exact or []) if k and len(str(k).split()) > 1]
    if not exact:
        exact = heur_exact[:6]

    # Tek kelimelik parcalar ("loss", "natural", "clinic") gercek arama terimi
    # degildir; para yakar. Sadece kategori nounu (shampoo gibi) kesif amacli
    # broad'da kalir. Kendi marka kelimeleri de elenir: yeni markanin adi
    # aranmiyor, o trafige odeme yapmak anlamsiz.
    head_tok = set(_head_tokens(title, competitors) or [])
    _brand_toks = {w for w in re.sub(r"[^a-z0-9]+", " ",
                                     str(product.get("brand") or "").lower()).split() if w}
    _head_sorted = list(head_tok)

    def _clean_terms(terms, allow_head_noun=True):
        out, head_used = [], False
        for t in terms or []:
            k = str(t).strip()
            if not k:
                continue
            ws = k.lower().split()
            if all(w in _brand_toks for w in ws):
                continue                      # tamamen kendi markasi
            if len(ws) == 1:
                if not allow_head_noun or head_used or ws[0] not in _head_sorted:
                    continue                  # tek kelimelik parca
                head_used = True              # kategori nounundan sadece bir tane
            if k not in out:
                out.append(k)
        return out

    phrase = _clean_terms(phrase, allow_head_noun=False)
    broad = _clean_terms(broad, allow_head_noun=True)
    exact = _clean_terms(exact, allow_head_noun=False)

    # AI negatif vermediyse (veya kapaliysa) taban negatifleri uygula
    negatives = baseline_negatives(title, negatives)

    # Traffic sculpting: exact'teki terimler broad/phrase'de negativeExact
    # olur; ayni terim icin kampanyalar birbiriyle yarismaz.
    exact_set = [k for k in exact if k]
    sculpt = list(dict.fromkeys(negatives + exact_set))

    # ASIN targeting: rakip ASIN'ler + Keepa "birlikte alinan"
    comp_asins = [c.get("asin") for c in competitors if c.get("asin")]
    if keepa_engine and asin and config.KEEPA_API_KEY:
        try:
            for rec in keepa_engine.search_related_asins(asin):
                a = rec.get("keyword")
                if a and a not in comp_asins:
                    comp_asins.append(a)
        except Exception:
            pass
    comp_asins = list(dict.fromkeys(comp_asins))[:20]

    if bid_strategy not in BID_STRATEGIES:
        bid_strategy = "profit"
    # Referanslar: markanin kendi olculmus verisi > kalibre varsayilan.
    # MARKA IZOLASYONU: report_rows/ba_rows cagiran tarafta brand_id ile
    # filtrelenmis olarak gelir. Baska markanin verisi buraya giremez.
    bench = benchmarks.resolve(
        rows=report_rows, ba_rows=ba_rows,
        brand_id=brand_id, brand_name=brand_name or product.get("brand"),
        category_tokens=head_tok, override_cpc=measured_cpc,
        assumed_cvr=assumed_cvr, asin=asin)
    bids = suggest_bids_v2(price, econ, bench,
                           strategy=(bid_strategy if bench.get("has_cpc") else "profit"))
    budgets = suggest_budgets(price, bids)
    outlook = bid_outlook_v2(price, econ, bids, bench)
    # Uyari her zaman "karli" bid'e gore hesaplanir: strateji degistirmek
    # fiyat/maliyet gercegini degistirmez, sadece ne kadar zarara razi
    # oldugunu degistirir.
    # SECILEN vs GERCEKLESEN strateji: CPC olcumu yoksa pazara capa atilamaz,
    # balanced/aggressive zorunlu olarak profit gibi davranir. Bunu sessizce
    # yapmak yaniltir - kullanici 'Dengeli' sectigini sanip Karli bid alir.
    effective_strategy = bid_strategy if bench.get("has_cpc") else "profit"
    if effective_strategy != bid_strategy:
        bench.setdefault("warnings", []).append(
            f"'{BID_STRATEGIES[bid_strategy]['label']}' secildi ama pazar CPC'si "
            f"olculmedigi icin uygulanamadi; bid'ler '{BID_STRATEGIES['profit']['label']}' "
            f"gibi hesaplandi. Pazara capa atmak icin once Faz 0 ile gercek "
            f"CPC'yi olc.")
    profit_bids = suggest_bids_v2(price, econ, bench, strategy="profit")
    feasibility = bid_feasibility_v2(price, econ, profit_bids, bench)
    prefix = _campaign_prefix(product.get("brand"), title, asin)

    campaigns = [
        {"key": "auto", "name": f"{prefix} | Auto | Discovery",
         "targeting_type": "Auto", "budget": budgets["auto"],
         "default_bid": bids["auto"], "match": None, "keywords": [],
         "auto_groups": True, "product_targets": [], "negatives": sculpt},
        {"key": "broad", "name": f"{prefix} | Manual | Broad Research",
         "targeting_type": "Manual", "budget": budgets["broad"],
         "default_bid": bids["broad"], "match": "broad", "keywords": broad,
         "auto_groups": False, "product_targets": [], "negatives": sculpt},
        {"key": "phrase", "name": f"{prefix} | Manual | Phrase",
         "targeting_type": "Manual", "budget": budgets["phrase"],
         "default_bid": bids["phrase"], "match": "phrase", "keywords": phrase,
         "auto_groups": False, "product_targets": [], "negatives": sculpt},
        {"key": "exact", "name": f"{prefix} | Manual | Exact (Scale)",
         "targeting_type": "Manual", "budget": budgets["exact"],
         "default_bid": bids["exact"], "match": "exact", "keywords": exact,
         "auto_groups": False, "product_targets": [], "negatives": negatives},
    ]
    if comp_asins:
        campaigns.append(
            {"key": "pt", "name": f"{prefix} | Manual | ASIN Targeting",
             "targeting_type": "Manual", "budget": budgets["pt"],
             "default_bid": bids["pt"], "match": None, "keywords": [],
             "auto_groups": False, "product_targets": comp_asins, "negatives": []})

    # Kelime bazli bid: ayni kampanyadaki her kelime artik ayni bid'i almaz.
    for c in campaigns:
        if not c.get("keywords"):
            continue
        sig = keyword_signals(c["keywords"], search_suggestions, competitors, head_tok)
        c["keyword_signals"] = sig
        c["keyword_bids"] = {
            kw: keyword_bid(kw, c["default_bid"], c["key"], head_tok, sig)
            for kw in c["keywords"]
        }

    total_budget = sum(c["budget"] for c in campaigns)
    return {
        "product": {"title": title, "asin": asin, "sku": sku,
                    "price": price, "brand": product.get("brand")},
        "keyword_source": source,
        "competitor_count": len(competitors),
        "competitor_asins": comp_asins,
        "keywords": {"exact": exact, "phrase": phrase, "broad": broad},
        "negatives": negatives,
        "economics": econ,
        "rationale": rationale,
        "expert_reasoning": expert_reasoning,
        "launch_phases": launch_phases,
        "action_plan": action_plan,
        "competitor_intel": intel_data,
        "keyword_analysis": kw_analysis,
        "market_assessment": market_assess,
        "bids": bids,
        "bid_strategy": bid_strategy,
        "bid_strategy_label": BID_STRATEGIES[bid_strategy]["label"],
        # Gercekten uygulanan strateji (CPC yoksa profit'e duser)
        "effective_strategy": ("profit" if not bench.get("has_cpc") else bid_strategy),
        "effective_strategy_label": BID_STRATEGIES[
            "profit" if not bench.get("has_cpc") else bid_strategy]["label"],
        "bid_explanation": bid_explanation(price, econ, bids, bench),
        "data_warnings": bench.get("warnings") or [],
        "keyword_risks": keyword_risks(
            list(dict.fromkeys((exact or []) + (phrase or []) + (broad or [])))),
        "discovery_plan": (discovery_plan(price, econ, bench,
                                          keywords=(exact or []) + (phrase or []))
                           if not bench.get("has_cpc") else None),
        "data_scope": bench.get("scope"),
        "benchmarks": {"cvr_pct": {k: (round(v*100,2) if v is not None else None)
                                   for k,v in bench["cvr"].items()},
                       "cpc": bench["cpc"], "cvr_source": bench["cvr_source"],
                       "cpc_source": bench["cpc_source"],
                       "cvr_basis": bench.get("cvr_basis"),
                       "ramp_pct": round(bench["ramp"]*100),
                       "account": bench.get("account"),
                       "calibration": bench["calibration_note"]},
        "bid_feasibility": feasibility,
        "bid_outlook": outlook,
        "budgets": budgets,
        "daily_budget_total": round(total_budget, 2),
        "campaigns": campaigns,
        "notes": _launch_notes(sku, asin, source),
    }


def _launch_notes(sku, asin, source):
    notes = []
    if sku == asin:
        notes.append("⚠️ SKU girilmedi — bulk sheet'te ASIN placeholder olarak kullanildi. "
                     "Seller iseniz Product Ad satirlarinda GERCEK SKU'nuzu yazin.")
    if source == "heuristic":
        err = globals().get("_LAST_AI_ERROR") or ""
        if not config.ANTHROPIC_API_KEY:
            notes.append("ℹ️ ANTHROPIC_API_KEY tanimli degil — keyword'ler sezgisel cikarildi.")
        elif "credit balance" in err.lower():
            notes.append("💳 AI CALISMADI: Anthropic hesabinda kredi yok. Plans & Billing'den "
                         "kredi yukleyin; keyword'ler simdilik sezgisel uretildi.")
        elif "model" in err.lower() and "not_found" in err.lower():
            notes.append(f"⚠️ AI CALISMADI: model adi gecersiz ({config.LAUNCH_MODEL}). "
                         ".env'deki LAUNCH_MODEL'i duzeltin.")
        elif err:
            notes.append(f"⚠️ AI CALISMADI: {err[:180]} — keyword'ler sezgisel uretildi.")
        else:
            notes.append("ℹ️ AI kapali — keyword'ler rakip basliklarindan sezgisel cikarildi.")
    notes.append("💡 Launch stratejisi: 2 hafta veri topla, exact'a kazananlari tasi, "
                 "auto/broad'daki alakasiz terimleri negatif yap.")
    return notes


# ------------------------------------------------------------------ bulksheet
_AUTO_EXPR = [
    ("close-match", "loose-match", "substitutes", "complements"),
]


def _blank():
    return {h: "" for h in BULK_HEADERS}


# Amazon'un resmi sablonundaki (Config sayfasi) gecerli degerler.
# Kaynak: AmazonAdvertisingBulksheetSellerTemplate.xlsx
SP_VALID = {
    "Product": {"Sponsored Products"},
    "Entity": {"Campaign", "Ad Group", "Bidding Adjustment",
               "Campaign Negative Keyword", "Keyword", "Negative Keyword",
               "Product Targeting", "Negative Product Targeting", "Product Ad"},
    "Operation": {"Create", "Update", "Archive"},
    "Targeting Type": {"AUTO", "MANUAL"},
    "Bidding Strategy": {"Dynamic bids - down only",
                         "Dynamic bids - up and down", "Fixed bid"},
}
SP_MATCH_TYPES = {
    "Keyword": {"exact", "phrase", "broad"},
    "Negative Keyword": {"negativeExact", "negativePhrase"},
    "Campaign Negative Keyword": {"negativeExact", "negativePhrase"},
}
# Create islemi icin entity bazinda zorunlu alanlar.
SP_REQUIRED = {
    "Campaign": ["Campaign ID", "Campaign Name", "Daily Budget",
                 "Targeting Type", "State", "Start Date", "Bidding Strategy"],
    "Ad Group": ["Campaign ID", "Ad Group ID", "Ad Group Name",
                 "Ad Group Default Bid", "State"],
    "Product Ad": ["Campaign ID", "Ad Group ID", "SKU", "State"],
    "Keyword": ["Campaign ID", "Ad Group ID", "State", "Keyword Text", "Match Type"],
    "Negative Keyword": ["Campaign ID", "Ad Group ID", "State", "Keyword Text",
                         "Match Type"],
    "Product Targeting": ["Campaign ID", "Ad Group ID", "State",
                          "Product Targeting Expression"],
}


def validate_bulk_row(d):
    """Tek bir bulksheet satirini Amazon kurallarina gore denetler.

    Bozuk dosyayi kullaniciya vermektense burada yakalamak ucuzdur; Amazon'un
    hata raporu satir satir okumak zorunda birakiyor.
    """
    errs = []
    for field, allowed in SP_VALID.items():
        v = d.get(field)
        if v not in (None, "") and v not in allowed:
            errs.append(f"{field}={v!r} gecersiz (beklenen: {sorted(allowed)})")

    ent = d.get("Entity")
    mt = d.get("Match Type")
    if ent in SP_MATCH_TYPES and mt not in (None, ""):
        if mt not in SP_MATCH_TYPES[ent]:
            errs.append(f"{ent} icin Match Type={mt!r} gecersiz "
                        f"(beklenen: {sorted(SP_MATCH_TYPES[ent])})")

    if d.get("Operation") == "Create":
        for f in SP_REQUIRED.get(ent, []):
            if d.get(f) in (None, ""):
                errs.append(f"{ent} icin zorunlu alan bos: {f}")
    return errs


def build_discovery_campaign(plan):
    """FAZ 0 bulksheet'i: gercek CPC'yi olcen tek, kucuk, exact kampanya.

    Ana plandan ayridir ve bilerek minimaldir:
      - tek kampanya, tek ad group, sadece EXACT match
      - en alakali 8 kelime (yuksek teklif + genis eslesme = pahali cop trafik)
      - sabit teklif (Fixed bid): dinamik teklif olcumu bozar, biz saf
        acik artirma fiyatini gormek istiyoruz
      - baslangic negatifleri: bariz alakasiz trafigi keser
    """
    d = plan.get("discovery_plan") or {}
    if not d:
        return None

    p = plan.get("product") or {}
    asin = (p.get("asin") or "").strip()
    sku = (p.get("sku") or "").strip() or asin
    # Kesif icin en yuksek niyetli kelimeler: exact listesi, yoksa phrase.
    # Kesifte her tiklama pahali (yuksek teklif). Sadece GERCEK arama sorgusu
    # gibi duran kelimelere odenir: cok kelimeli, kategori kok kelimesini
    # iceren, marka kelimesi olmayan.
    # competitor_intel bir ozet dict'tir, rakip listesi degil - basliktan tureт.
    head = set(_head_tokens(p.get("title") or "", None))
    brand_toks = {w for w in re.sub(r"[^a-z0-9]+", " ",
                                    str(p.get("brand") or "").lower()).split() if w}

    def _good(k):
        ws_ = str(k).lower().split()
        if len(ws_) < 2:
            return False
        if any(w in brand_toks for w in ws_):
            return False
        return (not head) or bool(set(ws_) & head)

    pool = (plan.get("keywords", {}).get("exact") or []) + \
           (plan.get("keywords", {}).get("phrase") or []) + \
           (plan.get("keywords", {}).get("broad") or [])
    kws = [k for k in dict.fromkeys(pool) if k and _good(k)]
    if len(kws) < 4:
        # Cok az kelime kaldiysa olcum guvenilmez olur: kategori kok kelimesi
        # sartini gevset, cok kelimeli + marka disi olmak yeterli.
        relaxed = [k for k in dict.fromkeys(pool)
                   if k and len(str(k).split()) >= 2
                   and not any(w in brand_toks for w in str(k).lower().split())]
        for k in relaxed:
            if k not in kws:
                kws.append(k)
    # Uzun (daha spesifik) olanlar once: niyet net, olcum temiz.
    kws.sort(key=lambda k: (-len(str(k).split()), str(k)))
    kws = kws[:8]
    if not kws:
        return None

    bid = d.get("probe_bid") or 3.00
    budget = d.get("budget_per_day") or 25
    name = f"{_campaign_prefix(p.get('brand'), p.get('title'), asin)} | FAZ0 | CPC Kesif"
    cid = _sanitize_name(name)
    agid = f"{cid} - AG"
    today = datetime.now().strftime("%Y%m%d")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Sponsored Products Campaigns")
    ws.append(BULK_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor="1F2937")
    ws.freeze_panes = "A2"

    problems = []

    def emit(d_):
        problems.extend(validate_bulk_row(d_))
        ws.append([d_.get(h, "") for h in BULK_HEADERS])

    r = _blank(); r.update({
        "Product": "Sponsored Products", "Entity": "Campaign", "Operation": "Create",
        "Campaign ID": cid, "Campaign Name": name, "Start Date": today,
        "Targeting Type": "MANUAL", "State": "enabled", "Daily Budget": budget,
        # Fixed bid: dinamik ayarlama olculen CPC'yi bozar.
        "Bidding Strategy": "Fixed bid"})
    emit(r)

    r = _blank(); r.update({
        "Product": "Sponsored Products", "Entity": "Ad Group", "Operation": "Create",
        "Campaign ID": cid, "Ad Group ID": agid, "Ad Group Name": "CPC Kesif",
        "State": "enabled", "Ad Group Default Bid": bid})
    emit(r)

    r = _blank(); r.update({
        "Product": "Sponsored Products", "Entity": "Product Ad", "Operation": "Create",
        "Campaign ID": cid, "Ad Group ID": agid, "State": "enabled",
        "SKU": sku, "ASIN (Informational only)": asin})
    emit(r)

    for kw in kws:
        r = _blank(); r.update({
            "Product": "Sponsored Products", "Entity": "Keyword", "Operation": "Create",
            "Campaign ID": cid, "Ad Group ID": agid, "State": "enabled",
            "Bid": bid, "Keyword Text": kw, "Match Type": "exact"})
        emit(r)

    for nk in (plan.get("negatives") or [])[:25]:
        r = _blank(); r.update({
            "Product": "Sponsored Products", "Entity": "Negative Keyword",
            "Operation": "Create", "Campaign ID": cid, "Ad Group ID": agid,
            "State": "enabled", "Keyword Text": nk, "Match Type": "negativeExact"})
        emit(r)

    # ---- GUVENLIK AGI: AUTO kampanya ----
    # Yepyeni bir ASIN'de exact keyword'ler gosterim almayabilir: Amazon'un
    # o kelimeler icin alaka gecmisi yoktur. Auto targeting'de yerlesimi
    # Amazon secer, bu yuzden ilk gosterimi almanin EN GUVENILIR yoludur.
    # Kucuk butceyle yanina konur: exact hic veri vermezse bile CPC olculur.
    acid = _sanitize_name(f"{name} | AUTO")
    aagid = f"{acid} - AG"
    auto_budget = max(10, int(round(budget * 0.5)))
    r = _blank(); r.update({
        "Product": "Sponsored Products", "Entity": "Campaign", "Operation": "Create",
        "Campaign ID": acid, "Campaign Name": f"{name} | AUTO", "Start Date": today,
        "Targeting Type": "AUTO", "State": "enabled", "Daily Budget": auto_budget,
        "Bidding Strategy": "Fixed bid"})
    emit(r)
    r = _blank(); r.update({
        "Product": "Sponsored Products", "Entity": "Ad Group", "Operation": "Create",
        "Campaign ID": acid, "Ad Group ID": aagid, "Ad Group Name": "Auto Kesif",
        "State": "enabled", "Ad Group Default Bid": bid})
    emit(r)
    r = _blank(); r.update({
        "Product": "Sponsored Products", "Entity": "Product Ad", "Operation": "Create",
        "Campaign ID": acid, "Ad Group ID": aagid, "State": "enabled",
        "SKU": sku, "ASIN (Informational only)": asin})
    emit(r)
    # Sadece yakin eslesme: gosterim garantisi isterken alakasiz trafige
    # yuksek teklifle para vermeyelim.
    for grp in ("close-match", "loose-match"):
        r = _blank(); r.update({
            "Product": "Sponsored Products", "Entity": "Product Targeting",
            "Operation": "Create", "Campaign ID": acid, "Ad Group ID": aagid,
            "State": "enabled", "Bid": bid,
            "Product Targeting Expression": grp})
        emit(r)
    for nk in (plan.get("negatives") or [])[:25]:
        r = _blank(); r.update({
            "Product": "Sponsored Products", "Entity": "Negative Keyword",
            "Operation": "Create", "Campaign ID": acid, "Ad Group ID": aagid,
            "State": "enabled", "Keyword Text": nk, "Match Type": "negativeExact"})
        emit(r)

    if problems:
        raise ValueError("Kesif bulksheet'i Amazon kurallarina uymuyor: "
                         + " | ".join(list(dict.fromkeys(problems))[:6]))
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_bulksheet(plan):
    """plan -> BytesIO(xlsx). Amazon SP Create operasyonlari."""
    p = plan["product"]
    asin = p.get("asin") or ""
    sku = p.get("sku") or asin
    today = datetime.now().strftime("%Y%m%d")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Sponsored Products Campaigns")
    ws.append(BULK_HEADERS)
    head_font = Font(bold=True, color="FFFFFF", size=10)
    head_fill = PatternFill("solid", fgColor="1F2937")
    for cell in ws[1]:
        cell.font, cell.fill = head_font, head_fill
    ws.freeze_panes = "A2"

    problems = []

    def emit(d):
        problems.extend(validate_bulk_row(d))
        ws.append([d.get(h, "") for h in BULK_HEADERS])

    for c in plan["campaigns"]:
        cid = _sanitize_name(c["name"])          # placeholder ID = kampanya adi
        agid = f"{cid} - AG"
        ag_name = "Ad Group 1"

        # 1) Campaign
        r = _blank()
        r.update({
            "Product": "Sponsored Products", "Entity": "Campaign",
            "Operation": "Create", "Campaign ID": cid, "Campaign Name": c["name"],
            "Start Date": today,
            # Amazon sablonu AUTO/MANUAL (buyuk harf) bekler; "Auto"/"Manual"
            # yazilirsa kampanya satiri reddedilebilir.
            "Targeting Type": str(c["targeting_type"]).upper(),
            "State": "enabled", "Daily Budget": c["budget"],
            "Bidding Strategy": "Dynamic bids - down only",
        })
        emit(r)

        # 2) Ad Group
        r = _blank()
        r.update({
            "Product": "Sponsored Products", "Entity": "Ad Group",
            "Operation": "Create", "Campaign ID": cid, "Ad Group ID": agid,
            "Ad Group Name": ag_name, "State": "enabled",
            "Ad Group Default Bid": c["default_bid"],
        })
        emit(r)

        # 3) Product Ad
        # Kampanya kendi ASIN'ini belirtmisse onu kullan (Firsat Radari her
        # kampanyayi dogru urune baglar); belirtmemisse plandaki tek urune dus.
        c_asin = c.get("asin") or asin
        c_sku = c.get("sku") or (sku if c_asin == asin else c_asin)
        r = _blank()
        r.update({
            "Product": "Sponsored Products", "Entity": "Product Ad",
            "Operation": "Create", "Campaign ID": cid, "Ad Group ID": agid,
            "State": "enabled", "SKU": c_sku, "ASIN (Informational only)": c_asin,
        })
        emit(r)

        # 3b) Negatif exact keyword'ler (bosa harcamayi keser)
        for nk in c.get("negatives", []):
            r = _blank()
            r.update({
                "Product": "Sponsored Products", "Entity": "Negative Keyword",
                "Operation": "Create", "Campaign ID": cid, "Ad Group ID": agid,
                "State": "enabled", "Keyword Text": nk, "Match Type": "negativeExact",
            })
            emit(r)

        # 4a) Auto targeting gruplari
        if c.get("auto_groups"):
            for grp in ("close-match", "loose-match", "substitutes", "complements"):
                r = _blank()
                r.update({
                    "Product": "Sponsored Products", "Entity": "Product Targeting",
                    "Operation": "Create", "Campaign ID": cid, "Ad Group ID": agid,
                    "State": "enabled", "Bid": c["default_bid"],
                    "Product Targeting Expression": grp,
                })
                emit(r)

        # 4b) Keyword'ler
        for kw in c.get("keywords", []):
            r = _blank()
            r.update({
                "Product": "Sponsored Products", "Entity": "Keyword",
                "Operation": "Create", "Campaign ID": cid, "Ad Group ID": agid,
                "State": "enabled",
                "Bid": (c.get("keyword_bids") or {}).get(kw, c["default_bid"]),
                "Keyword Text": kw, "Match Type": c["match"],
            })
            emit(r)

        # 4c) ASIN product targeting
        for a in c.get("product_targets", []):
            r = _blank()
            r.update({
                "Product": "Sponsored Products", "Entity": "Product Targeting",
                "Operation": "Create", "Campaign ID": cid, "Ad Group ID": agid,
                "State": "enabled", "Bid": c["default_bid"],
                "Product Targeting Expression": f'asin="{a}"',
            })
            emit(r)

    if problems:
        # Bozuk dosya uretip Amazon'a yukletmektense burada dur.
        uniq = list(dict.fromkeys(problems))[:10]
        raise ValueError("Bulksheet Amazon kurallarina uymuyor: " + " | ".join(uniq))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
