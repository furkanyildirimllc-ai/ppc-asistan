import openpyxl

wb = openpyxl.load_workbook("/Users/furkanyildirimllc/Claude/BATCI_amazon_bulksheet_20260722.xlsx")
print("Sheets:", wb.sheetnames)

if "Sponsored Products Campaigns" in wb.sheetnames:
    ws = wb["Sponsored Products Campaigns"]
    
    # Read headers
    headers = [cell.value for cell in ws[1]]
    print("Headers match Amazon required structure?:", "Product" in headers and "Entity" in headers and "Operation" in headers)
    
    print("\nFirst 3 data rows:")
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=4, values_only=True)):
        print(f"Row {i+1}:")
        row_dict = dict(zip(headers, row))
        # Print only non-none values
        print({k: v for k, v in row_dict.items() if v is not None})
        print("-" * 40)
        
